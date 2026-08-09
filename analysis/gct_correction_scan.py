# Scans all raw billing extracts (RT40+ premise-level only) and produces
# corrected revenue_jmd (=net_revenue) + gct_jmd (=GCT) per (jps_ac, year, month,
# rate_class), matched to jps_actuals' own key convention, for a targeted UPDATE.
# Does NOT touch RT10/RT20 (bucket/segment-grain in jps_actuals — different problem).
import openpyxl, csv, json, os, time

FILES = {
    '2025-01': 'Jan 25.csv', '2025-02': 'Feb 25.csv', '2025-03': 'Mar 25.csv', '2025-04': 'Apr 25.csv',
    '2025-05': 'Billing Details Report May 2025.xlsx',
    '2025-06': 'Jun 25.csv', '2025-07': 'Jul 25.csv', '2025-08': 'Aug 25.csv', '2025-09': 'Sep 25.csv',
    '2025-10': 'Billing Details Report - October-2025.xlsx',
    '2025-11': 'Billing Details Report - November-2025.xlsx',
    '2025-12': 'Billing Details Report - December-2025.xlsx',
    '2026-01': 'Billing Details Report_Jan 2026.xlsx',
    '2026-02': 'Billing Details Report_Feb 2026.xlsx',
    '2026-03': 'Billing Details Report_Mar 2026.xlsx',
    '2026-04': 'Billing Details Report_Apr 2026.xlsx',
    '2026-05': 'Billing Details Report_May 2026.xlsx',
}

RMAP = {}
RCMAP = {}
for r in list(csv.reader(open(r'C:\Users\jwilson\Downloads\Rate categorry Data mapping.csv')))[1:]:
    if len(r) >= 3:
        RMAP[r[2].strip().upper()] = r[0]
        if r[1].strip().upper() not in RCMAP:
            RCMAP[r[1].strip().upper()] = r[0]

def title_of(rc, srat):
    return RCMAP.get(str(rc).strip().upper()) or RMAP.get(str(srat).strip().upper())

def is_zip_xlsx(path):
    with open(path, 'rb') as f:
        return f.read(2) == b'PK'

def rows_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for r in ws.iter_rows(values_only=True):
        yield r
    wb.close()

def rows_csv(path):
    with open(path, encoding='utf-8', errors='replace', newline='') as f:
        for r in csv.reader(f):
            yield tuple(r)

def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        v = v.strip()
        if v == '':
            return 0.0
        try:
            return float(v)
        except ValueError:
            return 0.0
    return 0.0

t0 = time.time()
for mo, fname in sorted(FILES.items()):
    outpath = 'gct_correction_%s.json' % mo
    if os.path.exists(outpath):
        print(mo, 'already done, skipping', flush=True)
        continue
    out = []
    tf = time.time()
    path = fname
    rows = rows_csv(path) if path.lower().endswith('.csv') else (
        rows_xlsx(path) if is_zip_xlsx(path) else None)
    hdr = None
    idx = None
    n = 0
    for r in rows:
        if r and r[0] == 'Cust_Code':
            hdr = list(r)
            idx = {name: i for i, name in enumerate(hdr) if name}
            continue
        if hdr is None:
            continue
        code = r[idx['Cust_Code']] if idx.get('Cust_Code', -1) < len(r) else None
        if code in (None, '', 'Cust_Code'):
            continue
        cb = r[idx['cust_billed']] if 'cust_billed' in idx and idx['cust_billed'] < len(r) else None
        if cb is not None and str(cb).strip() in ('0', '0.0'):
            continue
        srat = str(r[idx['Srat_Code']]) if idx.get('Srat_Code', -1) < len(r) else ''
        rc = str(r[idx['rate_class']]) if idx.get('rate_class', -1) < len(r) else ''
        title = title_of(rc, srat)
        if title in (None, 'RT10', 'RT20', 'UNMAPPED'):
            continue
        prem = str(r[idx['Prem_Code']] or '').strip() if idx.get('Prem_Code', -1) < len(r) else ''
        if not prem:
            continue
        jps_ac = f"{code}-{prem}"
        nr = num(r[idx['net_revenue']]) if 'net_revenue' in idx else 0.0
        gct = num(r[idx['GCT']]) if 'GCT' in idx else 0.0
        out.append((jps_ac, mo, title, round(nr, 2), round(gct, 2)))
        n += 1
    json.dump(out, open(outpath, 'w'))
    print(mo, fname, 'rows kept:', n, 'in', round(time.time() - tf, 1), 's', flush=True)
    del out

print('ALL MONTHS DONE in', round(time.time() - t0, 1), 's')
