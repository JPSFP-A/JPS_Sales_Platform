import openpyxl, json, time
PATH='Billing Details Report May 2025.xlsx'
NUM=lambda v: float(v) if isinstance(v,(int,float)) else 0.0
wb=openpyxl.load_workbook(PATH, read_only=True)
# find the detail sheet (one whose rows contain a 'Cust_Code' header)
detail=None
for sn in wb.sheetnames:
    ws=wb[sn]
    for i,r in enumerate(ws.iter_rows(values_only=True)):
        if r and r[0]=='Cust_Code': detail=sn; break
        if i>8: break
    if detail: break
print('detail sheet:',detail,flush=True)
ws=wb[detail]; it=ws.iter_rows(values_only=True)
hdr=None
for r in it:
    if r and r[0]=='Cust_Code': hdr=list(r); break
I={n:k for k,n in enumerate(hdr) if n}
g=lambda r,n: NUM(r[I[n]])
energy=lambda r: g(r,'KWHP_KWH_Energy')+g(r,'KWHL_Energy')+g(r,'KWHO_Energy')
demand=lambda r: g(r,'KVAP_KVA_Demand')+g(r,'KVAL_Demand')+g(r,'KVAO_Demand')
fuel  =lambda r: g(r,'fuel')+g(r,'FuelOffPeak')+g(r,'FuelPartialPeak')+g(r,'FuelOnPeak')
byc={}; rt70=[]; cement=[]; n=0; t=time.time()
for r in it:
    if not r or r[I['Cust_Code']] in (None,'','Cust_Code'): continue
    n+=1
    rate=r[I['Srat_Code']] or 'NA'
    nk=g(r,'net_kwh_billed_consump'); nr=g(r,'net_revenue')
    en=energy(r); dm=demand(r); fu=fuel(r); ip=g(r,'IPP_Charge'); cc=g(r,'Cust_Charge')
    radj=g(r,'revenue_adj'); nadj=g(r,'Net_Billing_Adj'); tadj=g(r,'Tariff_Adj')
    kvab=g(r,'kva_billed_consump'); kvaa=g(r,'kva_actual_consump'); cb=g(r,'cust_billed')
    a=byc.setdefault(rate,[0]*13)
    a[0]+=1; a[1]+=nk; a[2]+=nr; a[3]+=en; a[4]+=dm; a[5]+=fu; a[6]+=ip; a[7]+=cc
    a[8]+=radj; a[9]+=nadj; a[10]+=tadj; a[11]+=kvab; a[12]+=cb
    nm=r[I['Name']] or ''
    rec=[nm,round(nk),round(nr),round(en),round(dm),round(fu),round(ip),round(cc),round(kvab),round(kvaa),round(radj+nadj+tadj)]
    if rate=='RT70': rt70.append(rec)
    if 'cement' in nm.lower(): cement.append([rate]+rec)
wb.close()
res=dict(nrows=n,secs=round(time.time()-t),byclass=byc,rt50=[],rt70=rt70,cement=cement)
out=json.load(open('billing_agg.json'))
out['May25']=res
json.dump(out, open('billing_agg.json','w'))
print('done May25 rows',n,'in',res['secs'],'s — merged',flush=True)
