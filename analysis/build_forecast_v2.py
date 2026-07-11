import csv, re, json, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
ITALIC_RED = Font(name=FONT, italic=True, color="C00000")
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBHDR_FILL = PatternFill("solid", fgColor="D9E2F3")
YELLOW = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="B7C6D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUMFMT = "#,##0;(#,##0);\"-\""
PCTFMT = "0.0%"
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

SRC_COMM = r"D:\Projects\Sales_Platform\analysis\commercial_actuals_2024_2026.csv"
SRC_RES = r"D:\Projects\Sales_Platform\analysis\residential_actuals_2024_2026.csv"
OUT = r"D:\Projects\Sales_Platform\analysis\Commercial_RateClass_Forecast_2026.xlsx"

COMMERCIAL_RCS = [("RT40", "LV Commercial"), ("RT50", "MV Commercial"),
                  ("RT60-ST", "Street Lights"), ("RT70", "Large Industrial")]
RESIDENTIAL_RCS = [("RT10", "Residential"), ("RT20", "SME / Gen. Service")]
ALL_RCS = RESIDENTIAL_RCS + COMMERCIAL_RCS


def is_individual(ac):
    return bool(re.match(r"^[0-9]", ac))


def load_commercial():
    rows = list(csv.DictReader(open(SRC_COMM, encoding="utf-8")))
    indiv, bucket = {}, {}
    for r in rows:
        ac, rc, parish = r["jps_ac"], r["rate_class"], r["parish"]
        if rc not in dict(COMMERCIAL_RCS):
            continue
        y, m = int(r["year"]), int(r["month"])
        kwh = float(r["kwh"] or 0)
        if is_individual(ac):
            key = (ac, rc)
            d = indiv
        else:
            key = (ac, rc, parish)
            d = bucket
        if key not in d:
            d[key] = {"name": r["name"], "parish": parish, "kwh": {}}
        d[key]["kwh"][(y, m)] = kwh
        if r["name"]:
            d[key]["name"] = r["name"]
    return indiv, bucket


def load_residential():
    if not os.path.exists(SRC_RES):
        return {}
    rows = list(csv.DictReader(open(SRC_RES, encoding="utf-8")))
    bucket = {}
    for r in rows:
        rc = r["rate_class"]
        if rc not in dict(RESIDENTIAL_RCS):
            continue
        ac, parish, cb = r["jps_ac"], r["parish"], r.get("consumption_bucket", "")
        y, m = int(r["year"]), int(r["month"])
        kwh = float(r["kwh"] or 0)
        key = (ac, rc, parish, cb)
        if key not in bucket:
            bucket[key] = {"name": r["name"], "parish": parish, "bucket": cb, "kwh": {}}
        bucket[key]["kwh"][(y, m)] = kwh
    return bucket


indiv_all, bucket_all = load_commercial()
res_bucket_all = load_residential()

wb = Workbook()
wb.remove(wb.active)

SHEET_LAYOUT = {}  # rc -> {'sheet':name, 'drv_row':4, 'first_data_row':.., 'last_data_row':.., 'total_row_formula_range':(sheet,colstart..)}


def style_header(ws, row, headers, col_start=1):
    for i, h in enumerate(headers, col_start):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 28


def add_drivers_block(ws, rc, first_data_row_hint_cell):
    ws["A1"] = f"{rc}"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws["B1"] = ""
    ws["A2"] = "Drivers (edit blue cells)"
    ws["A2"].font = Font(name=FONT, bold=True, italic=True)
    headers = ["Trailing Growth % (Apr-Jun YoY, calc)", "Growth Adjustment % (toggle)",
                "Customer Count Growth % (toggle)", "Other Monthly Adjustment (kWh, toggle)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = HDR_FONT
        c.fill = SUBHDR_FILL
        c.font = Font(name=FONT, bold=True, color="1F4E78")
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BORDER
    ws.row_dimensions[3].height = 34
    # trailing growth placeholder formula filled in later once we know data row range
    d = ws.cell(row=4, column=2, value=0.0); d.font = BLUE; d.number_format = PCTFMT; d.fill = YELLOW; d.border = BORDER
    e = ws.cell(row=4, column=3, value=0.0); e.font = BLUE; e.number_format = PCTFMT; e.fill = YELLOW; e.border = BORDER
    g = ws.cell(row=4, column=4, value=0); g.font = BLUE; g.number_format = NUMFMT; g.fill = YELLOW; g.border = BORDER
    for col in (1, 2, 3, 4):
        ws.cell(row=4, column=col).border = BORDER
    for i in range(1, 5):
        ws.column_dimensions[get_column_letter(i)].width = 24
    return {"trailing": "A4", "growth_adj": "B4", "cust_growth": "C4", "other_adj": "D4"}


def write_customer_table(ws, start_row, keys_data, id_label, use_bucket_cols=False):
    """keys_data: list of (row_key, data_dict) with data['kwh'] map, data['name'], data['parish']
    Returns dict of column indices used, and last row written."""
    header_row = start_row
    cols = ["Account #" if not use_bucket_cols else "Bucket Label", "Name", "Parish"]
    if use_bucket_cols:
        cols.append("Consumption Bucket")
    id_cols = len(cols)
    cols += [f"2026 {m}" for m in MONTHS[:6]]
    cols += [f"2025 {m}" for m in MONTHS[3:6]]
    cols += [f"2025 {m}" for m in MONTHS[6:12]]
    cols += [f"2024 {m}" for m in MONTHS[6:12]]
    cols += [f"2026 {m} (proj)" for m in MONTHS[6:12]]
    cols += ["FY2026 Total"]
    style_header(ws, header_row, cols)

    COL_JANJUN = list(range(id_cols + 1, id_cols + 7))
    COL_AMJ25 = list(range(id_cols + 7, id_cols + 10))
    COL_JD25 = list(range(id_cols + 10, id_cols + 16))
    COL_JD24 = list(range(id_cols + 16, id_cols + 22))
    COL_PROJ = list(range(id_cols + 22, id_cols + 28))
    COL_TOTAL = id_cols + 28

    r = header_row + 1
    for key, d in keys_data:
        ws.cell(row=r, column=1, value=key[0]).font = BLACK
        ws.cell(row=r, column=2, value=d.get("name", "")).font = BLACK
        ws.cell(row=r, column=3, value=d.get("parish", "")).font = BLACK
        if use_bucket_cols:
            ws.cell(row=r, column=4, value=d.get("bucket", "")).font = BLACK
        for j, m in enumerate(range(1, 7)):
            v = d["kwh"].get((2026, m), 0)
            c = ws.cell(row=r, column=COL_JANJUN[j], value=round(v)); c.font = BLACK; c.number_format = NUMFMT; c.border = BORDER
        for j, m in enumerate(range(4, 7)):
            v = d["kwh"].get((2025, m), 0)
            c = ws.cell(row=r, column=COL_AMJ25[j], value=round(v)); c.font = BLACK; c.number_format = NUMFMT
        for j, m in enumerate(range(7, 13)):
            v = d["kwh"].get((2025, m), 0)
            c = ws.cell(row=r, column=COL_JD25[j], value=round(v)); c.font = BLACK; c.number_format = NUMFMT
        for j, m in enumerate(range(7, 13)):
            v = d["kwh"].get((2024, m), 0)
            c = ws.cell(row=r, column=COL_JD24[j], value=round(v)); c.font = BLACK; c.number_format = NUMFMT
        r += 1
    last_row = r - 1
    return {"header_row": header_row, "first_row": header_row + 1, "last_row": last_row,
            "COL_JANJUN": COL_JANJUN, "COL_AMJ25": COL_AMJ25, "COL_JD25": COL_JD25,
            "COL_JD24": COL_JD24, "COL_PROJ": COL_PROJ, "COL_TOTAL": COL_TOTAL, "id_cols": id_cols}


def fill_projection_formulas(ws, tbl, drv_cells):
    if tbl["first_row"] > tbl["last_row"]:
        return
    jun_col = get_column_letter(tbl["COL_JANJUN"][5])
    id_col_letter = "A"
    for r in range(tbl["first_row"], tbl["last_row"] + 1):
        for j in range(6):
            h25 = get_column_letter(tbl["COL_JD25"][j])
            h24 = get_column_letter(tbl["COL_JD24"][j])
            proj_col = tbl["COL_PROJ"][j]
            f = (f"=IF({h24}{r}=0,{h25}{r},AVERAGE({h25}{r},{h24}{r}))"
                 f"*(1+$B$4+$A$4)"
                 f"+$D$4*IFERROR({jun_col}{r}/SUM(${jun_col}${tbl['first_row']}:${jun_col}${tbl['last_row']}),0)")
            c = ws.cell(row=r, column=proj_col, value=f); c.font = BLACK; c.number_format = NUMFMT; c.border = BORDER
        fa = get_column_letter(tbl["COL_JANJUN"][0])
        la = get_column_letter(tbl["COL_JANJUN"][-1])
        fp = get_column_letter(tbl["COL_PROJ"][0])
        lp = get_column_letter(tbl["COL_PROJ"][-1])
        tot = ws.cell(row=r, column=tbl["COL_TOTAL"], value=f"=SUM({fa}{r}:{la}{r})+SUM({fp}{r}:{lp}{r})")
        tot.font = BOLD; tot.number_format = NUMFMT; tot.border = BORDER


def fill_trailing_growth(ws, drv_cells, table_ranges):
    """table_ranges: list of dicts (from write_customer_table) to SUM across for growth calc"""
    num_parts, den_parts = [], []
    for tbl in table_ranges:
        if tbl["first_row"] > tbl["last_row"]:
            continue
        for j in range(3):
            col = get_column_letter(tbl["COL_JANJUN"][3 + j])
            num_parts.append(f"SUM({col}{tbl['first_row']}:{col}{tbl['last_row']})")
        for j in range(3):
            col = get_column_letter(tbl["COL_AMJ25"][j])
            den_parts.append(f"SUM({col}{tbl['first_row']}:{col}{tbl['last_row']})")
    if not num_parts:
        ws["A4"] = 0
        return
    f = f"=IFERROR(({'+'.join(num_parts)})/({'+'.join(den_parts)})-1,0)"
    ws["A4"] = f
    ws["A4"].number_format = PCTFMT
    ws["A4"].font = BLACK


sheet_meta = {}  # rc -> dict with sheet, tables list (for Total tab references)

for rc, desc in ALL_RCS:
    title = rc.replace("-", "_")[:31]
    ws = wb.create_sheet(title)
    drv_cells = add_drivers_block(ws, f"{rc} - {desc}", None)

    tables = []
    row_cursor = 6
    if rc in dict(COMMERCIAL_RCS):
        indiv_keys = sorted(k for k in indiv_all if k[1] == rc)
        indiv_data = [(k, indiv_all[k]) for k in indiv_keys]
        ws.cell(row=row_cursor - 1, column=1, value=f"Individual Commercial Customers ({len(indiv_data)} accounts)").font = BOLD
        tbl1 = write_customer_table(ws, row_cursor, indiv_data, "Account #")
        tables.append(tbl1)
        row_cursor = tbl1["last_row"] + 3

        bkeys = sorted(k for k in bucket_all if k[1] == rc)
        bdata = [(k, bucket_all[k]) for k in bkeys]
        if bdata:
            note = ws.cell(row=row_cursor - 1, column=1,
                            value=f"Aggregate / Bucket Rows - not individually-identified customers ({len(bdata)} rows; e.g. Postpaid/Prepaid/Zero by parish). Flagged in Read Me.")
            note.font = ITALIC_RED
            tbl2 = write_customer_table(ws, row_cursor, bdata, "Bucket Label")
            tables.append(tbl2)
            row_cursor = tbl2["last_row"] + 3
    else:
        bkeys = sorted(k for k in res_bucket_all if k[1] == rc)
        bdata = [(k, res_bucket_all[k]) for k in bkeys]
        ws.cell(row=row_cursor - 1, column=1,
                value=f"Parish x Consumption-Bucket rows ({len(bdata)}) - residential classes have no individual account detail in jps_actuals").font = BOLD
        tbl1 = write_customer_table(ws, row_cursor, bdata, "Bucket Label", use_bucket_cols=True)
        tables.append(tbl1)
        row_cursor = tbl1["last_row"] + 3

    for tbl in tables:
        fill_projection_formulas(ws, tbl, drv_cells)
    fill_trailing_growth(ws, drv_cells, tables)

    for i in range(1, 32):
        w = 12
        if i <= 4:
            w = [14, 26, 14, 16][i - 1] if i <= 4 else 12
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A6"

    sheet_meta[rc] = {"title": title, "tables": tables}

print("sheet build done:", list(sheet_meta.keys()))

# ── TOTAL SHEET ──────────────────────────────────────────────────────
tot = wb.create_sheet("Total", 0)
tot.column_dimensions["A"].width = 26
for i in range(2, 15):
    tot.column_dimensions[get_column_letter(i)].width = 12
style_header(tot, 1, ["Rate Class"] + [f"2026 {m}" for m in MONTHS])

row = 2
rc_total_rows = {}
for rc, desc in ALL_RCS:
    tot.cell(row=row, column=1, value=f"{rc} - {desc}").font = BOLD
    meta = sheet_meta[rc]
    sheet_title = meta["title"]
    for j in range(6):
        parts = []
        for tbl in meta["tables"]:
            if tbl["first_row"] > tbl["last_row"]:
                continue
            col = get_column_letter(tbl["COL_JANJUN"][j])
            parts.append(f"SUM('{sheet_title}'!{col}{tbl['first_row']}:{col}{tbl['last_row']})")
        f = "=" + "+".join(parts) if parts else "=0"
        c = tot.cell(row=row, column=2 + j, value=f); c.font = GREEN; c.number_format = NUMFMT; c.border = BORDER
    for j in range(6):
        parts = []
        for tbl in meta["tables"]:
            if tbl["first_row"] > tbl["last_row"]:
                continue
            col = get_column_letter(tbl["COL_PROJ"][j])
            parts.append(f"SUM('{sheet_title}'!{col}{tbl['first_row']}:{col}{tbl['last_row']})")
        f = "=" + "+".join(parts) if parts else "=0"
        c = tot.cell(row=row, column=8 + j, value=f); c.font = GREEN; c.number_format = NUMFMT; c.border = BORDER
    rc_total_rows[rc] = row
    row += 1

cust_growth_row = row + 1
tot.cell(row=cust_growth_row, column=1, value="Customer Count Growth Add-On (Jul-Dec, kWh)").font = Font(name=FONT, italic=True)
for j in range(6):
    parts = []
    for rc, desc in ALL_RCS:
        meta = sheet_meta[rc]
        sheet_title = meta["title"]
        for tbl in meta["tables"]:
            if tbl["first_row"] > tbl["last_row"]:
                continue
            jun_col = get_column_letter(tbl["COL_JANJUN"][5])
            n = tbl["last_row"] - tbl["first_row"] + 1
            parts.append(
                f"'{sheet_title}'!$C$4*{n}*IFERROR(SUM('{sheet_title}'!{jun_col}{tbl['first_row']}:{jun_col}{tbl['last_row']})/{n},0)"
            )
    f = "=" + "+".join(parts) if parts else "=0"
    c = tot.cell(row=cust_growth_row, column=8 + j, value=f); c.font = BLACK; c.number_format = NUMFMT

total_row = cust_growth_row + 1
tot.cell(row=total_row, column=1, value="TOTAL (incl. Count Growth)").font = BOLD
for j in range(6):
    col = get_column_letter(2 + j)
    c = tot.cell(row=total_row, column=2 + j, value=f"=SUM({col}2:{col}{cust_growth_row-2})")
    c.font = BOLD; c.number_format = NUMFMT
for j in range(6):
    col = get_column_letter(8 + j)
    c = tot.cell(row=total_row, column=8 + j, value=f"=SUM({col}2:{col}{cust_growth_row})")
    c.font = BOLD; c.number_format = NUMFMT

# ── BY PARISH ────────────────────────────────────────────────────────
parish_start = total_row + 3
tot.cell(row=parish_start, column=1, value="BY PARISH (all rate classes, Jan-Jun actual + Jul-Dec projected)").font = Font(name=FONT, bold=True, size=12)
hdr_row = parish_start + 1
style_header(tot, hdr_row, ["Parish"] + [f"2026 {m}" for m in MONTHS])

all_parishes = sorted(set(d["parish"] for k, d in indiv_all.items() if d.get("parish")) |
                       set(d["parish"] for k, d in bucket_all.items() if d.get("parish")) |
                       set(d["parish"] for k, d in res_bucket_all.items() if d.get("parish")))

pr = hdr_row + 1
for parish in all_parishes:
    tot.cell(row=pr, column=1, value=parish).font = BLACK
    for j in range(6):
        parts = []
        for rc, desc in ALL_RCS:
            meta = sheet_meta[rc]
            sheet_title = meta["title"]
            for tbl in meta["tables"]:
                if tbl["first_row"] > tbl["last_row"]:
                    continue
                col = get_column_letter(tbl["COL_JANJUN"][j])
                parish_col = "C"
                parts.append(f"SUMIF('{sheet_title}'!{parish_col}{tbl['first_row']}:{parish_col}{tbl['last_row']},\"{parish}\",'{sheet_title}'!{col}{tbl['first_row']}:{col}{tbl['last_row']})")
        f = "=" + "+".join(parts) if parts else "=0"
        c = tot.cell(row=pr, column=2 + j, value=f); c.font = GREEN; c.number_format = NUMFMT; c.border = BORDER
    for j in range(6):
        parts = []
        for rc, desc in ALL_RCS:
            meta = sheet_meta[rc]
            sheet_title = meta["title"]
            for tbl in meta["tables"]:
                if tbl["first_row"] > tbl["last_row"]:
                    continue
                col = get_column_letter(tbl["COL_PROJ"][j])
                parish_col = "C"
                parts.append(f"SUMIF('{sheet_title}'!{parish_col}{tbl['first_row']}:{parish_col}{tbl['last_row']},\"{parish}\",'{sheet_title}'!{col}{tbl['first_row']}:{col}{tbl['last_row']})")
        f = "=" + "+".join(parts) if parts else "=0"
        c = tot.cell(row=pr, column=8 + j, value=f); c.font = GREEN; c.number_format = NUMFMT; c.border = BORDER
    pr += 1

parish_total_row = pr
tot.cell(row=parish_total_row, column=1, value="TOTAL").font = BOLD
for j in range(6):
    col = get_column_letter(2 + j)
    c = tot.cell(row=parish_total_row, column=2 + j, value=f"=SUM({col}{hdr_row+1}:{col}{pr-1})")
    c.font = BOLD; c.number_format = NUMFMT
for j in range(6):
    col = get_column_letter(8 + j)
    c = tot.cell(row=parish_total_row, column=8 + j, value=f"=SUM({col}{hdr_row+1}:{col}{pr-1})")
    c.font = BOLD; c.number_format = NUMFMT

# ── READ ME ──────────────────────────────────────────────────────────
rm = wb.create_sheet("Read Me", 0)
rm.column_dimensions["A"].width = 110
lines = [
    ("Rate Class Forecast - 2026", BOLD),
    ("", None),
    ("Source: jps_actuals table, Supabase project bhrswnbenkvflpdjhfpa. Pulled 2026-07-06.", BLACK),
    ("Scope: all rate classes - RT10 (Residential), RT20 (SME/Gen. Service), RT40 (LV Commercial), RT50 (MV Commercial), RT60-ST (Street Lights), RT70 (Large Industrial).", BLACK),
    ("Jan-Jun 2026: real actuals. Jul-Dec 2026: formula-driven projection (see Methodology).", BLACK),
    ("Each rate class has its own tab with a local Drivers block (blue cells) at the top - edit those to update that class's whole forecast.", BLACK),
    ("", None),
    ("DATA QUALITY FLAGS - NOT FIXED, DISCLOSED HERE:", Font(name=FONT, bold=True, color="C00000")),
    ("1) jps_actuals 2026-06 contains 23,780 rows tagged rate_class='RT20', segment='Commercial' that are really residential", ITALIC_RED),
    ("   Prepaid/Zero bucket data (real parish customer counts, null revenue, blank bucket) mistagged as individual RT20 accounts.", ITALIC_RED),
    ("   Inserted 2026-07-03 14:45-14:48 UTC, no matching sync_runs entry. This entire batch is EXCLUDED from the RT20 tab.", ITALIC_RED),
    ("2) Within RT40 and RT50, real 'Postpaid'/'Zero'/'Zero-Postpaid' bucket rows (by parish) appear alongside individually-", ITALIC_RED),
    ("   identified commercial customers in every month except June 2026, where they are absent entirely. These are NOT", ITALIC_RED),
    ("   individual customers - they are shown in a separate 'Aggregate / Bucket Rows' section on the RT40 and RT50 tabs,", ITALIC_RED),
    ("   not mixed into the customer list. June's missing Postpaid volume for RT40/RT50 has not been investigated further.", ITALIC_RED),
    ("", None),
    ("Methodology (Jul-Dec 2026 projection, per row):", BOLD),
    ("  seasonal base = average of that row's real kWh in the same calendar month, 2024 and 2025 (2025 only if 2024 has no data)", BLACK),
    ("  projected kWh = seasonal base x (1 + trailing growth% + Growth Adjustment toggle) + (Other Monthly Adjustment toggle x row's share of Jun-2026 kWh within its table)", BLACK),
    ("  trailing growth% = (rate class Apr-Jun 2026 actual) / (rate class Apr-Jun 2025 actual) - 1, calculated live on each tab from that tab's own rows", BLACK),
    ("  Customer Count Growth toggle applies only on the Total tab, as a class-level add-on (new accounts), not spread across existing rows", BLACK),
    ("", None),
    ("Tabs:", BOLD),
    ("  Total - all rate classes, Jan-Jun actual + Jul-Dec projected, Customer Count Growth add-on, and a By-Parish breakdown.", BLACK),
    ("  RT10 / RT20 / RT40 / RT50 / RT60_ST / RT70 - one tab per rate class, Drivers block + detail rows + projection.", BLACK),
]
r = 1
for text, font in lines:
    c = rm.cell(row=r, column=1, value=text)
    if font:
        c.font = font
    r += 1

wb.move_sheet("Read Me", offset=-100)
wb.move_sheet("Total", offset=-99)
wb.active = 0
wb.save(OUT)
print("saved", OUT)
