# Streetlight (L-code / RT6x) account-level scan for Apr & May 2026, incl. billing-adjustment columns
# to test the true-up hypothesis behind RT60/L58.
import openpyxl, json, time
FILES={'Apr':'Billing Details Report_Apr 2026.xlsx','May':'Billing Details Report_May 2026.xlsx'}
NUM=lambda v: float(v) if isinstance(v,(int,float)) else 0.0
def isSL(c):
    c=str(c).upper(); return c[:1]=='L' or c[:3]=='RT6'
def proc(path):
    wb=openpyxl.load_workbook(path, read_only=True)
    detail=[s for s in wb.sheetnames if s.upper().startswith('BILLING_DETAILS_REPORT')][0]
    ws=wb[detail]; it=ws.iter_rows(values_only=True)
    hdr=None
    for r in it:
        if r and r[0]=='Cust_Code': hdr=list(r); break
    I={n:k for k,n in enumerate(hdr) if n}
    g=lambda r,n: NUM(r[I[n]]) if n in I else 0.0
    en=lambda r: g(r,'KWHP_KWH_Energy')+g(r,'KWHL_Energy')+g(r,'KWHO_Energy')
    fu=lambda r: g(r,'fuel')+g(r,'FuelOffPeak')+g(r,'FuelPartialPeak')+g(r,'FuelOnPeak')
    acc={}
    for r in it:
        code=r[I['Cust_Code']]
        if code in (None,'','Cust_Code'): continue
        if not isSL(r[I['Srat_Code']]): continue
        a=acc.get(code)
        if a is None: a=acc[code]={'name':(r[I['Name']] or '').strip(),'srat':r[I['Srat_Code']],'m':[0.0]*9}
        m=a['m']
        m[0]+=g(r,'net_kwh_billed_consump'); m[1]+=g(r,'kwh_billed_consump_adj')
        m[2]+=g(r,'net_revenue'); m[3]+=g(r,'revenue_chg'); m[4]+=g(r,'revenue_adj')
        m[5]+=g(r,'Net_Billing_Adj'); m[6]+=en(r); m[7]+=fu(r); m[8]+=g(r,'IPP_Charge')
    wb.close()
    return acc
out={'_legend':['net_kwh','kwh_adj','net_rev','rev_chg','rev_adj','net_bill_adj','energy','fuel','ipp']}
for mo,f in FILES.items():
    t=time.time(); print('processing',mo,flush=True)
    out[mo]={c:{'name':a['name'],'srat':a['srat'],'m':[round(x) for x in a['m']]} for c,a in proc(f).items()}
    json.dump(out,open('streetlight.json','w'))
    print('  done',mo,'accts',len(out[mo]),'in',round(time.time()-t),'s',flush=True)
print('WROTE streetlight.json',flush=True)
