# -*- coding: utf-8 -*-
# Scans all Check Consumption Detail/Report files for BILLED_CONSUMP_ADJ != 0 rows
# on the 40 "Special Accounts", cross-referenced with who/dept from the Special
# Accounts Directory notes (special_accounts_who.json, built by a prior step).
import csv, json, glob, time, zipfile
import openpyxl

who_map = json.load(open("special_accounts_who.json"))
special = set(who_map.keys())

MON = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}

def parse_month_year(v):
    if hasattr(v, 'year'):
        return v.year, v.month
    s = str(v).strip().upper()
    if '-' in s:
        mo, yr = s.split('-')
        return int(yr), MON[mo[:3]]
    return None, None

def is_real_xlsx(fp):
    try:
        with zipfile.ZipFile(fp):
            return True
    except Exception:
        return False

results = []
files = sorted(glob.glob("Check Consumption*.xls*"))
print("files:", len(files))
t0 = time.time()
for fp in files:
    tf = time.time()
    n = 0
    kept = 0
    if is_real_xlsx(fp):
        wb = openpyxl.load_workbook(fp, read_only=True)
        sn = [s for s in wb.sheetnames if "CONSUMP" in s.upper()]
        sn = sn[0] if sn else wb.sheetnames[0]
        rows_iter = wb[sn].iter_rows(values_only=True)
    else:
        f = open(fp, encoding="iso-8859-1")
        rows_iter = csv.reader(f, delimiter="\t")

    hdr = None
    idx = None
    for r in rows_iter:
        if r and r[0] == "CUST_CODE":
            hdr = list(r)
            idx = {name: i for i, name in enumerate(hdr)}
            continue
        if hdr is None:
            continue
        if len(r) <= idx.get("MONTH_YEAR", 11):
            continue
        n += 1
        ac = f"{r[idx['CUST_CODE']]}-{r[idx['PREM_CODE']]}"
        if ac not in special:
            continue
        try:
            adj = float(r[idx["BILLED_CONSUMP_ADJ"]])
        except (TypeError, ValueError):
            continue
        if adj == 0:
            continue
        year, month = parse_month_year(r[idx["MONTH_YEAR"]])
        if not year:
            continue
        results.append({
            "ac": ac, "year": year, "month": month,
            "rate_class_raw": r[idx["SERV_RATE"]], "scat": r[idx["SCAT_CODE"]],
            "adj": adj,
            "who": who_map[ac]["who"], "dept": who_map[ac]["dept"], "name": who_map[ac]["name"],
        })
        kept += 1
    if not is_real_xlsx(fp):
        f.close()
    else:
        wb.close()
    print(f"{fp}: {n} rows scanned, {kept} special-account adj rows kept, {time.time()-tf:.1f}s")

print("TOTAL elapsed", round(time.time()-t0, 1), "s")
print("total adjustment rows:", len(results))
json.dump(results, open("special_adjustments.json", "w"))
print("wrote special_adjustments.json")
