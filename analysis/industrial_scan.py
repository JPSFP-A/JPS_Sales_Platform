# Account-level scan of RT40/50/70 families for May-25 & May-26. Captures per-customer
# kwh/revenue/energy/demand/fuel + flags cement/windalco/alumina. Constant memory.
import openpyxl, json, time
FILES={'May25':'Billing Details Report May 2025.xlsx','May26':'Billing Details Report_10June2026.xlsx'}
NUM=lambda v: float(v) if isinstance(v,(int,float)) else 0.0
def grp(c):
    c=str(c).upper()
    if c[:3] in ('RT4','NB4'): return 'RT40'
    if c[:3]=='RT5': return 'RT50'
    if c[:3]=='RT7': return 'RT70'
    return None
def proc(path):
    wb=openpyxl.load_workbook(path, read_only=True)
    detail=None
    for sn in wb.sheetnames:
        ws=wb[sn]
        for i,r in enumerate(ws.iter_rows(values_only=True)):
            if r and r[0]=='Cust_Code': detail=sn; break
            if i>8: break
        if detail: break
    ws=wb[detail]; it=ws.iter_rows(values_only=True)
    hdr=None
    for r in it:
        if r and r[0]=='Cust_Code': hdr=list(r); break
    I={n:k for k,n in enumerate(hdr) if n}
    g=lambda r,n: NUM(r[I[n]])
    en=lambda r: g(r,'KWHP_KWH_Energy')+g(r,'KWHL_Energy')+g(r,'KWHO_Energy')
    dm=lambda r: g(r,'KVAP_KVA_Demand')+g(r,'KVAL_Demand')+g(r,'KVAO_Demand')
    fu=lambda r: g(r,'fuel')+g(r,'FuelOffPeak')+g(r,'FuelPartialPeak')+g(r,'FuelOnPeak')
    ind={'RT40':{},'RT50':{},'RT70':{}}; names=set(); n=0; t=time.time()
    for r in it:
        if not r or r[I['Cust_Code']] in (None,'','Cust_Code'): continue
        n+=1
        gr=grp(r[I['Srat_Code']])
        nm=(r[I['Name']] or '').strip()
        low=nm.lower()
        if any(k in low for k in ('windalco','alumina','jamalco','bauxite','alpart','cement')):
            names.add(nm)
        if not gr: continue
        a=ind[gr].setdefault(nm,[0.0]*7)  # kwh,rev,energy,demand,fuel,ipp,kva
        a[0]+=g(r,'net_kwh_billed_consump'); a[1]+=g(r,'net_revenue'); a[2]+=en(r)
        a[3]+=dm(r); a[4]+=fu(r); a[5]+=g(r,'IPP_Charge'); a[6]+=g(r,'kva_billed_consump')
    wb.close()
    return dict(ind={k:{nm:[round(x) for x in v] for nm,v in d.items()} for k,d in ind.items()},
                names=sorted(names), nrows=n, secs=round(time.time()-t))
out={'_legend':['kwh','rev','energy','demand','fuel','ipp','kva']}
for mo,f in FILES.items():
    print('processing',mo,flush=True); out[mo]=proc(f)
    json.dump(out, open('industrial.json','w'))
    print('  done',mo,'rows',out[mo]['nrows'],'in',out[mo]['secs'],'s; names hit:',out[mo]['names'][:12],flush=True)
print('WROTE industrial.json',flush=True)
