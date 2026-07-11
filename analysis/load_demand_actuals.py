# -*- coding: utf-8 -*-
# Parses monthly "Check Consumption Detail/Report" Excel files -> jps_demand_actuals
# rows: {year, month, rate_class, sub_class, demand_type, demand_kva}.
#
# SCAT_CODE -> demand_type mapping (confirmed against jps_tariff_rates' 4 demand-rate
# columns + corrected.json's srat_legend, which carries kva/kvap/kval/kvao as 4
# DISTINCT fields, not kvap+kval+kvao summing into a separate "kva"):
#   KVA  -> standard      (non-TOU demand billing; demand_rate_jmd_kva == demand_rate_standard_jmd)
#   KVAP -> on_peak
#   KVAO -> off_peak
#   KVAL -> partial_peak
#
# rate_class comes from SERV_RATE via the authoritative "Rate categorry Data mapping.csv"
# (Srat_Code -> Rate_Class), the same mapping already used in gen_demand_energy_xlsx.py
# earlier this session. sub_class = STYP_CODE (e.g. MT40) straight through.
import openpyxl, csv, json, time, glob, zipfile

DEMAND_SCAT = {"KVA":"standard","KVAP":"on_peak","KVAO":"off_peak","KVAL":"partial_peak"}
MON = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}

def norm(s): return str(s).strip().upper()

def parse_month_year(v):
    if hasattr(v, 'year'):
        return v.year, v.month
    s = str(v).strip().upper()
    if '-' in s:
        mo, yr = s.split('-')
        return int(yr), MON[mo[:3]]
    return None, None

def is_real_xlsx(fp):
    try:
        with zipfile.ZipFile(fp):
            return True
    except Exception:
        return False

# Title col (row[0]) holds the canonical rate_class in this CSV; Srat_Code (row[2]) is the key.
RCMAP = {}
with open(r"C:\Users\jwilson\Downloads\Rate categorry Data mapping.csv", encoding="utf-8-sig") as f:
    for row in list(csv.reader(f))[1:]:
        if len(row) >= 3:
            RCMAP[norm(row[2])] = row[0]
RCMAP['RT76'] = 'RT70'  # Jamalco-only wholesale tariff variant, same family as RT74/RT75 (per Special Accounts Directory notes: "RT 76 consists of Jamalco only")

# Some monthly extracts arrive with a .xls extension but are actually tab-delimited
# text (not real OOXML/BIFF) — is_real_xlsx() below sniffs the real format per file
# rather than trusting the extension.
FILES = sorted(set(glob.glob("Check Consumption*.xlsx")) | set(glob.glob("Check Consumption*.xls")))
print("files:", len(FILES))

agg = {}          # (year,month,rate_class,sub_class,demand_type) -> kva sum
unmapped_srat = {}  # SERV_RATE -> count, for anything not in RCMAP
row_counts = {}

t0 = time.time()
for fp in FILES:
    tf = time.time()
    n = 0
    kept = 0
    if is_real_xlsx(fp):
        wb = openpyxl.load_workbook(fp, read_only=True)
        sn = [s for s in wb.sheetnames if "CONSUMP" in s.upper()]
        sn = sn[0] if sn else wb.sheetnames[0]
        rows_iter = wb[sn].iter_rows(values_only=True)
    else:
        fh = open(fp, encoding="iso-8859-1")
        rows_iter = csv.reader(fh, delimiter="\t")
    hdr = None
    for r in rows_iter:
        if r and r[0] == "CUST_CODE":
            hdr = list(r)
            idx = {name: i for i, name in enumerate(hdr)}
            continue
        if hdr is None:
            continue
        if len(r) <= idx.get("MONTH_YEAR", 11):
            continue
        n += 1
        scat = r[idx["SCAT_CODE"]]
        if scat not in DEMAND_SCAT:
            continue
        serv_rate = norm(r[idx["SERV_RATE"]])
        rc = RCMAP.get(serv_rate)
        if not rc:
            unmapped_srat[serv_rate] = unmapped_srat.get(serv_rate, 0) + 1
            continue
        if rc == "EV":
            continue
        sub_class = r[idx["STYP_CODE"]]
        year, month = parse_month_year(r[idx["MONTH_YEAR"]])
        if not year:
            continue
        val = r[idx["NET_BILLED_CONSUMP"]]
        if val is None or val == '':
            continue
        key = (year, month, rc, sub_class, DEMAND_SCAT[scat])
        agg[key] = agg.get(key, 0.0) + float(val)
        kept += 1
    if is_real_xlsx(fp):
        wb.close()
    else:
        fh.close()
    row_counts[fp] = (n, kept)
    print(f"{fp}: {n} rows scanned, {kept} demand rows kept, {time.time()-tf:.1f}s")

print("TOTAL elapsed", round(time.time()-t0,1), "s")
print("aggregated keys:", len(agg))
print("unmapped SERV_RATE codes:", unmapped_srat)

with open("demand_actuals_agg.json", "w") as f:
    json.dump({"agg": [[list(k), v] for k, v in agg.items()], "unmapped": unmapped_srat, "row_counts": row_counts}, f)
print("wrote demand_actuals_agg.json")
