# Scan April-26 billing detail (RT40/50/70) keyed by Cust_Code; merge into cust_accounts.json as 'Apr26'.
import openpyxl, json, time
PATH='Billing Details Report_12May2026.xlsx'
NUM=lambda v: float(v) if isinstance(v,(int,float)) else 0.0
def grp(c):
    c=str(c).upper()
    if c[:3] in ('RT4','NB4'): return 'RT40'
    if c[:3]=='RT5': return 'RT50'
    if c[:3]=='RT7': return 'RT70'
    return None
wb=openpyxl.load_workbook(PATH, read_only=True)
detail=[s for s in wb.sheetnames if s.upper().startswith('BILLING_DETAILS_REPORT')][0]
ws=wb[detail]; it=ws.iter_rows(values_only=True)
hdr=None
for r in it:
    if r and r[0]=='Cust_Code': hdr=list(r); break
I={n:k for k,n in enumerate(hdr) if n}
g=lambda r,n: NUM(r[I[n]])
en=lambda r: g(r,'KWHP_KWH_Energy')+g(r,'KWHL_Energy')+g(r,'KWHO_Energy')
dm=lambda r: g(r,'KVAP_KVA_Demand')+g(r,'KVAL_Demand')+g(r,'KVAO_Demand')
fu=lambda r: g(r,'fuel')+g(r,'FuelOffPeak')+g(r,'FuelPartialPeak')+g(r,'FuelOnPeak')
acc={}; n=0; t=time.time()
for r in it:
    code=r[I['Cust_Code']]
    if code in (None,'','Cust_Code'): continue
    n+=1
    gr=grp(r[I['Srat_Code']])
    if not gr: continue
    a=acc.get(code)
    if a is None: a=acc[code]={'name':(r[I['Name']] or '').strip(),'m':[0.0]*7,'cls':{}}
    kwh=g(r,'net_kwh_billed_consump')
    a['cls'][gr]=a['cls'].get(gr,0)+kwh
    mm=a['m']
    mm[0]+=kwh; mm[1]+=g(r,'net_revenue'); mm[2]+=en(r); mm[3]+=dm(r); mm[4]+=fu(r); mm[5]+=g(r,'IPP_Charge'); mm[6]+=g(r,'kva_billed_consump')
wb.close()
out={}
for code,a in acc.items():
    cls=max(a['cls'],key=a['cls'].get) if a['cls'] else 'NA'
    out[str(code)]=[a['name'],cls]+[round(x) for x in a['m']]
res=json.load(open('cust_accounts.json'))
res['Apr26']={'acc':out,'nrows':n,'secs':round(time.time()-t)}
json.dump(res,open('cust_accounts.json','w'))
print('done Apr26 codes',len(out),'in',round(time.time()-t),'s',flush=True)
