# -*- coding: utf-8 -*-
import openpyxl
wb=openpyxl.load_workbook('Billing Details Report_May 2026.xlsx', read_only=True)
detail=None
for sn in wb.sheetnames:
    ws=wb[sn]
    for i,r in enumerate(ws.iter_rows(values_only=True)):
        if r and r[0]=='Cust_Code': detail=sn; break
        if i>8: break
    if detail: break
ws=wb[detail]; it=ws.iter_rows(values_only=True)
hdr=None
for r in it:
    if r and r[0]=='Cust_Code': hdr=list(r); break
print('SHEET:',detail,'| cols:',len(hdr))
for i,h in enumerate(hdr): print(f'  [{i}] {h}')
# sample 5 data rows for parish-like / srat columns
cand=[i for i,h in enumerate(hdr) if h and any(k in str(h).lower() for k in ['parish','town','district','region','area','loc'])]
print('PARISH-LIKE COLUMN INDEXES:',[(i,hdr[i]) for i in cand])
n=0
for r in it:
    if r[0] in (None,'','Cust_Code'): continue
    print('  sample srat=',r[hdr.index('Srat_Code')] if 'Srat_Code' in hdr else '?', '| parishvals=',{hdr[i]:r[i] for i in cand})
    n+=1
    if n>=6: break
wb.close()
