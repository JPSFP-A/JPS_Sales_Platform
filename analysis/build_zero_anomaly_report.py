# -*- coding: utf-8 -*-
# Cross-rate-class zero-consumption anomaly report.
# Two populations, two precision levels (see _zero_streak_analysis.py header):
#  - Premise-level (RT20-Commercial, RT40, RT50, RT60-ST, RT70): true per-account
#    consecutive-months-at-zero-kWh streak, computed from jps_actuals.
#  - RT10 / RT20-residential (bucket-aggregate, no per-customer identity): monthly
#    zero-consumption customer COUNT only, no per-customer streak possible.
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Calibri"
BLUE_FILL = PatternFill("solid", fgColor="1F4E78")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="1F4E78")
SUB_FONT = Font(name=FONT, bold=True, color="1F4E78", size=11)
BOLD = Font(name=FONT, bold=True)
THIN = Side(style="thin", color="B7C6D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUMFMT = "#,##0;(#,##0)"
MONEYFMT = "#,##0;(#,##0)"

d = json.load(open("_zero_streak_result.json"))
streaks = d["streaks"]
rt10_monthly = d["rt10_monthly_zero_count"]
rt20_res_monthly = d["rt20_residential_monthly_zero_count"]

wb = openpyxl.Workbook()
wb.remove(wb.active)


def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ===== Read Me =====
ws = wb.create_sheet("Read Me")
ws.sheet_view.showGridLines = False
ws["B2"] = "Zero-Consumption Anomaly Report — All Rate Classes"
ws["B2"].font = TITLE_FONT
notes = [
    "",
    "Scope: how many customers have zero consumption, and for how long, across every rate class.",
    "",
    "TWO DIFFERENT PRECISION LEVELS, BY DATA GRAIN:",
    "",
    "1. Premise-level classes (RT20-Commercial, RT40, RT50, RT60-ST, RT70) have real per-account (jps_ac)",
    "   identity across months in jps_actuals. 'Streaks' sheet shows a TRUE consecutive-months-at-zero-kWh",
    "   count per account, computed by walking each account's monthly kWh history backward from its most",
    "   recent record. A gap in an account's monthly presence breaks the streak (treated conservatively).",
    "",
    "2. RT10 and RT20's residential-style (no-NAICS) population are bucket-aggregates with NO per-customer",
    "   identity — only a monthly zero-consumption CUSTOMER COUNT is possible, not a per-customer streak.",
    "   See 'RT10-RT20 Residential Trend' sheet.",
    "",
    "KEY FINDING (CORRECTED from an earlier draft of this report — see below): zero-kWh accounts are NOT",
    "$0-revenue accounts. Traced a sample of the longest streaks (National Water Commission, Ministry of",
    "Health, Bank of Nova Scotia, JPS's own account, and others) back to the raw CIS billing extract and",
    "confirmed they bill a real minimum/standby demand charge every month even at 0 metered kWh — present",
    "in JPS's own source file, not a pipeline artifact. The tell: many unrelated customers (from JPS itself",
    "down to individual people) share bit-identical demand_jmd/ipp_jmd/customer_charge_jmd figures every",
    "month, which is the signature of a standardized minimum-bill tariff tier, not duplicate records. The",
    "'is_minbill_cluster' column on the 'Streaks' sheet flags these — 0 rows in this dataset actually have",
    "$0 current-month revenue.",
    "",
    "So what's still worth flagging: these premises have registered NO metered consumption for the ENTIRE",
    "dataset window (up to 19 straight months) and are being billed at a default minimum instead. Billing",
    "math looks correct; the open question is operational — is the meter at these premises actually being",
    "read at all, or has metering lapsed for over a year? Worth a CIS/field-ops check, not a revenue-",
    "integrity one. Accounts flagged is_minbill_cluster=False are the more interesting case: an isolated,",
    "unshared billing profile at zero for 12+ months looks more like a genuine dormant/closed account.",
]
r = 4
for line in notes:
    ws.cell(row=r, column=2, value=line)
    if line.startswith(("1.", "2.", "TWO", "KEY")):
        ws.cell(row=r, column=2).font = BOLD
    else:
        ws.cell(row=r, column=2).font = Font(name=FONT, size=10)
    r += 1
autosize(ws, [3, 118])

# ===== Summary =====
ws = wb.create_sheet("Summary")
ws.sheet_view.showGridLines = False
ws["B2"] = "Ongoing Zero-Consumption Streaks — Summary"
ws["B2"].font = TITLE_FONT
row = 4
ws.cell(row=row, column=2, value="By rate class").font = SUB_FONT
row += 1
hdr = ["Rate Class", "Accounts w/ ongoing zero streak", "2+ months", "3+ months", "6+ months", "12+ months (whole window)"]
for i, h in enumerate(hdr):
    ws.cell(row=row, column=2 + i, value=h)
style_header(ws, row, len(hdr), start_col=2)
row += 1
by_rc = {}
for s in streaks:
    by_rc.setdefault(s["rate_class"], []).append(s["streak_len"])
for rc in sorted(by_rc, key=lambda k: -len(by_rc[k])):
    lens = by_rc[rc]
    ws.cell(row=row, column=2, value=rc)
    ws.cell(row=row, column=3, value=len(lens)).number_format = NUMFMT
    ws.cell(row=row, column=4, value=sum(1 for x in lens if x >= 2)).number_format = NUMFMT
    ws.cell(row=row, column=5, value=sum(1 for x in lens if x >= 3)).number_format = NUMFMT
    ws.cell(row=row, column=6, value=sum(1 for x in lens if x >= 6)).number_format = NUMFMT
    ws.cell(row=row, column=7, value=sum(1 for x in lens if x >= 12)).number_format = NUMFMT
    row += 1
row += 2
all_lens = [s["streak_len"] for s in streaks]
ws.cell(row=row, column=2, value="TOTAL").font = BOLD
ws.cell(row=row, column=3, value=len(all_lens)).number_format = NUMFMT
ws.cell(row=row, column=4, value=sum(1 for x in all_lens if x >= 2)).number_format = NUMFMT
ws.cell(row=row, column=5, value=sum(1 for x in all_lens if x >= 3)).number_format = NUMFMT
ws.cell(row=row, column=6, value=sum(1 for x in all_lens if x >= 6)).number_format = NUMFMT
ws.cell(row=row, column=7, value=sum(1 for x in all_lens if x >= 12)).number_format = NUMFMT
for c in range(2, 8):
    ws.cell(row=row, column=c).font = BOLD
row += 3
ws.cell(row=row, column=2, value=("Note: 'streak' = consecutive months (no gaps) ending at the account's most recent "
                                    "billing record, all at exactly 0 kWh. 12+ months means the account has been at zero "
                                    "for the entire dataset window (Jan 2025-Jul 2026) with no positive-consumption month "
                                    "on record at all. See Read Me: this does NOT mean $0 revenue — most of these are on "
                                    "a minimum-bill tariff and still generate real revenue every month."))
ws.cell(row=row, column=2).font = Font(name=FONT, italic=True, size=10)
ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
row += 2
n_minbill = sum(1 for s in streaks if s["is_minbill_cluster"])
n_isolated = len(streaks) - n_minbill
n_zero_rev = sum(1 for s in streaks if s["current_monthly_revenue"] == 0)
ws.cell(row=row, column=2, value="Minimum-bill tariff vs. isolated zero-revenue-risk accounts").font = SUB_FONT
row += 1
hdr2 = ["", "Count", "Meaning"]
for i, h in enumerate(hdr2):
    ws.cell(row=row, column=2 + i, value=h)
style_header(ws, row, len(hdr2), start_col=2)
row += 1
for label, val, meaning in [
    ("On a shared minimum-bill tariff (is_minbill_cluster=True)", n_minbill, "Real revenue every month; billing is working as designed. Operational question: is the meter being read?"),
    ("Isolated billing profile (is_minbill_cluster=False)", n_isolated, "Not on a known shared minimum-bill default — could be a legitimate unique demand contract (several are real, large accounts) or worth a closer look. Not conclusive on its own."),
    ("Genuinely $0 current-month revenue", n_zero_rev, f"The unambiguous finding: these {n_zero_rev} accounts generate no revenue at all, not even a minimum bill. Worth reviewing individually — see names below."),
]:
    ws.cell(row=row, column=2, value=label)
    ws.cell(row=row, column=3, value=val).number_format = NUMFMT
    ws.cell(row=row, column=4, value=meaning)
    row += 1
row += 2
zero_rev_accts = [s for s in streaks if s["current_monthly_revenue"] == 0]
if zero_rev_accts:
    ws.cell(row=row, column=2, value="The genuinely $0-revenue accounts").font = SUB_FONT
    row += 1
    hdr3 = ["Account", "Name", "Rate Class", "Streak (months)", "Total Revenue During Streak J$"]
    for i, h in enumerate(hdr3):
        ws.cell(row=row, column=2 + i, value=h)
    style_header(ws, row, len(hdr3), start_col=2)
    row += 1
    for s in zero_rev_accts:
        ws.cell(row=row, column=2, value=s["jps_ac"])
        ws.cell(row=row, column=3, value=s["name"])
        ws.cell(row=row, column=4, value=s["rate_class"])
        ws.cell(row=row, column=5, value=s["streak_len"]).number_format = NUMFMT
        ws.cell(row=row, column=6, value=s["streak_total_revenue"]).number_format = MONEYFMT
        row += 1
autosize(ws, [3, 16, 24, 12, 12, 12, 22])

# ===== Streaks (detail) =====
ws = wb.create_sheet("Streaks")
ws.sheet_view.showGridLines = False
ws["B2"] = "Account-Level Detail — Ongoing Zero-Consumption Streaks (longest first)"
ws["B2"].font = TITLE_FONT
row = 4
hdr = ["Account", "Name", "Rate Class", "Streak (months)", "Streak Start", "Streak End",
       "Current Monthly Revenue J$", "Total Revenue During Streak J$", "On Minimum-Bill Tariff?",
       "Account Status (as of streak end)", "Suspended?"]
for i, h in enumerate(hdr):
    ws.cell(row=row, column=2 + i, value=h)
style_header(ws, row, len(hdr), start_col=2)
row += 1
for s in streaks:
    ws.cell(row=row, column=2, value=s["jps_ac"])
    ws.cell(row=row, column=3, value=s["name"])
    ws.cell(row=row, column=4, value=s["rate_class"])
    c = ws.cell(row=row, column=5, value=s["streak_len"])
    c.number_format = NUMFMT
    # The real priority flag: an ACTIVE, non-suspended account generating $0 revenue.
    # 12+ months isolated (not on a minbill tariff) is a secondary, softer flag.
    is_top_priority = s.get("account_status") == "A" and s["current_monthly_revenue"] == 0
    is_soft_flag = s["streak_len"] >= 12 and not s["is_minbill_cluster"]
    if is_top_priority:
        c.fill = WARN_FILL
    elif is_soft_flag:
        c.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=row, column=6, value=s["streak_start"])
    ws.cell(row=row, column=7, value=s["streak_end"])
    ws.cell(row=row, column=8, value=s["current_monthly_revenue"]).number_format = MONEYFMT
    ws.cell(row=row, column=9, value=s["streak_total_revenue"]).number_format = MONEYFMT
    ws.cell(row=row, column=10, value="Yes" if s["is_minbill_cluster"] else "No — isolated")
    cstatus = ws.cell(row=row, column=11, value=s.get("account_status_label", "?"))
    csusp = ws.cell(row=row, column=12, value=s.get("is_suspended") or "")
    if is_top_priority:
        cstatus.fill = WARN_FILL
        csusp.fill = WARN_FILL
    row += 1
autosize(ws, [3, 22, 34, 12, 16, 14, 14, 18, 20, 18, 20, 12])
ws.freeze_panes = "B5"

# ===== Account Status Breakdown =====
ws = wb.create_sheet("Account Status Breakdown")
ws.sheet_view.showGridLines = False
ws["B2"] = "Account Status of Flagged Zero-Consumption Accounts (from the raw CIS extract)"
ws["B2"].font = TITLE_FONT
ws["B4"] = ("Account_Status / is_suspended pulled from the raw billing file for each account's most recent "
            "(streak-end) month. This is the real breakdown behind the minimum-bill-tariff finding: Inactive/"
            "Final accounts correctly billing $0 need no action; an Active, non-suspended account billing $0 is "
            "the priority case.")
ws["B4"].font = Font(name=FONT, italic=True, size=9, color="7F7F7F")
ws["B4"].alignment = Alignment(wrap_text=True)
row = 6
ws.cell(row=row, column=2, value="By account status").font = SUB_FONT
row += 1
hdr = ["Account Status", "Count", "Of which: on minimum-bill tariff", "Of which: $0 current revenue"]
for i, h in enumerate(hdr):
    ws.cell(row=row, column=2 + i, value=h)
style_header(ws, row, len(hdr), start_col=2)
row += 1
from collections import Counter
status_labels = sorted(set(s.get("account_status_label", "?") for s in streaks),
                        key=lambda lbl: -sum(1 for s in streaks if s.get("account_status_label") == lbl))
for lbl in status_labels:
    grp = [s for s in streaks if s.get("account_status_label") == lbl]
    ws.cell(row=row, column=2, value=lbl)
    ws.cell(row=row, column=3, value=len(grp)).number_format = NUMFMT
    ws.cell(row=row, column=4, value=sum(1 for s in grp if s["is_minbill_cluster"])).number_format = NUMFMT
    c = ws.cell(row=row, column=5, value=sum(1 for s in grp if s["current_monthly_revenue"] == 0))
    c.number_format = NUMFMT
    if lbl == "Active" and c.value:
        c.fill = WARN_FILL
    row += 1
row += 2
top_priority = [s for s in streaks if s.get("account_status") == "A" and s["current_monthly_revenue"] == 0]
ws.cell(row=row, column=2, value=f"Top-priority accounts: Active status + $0 current revenue ({len(top_priority)})").font = SUB_FONT
row += 1
if top_priority:
    hdr2 = ["Account", "Name", "Rate Class", "Streak (months)", "Streak End"]
    for i, h in enumerate(hdr2):
        ws.cell(row=row, column=2 + i, value=h)
    style_header(ws, row, len(hdr2), start_col=2)
    row += 1
    for s in top_priority:
        ws.cell(row=row, column=2, value=s["jps_ac"])
        ws.cell(row=row, column=3, value=s["name"])
        ws.cell(row=row, column=4, value=s["rate_class"])
        ws.cell(row=row, column=5, value=s["streak_len"]).number_format = NUMFMT
        ws.cell(row=row, column=6, value=s["streak_end"])
        row += 1
else:
    ws.cell(row=row, column=2, value="None — every flagged account with $0 current revenue is formally Inactive/Final/etc., not Active.")
autosize(ws, [3, 30, 30, 16, 14])

# ===== RT10 / RT20 Residential Trend =====
ws = wb.create_sheet("RT10-RT20 Residential Trend")
ws.sheet_view.showGridLines = False
ws["B2"] = "RT10 / RT20-Residential — Monthly Zero-Consumption Customer Count"
ws["B2"].font = TITLE_FONT
ws["B4"] = "No per-customer identity exists for this population (bucket-aggregate source) — this is a monthly count, not a per-customer streak."
ws["B4"].font = Font(name=FONT, italic=True, size=9, color="7F7F7F")
row = 6
hdr = ["Month", "RT10 Zero-Consumption Customers", "RT20-Residential Zero-Consumption Customers"]
for i, h in enumerate(hdr):
    ws.cell(row=row, column=2 + i, value=h)
style_header(ws, row, len(hdr), start_col=2)
row += 1
months = sorted(set(rt10_monthly) | set(rt20_res_monthly))
for mo in months:
    ws.cell(row=row, column=2, value=mo)
    ws.cell(row=row, column=3, value=rt10_monthly.get(mo)).number_format = NUMFMT
    ws.cell(row=row, column=4, value=rt20_res_monthly.get(mo)).number_format = NUMFMT
    row += 1
autosize(ws, [3, 14, 30, 34])

wb.save("Zero_Consumption_Anomaly_Report.xlsx")
print("wrote Zero_Consumption_Anomaly_Report.xlsx")
