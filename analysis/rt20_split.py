# Splits RT20 (General Service) per business rule: customers WITH a NAICS Code
# are real registered businesses -> individual commercial accounts (premise-level,
# matching the RT40+ pattern). Customers WITHOUT a NAICS Code are residential-style
# -> bucketed by kWh consumption range + parish, matching the scheme already used
# for RT10: <Zero, Zero, <150, 150>350, 350>550, 550>750, 750>950, over 950.
#
# Generalized from rt20_split_jun26.py (originally hardcoded to June 2026) so it
# can be rerun each month against that month's "Billing Details Report_<Mon> <YYYY>"
# file. Usage: python rt20_split.py <YYYY> <M> "<billing file path>"
import json, sys, csv

YEAR = int(sys.argv[1])
MONTH = int(sys.argv[2])
SRC = sys.argv[3]
OUT = f'rt20_split_{YEAR}_{MONTH:02d}.json'

# Row inclusion MUST match corrected_scan.py's title_of() exactly (rate_class checked
# first, Srat_Code only as fallback) — a naive Srat_Code=='RT20' filter disagrees with
# it on some rows (found via a ~5M kWh reconciliation gap against corrected.json for
# Jan 2025), silently mis-scoping which rows count as RT20 at all.
def norm(s):
    return str(s).strip().upper()


_RMAP = {}
_RCMAP = {}
for _r in list(csv.reader(open(r'C:\Users\jwilson\Downloads\Rate categorry Data mapping.csv')))[1:]:
    if len(_r) >= 3:
        _RMAP[norm(_r[2])] = _r[0]
        if norm(_r[1]) not in _RCMAP:
            _RCMAP[norm(_r[1])] = _r[0]


def title_of(rc, srat):
    return _RCMAP.get(norm(rc)) or _RMAP.get(norm(srat))


def num(v):
    if v is None:
        return 0.0
    v = str(v).strip()
    if v == '':
        return 0.0
    try:
        return float(v)
    except ValueError:
        return 0.0


def bucket_of(kwh):
    if kwh < 0:
        return '<Zero'
    if kwh == 0:
        return 'Zero'
    if kwh < 150:
        return '<150'
    if kwh < 350:
        return '150>350'
    if kwh < 550:
        return '350>550'
    if kwh < 750:
        return '550>750'
    if kwh < 950:
        return '750>950'
    return 'over 950'


comm = {}   # jps_ac -> {name, parish, v:[kwh,rev,dem,fuel,energy,ipp,cust,gct]}
res = {}    # (parish, bucket) -> [kwh,rev,gct,cust_count]


def _is_zip_xlsx(path):
    with open(path, 'rb') as fh:
        return fh.read(2) == b'PK'


def _rows():
    if _is_zip_xlsx(SRC):
        import openpyxl
        wb = openpyxl.load_workbook(SRC, read_only=True)
        detail = None
        for sn in wb.sheetnames:
            ws = wb[sn]
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if r and r[0] == 'Cust_Code':
                    detail = sn
                    break
                if i > 8:
                    break
            if detail:
                break
        ws = wb[detail]
        for r in ws.iter_rows(values_only=True):
            yield [('' if c is None else c) for c in r]
        wb.close()
    elif SRC.lower().endswith('.csv'):
        # 2025-vintage raw exports: comma-delimited, quoted fields (addresses contain commas).
        import csv
        with open(SRC, encoding='utf-8', errors='replace', newline='') as f:
            for r in csv.reader(f):
                yield r
    else:
        with open(SRC, encoding='latin-1', errors='replace') as f:
            for line in f:
                yield line.rstrip('\n').split('\t')


hdr = None
idx = None
n_total = 0
n_billed = 0
for r in _rows():
    if r and r[0] == 'Cust_Code':
        hdr = r
        idx = {name: i for i, name in enumerate(hdr)}
        continue
    if hdr is None:
        continue
    code = r[idx['Cust_Code']] if idx.get('Cust_Code', -1) < len(r) else None
    if code in (None, '', 'Cust_Code'):
        continue
    srat = r[idx['Srat_Code']] if idx.get('Srat_Code', -1) < len(r) else ''
    rate_class_raw = r[idx['rate_class']] if idx.get('rate_class', -1) < len(r) else ''
    title = title_of(rate_class_raw, srat)
    if title != 'RT20':
        continue
    n_total += 1
    cb = r[idx['cust_billed']] if 'cust_billed' in idx and idx['cust_billed'] < len(r) else None
    if cb is not None and str(cb).strip() in ('0', '0.0'):
        continue
    n_billed += 1

    # 2026 exports call this column 'NAICS Code'; 2025-vintage exports call the same
    # business-classification field 'sicc_code' (Standard Industrial Classification,
    # NAICS's predecessor) — same business rule, different column name.
    naics_col = 'NAICS Code' if 'NAICS Code' in idx else ('sicc_code' if 'sicc_code' in idx else None)
    naics = str(r[idx[naics_col]] if naics_col and idx[naics_col] < len(r) else '').strip()
    prem = str(r[idx['Prem_Code']] or '').strip() if idx.get('Prem_Code', -1) < len(r) else ''
    parish = str(r[idx['Parish']] or '').strip() if idx.get('Parish', -1) < len(r) else 'UNMAPPED'
    name = str(r[idx['Name']] or '').strip() if idx.get('Name', -1) < len(r) else ''

    kwh = num(r[idx['net_kwh_billed_consump']]) if 'net_kwh_billed_consump' in idx else 0.0
    rev = num(r[idx['net_revenue']]) if 'net_revenue' in idx else 0.0
    gct = num(r[idx['GCT']]) if 'GCT' in idx else 0.0
    dem = (num(r[idx['KVAP_KVA_Demand']]) if 'KVAP_KVA_Demand' in idx else 0.0)
    en = (num(r[idx['KWHP_KWH_Energy']]) if 'KWHP_KWH_Energy' in idx else 0.0) + \
         (num(r[idx['KWHL_Energy']]) if 'KWHL_Energy' in idx else 0.0) + \
         (num(r[idx['KWHO_Energy']]) if 'KWHO_Energy' in idx else 0.0)
    fu = (num(r[idx['fuel']]) if 'fuel' in idx else 0.0) + \
         (num(r[idx['FuelOffPeak']]) if 'FuelOffPeak' in idx else 0.0) + \
         (num(r[idx['FuelPartialPeak']]) if 'FuelPartialPeak' in idx else 0.0) + \
         (num(r[idx['FuelOnPeak']]) if 'FuelOnPeak' in idx else 0.0)
    ipp = num(r[idx['IPP_Charge']]) if 'IPP_Charge' in idx else 0.0
    cust_chg = num(r[idx['Cust_Charge']]) if 'Cust_Charge' in idx else 0.0

    if naics:
        if not prem:
            continue
        jps_ac = str(code) + '-' + prem
        b = comm.get(jps_ac)
        if b is None:
            b = comm[jps_ac] = {'name': name, 'parish': parish, 'v': [0.0]*8}
        v = b['v']
        v[0] += kwh; v[1] += rev; v[2] += dem; v[3] += fu; v[4] += en; v[5] += ipp; v[6] += cust_chg; v[7] += gct
    else:
        bucket = bucket_of(kwh)
        key = (parish, bucket)
        b = res.get(key)
        if b is None:
            b = res[key] = [0.0, 0.0, 0.0, 0]
        b[0] += kwh; b[1] += rev; b[2] += gct; b[3] += 1

print('RT20 raw rows:', n_total, '| billed (cust_billed<>0):', n_billed)
print('commercial (NAICS) accounts:', len(comm))
print('residential bucket cells:', len(res))

json.dump({'year': YEAR, 'month': MONTH, 'comm': comm,
           'res': {f'{k[0]}||{k[1]}': v for k, v in res.items()}},
          open(OUT, 'w'))
print('wrote', OUT)
