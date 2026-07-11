# -*- coding: utf-8 -*-
# Apr & May 2026 per (customer, rate-class) — so multi-class customers (e.g. Bay Roc RT40+RT50)
# split correctly. Captures kWh, demand charge (J$), billed kVA. Non-RT10/RT20 (by-customer set).
import openpyxl, json, time, csv
DL=r'C:\Users\jwilson\Downloads'
FILES={'Apr':'Billing Details Report_Apr 2026.xlsx','May':'Billing Details Report_May 2026.xlsx'}
NUM=lambda v: float(v) if isinstance(v,(int,float)) else 0.0
def norm(s): return str(s).strip().upper()
RMAP={};RCMAP={}
for r in list(csv.reader(open(DL+r'\Rate categorry Data mapping.csv')))[1:]:
    if len(r)>=3:
        RMAP[norm(r[2])]=r[0]
        if norm(r[1]) not in RCMAP: RCMAP[norm(r[1])]=r[0]
def baseT(t): return 'RT60' if t=='RT60-ST' else t
def title_of(srat,rc): return baseT(RCMAP.get(norm(rc)) or RMAP.get(norm(srat)) or 'Other')
def proc(path):
    wb=openpyxl.load_workbook(path, read_only=True)
    detail=None
    for sn in wb.sheetnames:
        ws=wb[sn]
        for i,r in enumerate(ws.iter_rows(values_only=True)):
            if r and r[0]=='Cust_Code': detail=sn; break
            if i>8: break
        if detail: break
    ws=wb[detail]; it=ws.iter_rows(values_only=True)
    hdr=None
    for r in it:
        if r and r[0]=='Cust_Code': hdr=list(r); break
    I={n:k for k,n in enumerate(hdr) if n}
    g=lambda r,n: NUM(r[I[n]]) if n in I else 0.0
    acc={}
    for r in it:
        code=r[I['Cust_Code']]
        if code in (None,'','Cust_Code'): continue
        srat=str(r[I['Srat_Code']]); rc=str(r[I['rate_class']]); t=title_of(srat,rc)
        if t in ('RT10','RT20'): continue
        key=(str(code),t)
        a=acc.get(key)
        if a is None: a=acc[key]={'name':str(r[I['Name']] or '').strip(),'kwh':0.0,'dem':0.0,'kva':0.0}
        a['kwh']+=g(r,'net_kwh_billed_consump')
        a['dem']+=g(r,'KVAP_KVA_Demand')+g(r,'KVAL_Demand')+g(r,'KVAO_Demand')
        a['kva']+=g(r,'kva_billed_consump')
    wb.close()
    return acc
OUT={}
for mo,f in FILES.items():
    t=time.time(); print('processing',mo,flush=True)
    acc=proc(f)
    for (code,title),a in acc.items():
        rec=OUT.setdefault(code+'|'+title,{'code':code,'title':title,'name':a['name']})
        rec[mo]=[round(a['kwh'],1),round(a['dem'],1),round(a['kva'],1)]
    json.dump(OUT,open('custclass.json','w'))
    print('  done',mo,'rows',len(acc),'in',round(time.time()-t),'s',flush=True)
print('WROTE custclass.json keys',len(OUT),flush=True)
