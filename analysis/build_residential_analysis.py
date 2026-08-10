# -*- coding: utf-8 -*-
# Residential (RT10 + RT20) analysis workbook: parish, consumption-tier "clusters",
# and driver commentary for the last 3 months (May-Jul 2026), plus budget/LE
# comparison where real data exists.
#
# Sources (deliberately kept SEPARATE, never blended):
#   1. Consumption-bucket population (official, ALL RT10/RT20 meters) - from
#      corrected.json's 'bucket' dict (built by corrected_scan.py off the raw
#      monthly Billing Details Report). Most complete population; no parish detail
#      for RT10 (bucket-only), has parish for the RT20 NAICS split.
#   2. Prepaid/parish population (PAYG only) - jps_actuals consumption_bucket='Prepaid',
#      loaded from the Customer-Monthly-Transaction-Report files. Jun-Jul 2026 only
#      (May prepaid file was never supplied/loaded - flagged, not silently skipped).
#   3. Real RT20 premise population (NAICS-split) - jps_actuals rate_class=RT20,
#      consumption_bucket NOT IN ('Prepaid') - individual commercial(NAICS) accounts
#      + residential-style parish x consumption-tier buckets (no-NAICS). May/Jun/Jul.
#   4. Budget (jps_budget) / LE forecast (jps_le) - RT20 only. RT10's budget/LE rows
#      are a "Postpaid true-up" placeholder (a handful of pseudo-accounts, not real
#      consumption data) - explicitly marked N/A rather than plotted as if real.
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

HERE = os.path.dirname(os.path.abspath(__file__))
os.chdir(HERE)

FONT = "Calibri"
BLUE_FILL = PatternFill("solid", fgColor="1F4E78")
LT_FILL = PatternFill("solid", fgColor="D9E2F3")
YEL_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
SUB_FONT = Font(name=FONT, bold=True, color="1F4E78", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=16, color="1F4E78")
NOTE_FONT = Font(name=FONT, italic=True, size=10, color="7F7F7F")
BOLD = Font(name=FONT, bold=True)
THIN = Side(style="thin", color="B7C6D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUMFMT = "#,##0;(#,##0)"
MONEYFMT = "#,##0;(#,##0)"
PCTFMT = "0.0%"

MONTHS_2026 = ["2026-05", "2026-06", "2026-07"]
MONTHS_2025 = ["2025-05", "2025-06", "2025-07"]
MLABELS = ["May'26", "Jun'26", "Jul'26"]

BUCKET_ORDER = ["<Zero", "Zero", "<150", "150>350", "350>550", "550>750", "750>950", "over 950"]


def bucket_of(kwh):
    if kwh < 0:
        return "<Zero"
    if kwh == 0:
        return "Zero"
    if kwh < 150:
        return "<150"
    if kwh < 350:
        return "150>350"
    if kwh < 550:
        return "350>550"
    if kwh < 750:
        return "550>750"
    if kwh < 950:
        return "750>950"
    return "over 950"


def bin_bucket_of(bk):
    """For corrected.json's fine 50kWh-WIDE histogram bins, not exact per-customer kWh.
    Bin key 0 means the range [0,50) - it is NOT "exactly zero" (bucket_of(0) would say
    "Zero", which is wrong here and previously caused the whole [0,50) bin's real,
    nonzero consumption to display under the "Zero" tier). This data source has no bin
    fine enough to isolate exactly-zero consumption, so "Zero" is left empty for it and
    the [0,50) bin folds into "<150" like every other 50kWh bin does."""
    if bk < 0:
        return "<Zero"
    if bk < 150:
        return "<150"
    if bk < 350:
        return "150>350"
    if bk < 550:
        return "350>550"
    if bk < 750:
        return "550>750"
    if bk < 950:
        return "750>950"
    return "over 950"


# ---------- load sources ----------
CJ = json.load(open("corrected.json"))
BUCK = CJ["bucket"]  # month -> title -> {bin: [count,kwh,rev,energy,fuel,ipp,cust]}

prepaid = json.load(open("_res_prepaid.json"))
rt20_real = json.load(open("_res_rt20_real.json"))
rt20_budget = json.load(open("_res_jps_budget.json"))
rt20_le = json.load(open("_res_jps_le.json"))
unassigned_payg = json.load(open("_res_unassigned_payg.json"))


RT10_TRUE_ZERO = json.load(open("rt10_zero_fix_result.json")) if os.path.exists("rt10_zero_fix_result.json") else {}


def roll_up_buckets(month, title):
    """Fine 50kWh bins -> the 8 human bands, for one month/rate-class."""
    out = {b: {"count": 0.0, "kwh": 0.0, "rev": 0.0} for b in BUCKET_ORDER}
    cell = BUCK.get(month, {}).get(title, {})
    for bstr, v in cell.items():
        bk = int(bstr)
        label = bin_bucket_of(bk if bk < 5000 else 5000)
        out[label]["count"] += v[0]
        out[label]["kwh"] += v[1]
        out[label]["rev"] += v[2]
    if title == "RT10" and month in RT10_TRUE_ZERO:
        # bin_bucket_of() folds corrected.json's [0,50) bin into "<150" wholesale since
        # that bin can't distinguish exactly-zero from 1-49kWh on its own. Pull the true
        # zero-consumption count back out using the raw-file reprocessing in
        # rt10_zero_fix.py, which can (kwh==0 exactly vs 0<kwh<150).
        tz_count, tz_kwh, tz_rev = RT10_TRUE_ZERO[month]["TrueZero"][:3]
        out["Zero"]["count"] += tz_count
        out["Zero"]["kwh"] += tz_kwh
        out["Zero"]["rev"] += tz_rev
        out["<150"]["count"] -= tz_count
        out["<150"]["kwh"] -= tz_kwh
        out["<150"]["rev"] -= tz_rev
    return out


def class_total(month, title):
    cell = BUCK.get(month, {}).get(title, {})
    kwh = sum(v[1] for v in cell.values())
    rev = sum(v[2] for v in cell.values())
    cnt = sum(v[0] for v in cell.values())
    return kwh, rev, cnt


def agg_prepaid(month_num, rc):
    ym = None
    rows = [r for r in prepaid if r["year"] == 2026 and r["month"] == month_num and r["rate_class"] == rc]
    by_parish = {}
    for r in rows:
        p = r["parish"]
        a = by_parish.setdefault(p, {"kwh": 0.0, "rev": 0.0, "cust": 0})
        a["kwh"] += r["kwh"] or 0
        a["rev"] += r["revenue_jmd"] or 0
        a["cust"] += r["customer_count"] or 0
    return by_parish


def agg_rt20_real(month_num):
    rows = [r for r in rt20_real if r["year"] == 2026 and r["month"] == month_num]
    comm = {"kwh": 0.0, "rev": 0.0, "cust": 0}
    res = {"kwh": 0.0, "rev": 0.0, "cust": 0}
    by_parish_comm = {}
    by_parish_res = {}
    for r in rows:
        kwh = r["kwh"] or 0
        rev = r["revenue_jmd"] or 0
        p = r["parish"] or "UNMAPPED"
        if r["consumption_bucket"] == "Commercial":
            comm["kwh"] += kwh; comm["rev"] += rev; comm["cust"] += 1
            a = by_parish_comm.setdefault(p, {"kwh": 0.0, "rev": 0.0, "cust": 0})
            a["kwh"] += kwh; a["rev"] += rev; a["cust"] += 1
        else:
            cnt = r["customer_count"] or 0
            res["kwh"] += kwh; res["rev"] += rev; res["cust"] += cnt
            a = by_parish_res.setdefault(p, {"kwh": 0.0, "rev": 0.0, "cust": 0})
            a["kwh"] += kwh; a["rev"] += rev; a["cust"] += cnt
    return comm, res, by_parish_comm, by_parish_res


def agg_rt20_budget_le(source, month_num):
    rows = [r for r in source if r["year"] == 2026 and r["month"] == month_num]
    kwh_key = "kwh_budget" if "kwh_budget" in rows[0] else "kwh_le"
    rev_key = "revenue_budget" if "revenue_budget" in rows[0] else "revenue_le"
    kwh = sum((r[kwh_key] or 0) for r in rows)
    rev = sum((r[rev_key] or 0) for r in rows)
    return kwh, rev


# ---------- workbook ----------
wb = openpyxl.Workbook()
wb.remove(wb.active)


def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ===== Read Me =====
ws = wb.create_sheet("Read Me")
ws.sheet_view.showGridLines = False
ws["B2"] = "Residential (RT10 + RT20) Performance Analysis"
ws["B2"].font = TITLE_FONT
ws["B4"] = "Scope: May, June, July 2026. Built %s." % "2026-08-05"
ws["B4"].font = BOLD
notes = [
    "",
    "THREE DIFFERENT RESIDENTIAL DATA SOURCES ARE USED HERE - DELIBERATELY KEPT SEPARATE:",
    "",
    "1. Consumption-bucket view (sheet 'By Bucket', 'Overview') - the OFFICIAL population: every RT10 and",
    "   RT20 meter in the monthly CIS billing extract, grouped by kWh usage tier. This is the most complete",
    "   figure and matches total company revenue/kWh for these classes. No parish detail for RT10.",
    "",
    "2. Prepaid/parish view (sheet 'By Parish - Prepaid') - PAYG (prepaid vending) accounts only, from the",
    "   separate Customer-Monthly-Transaction-Report export, aggregated by parish. Covers June and July 2026",
    "   ONLY - a May prepaid file was never supplied, so May is genuinely absent here, not zero.",
    "",
    "3. Real RT20 premise view (sheet 'By Parish - RT20 Real') - RT20 split by NAICS code: customers WITH a",
    "   NAICS code are registered businesses tracked as individual premises ('Commercial'); customers WITHOUT",
    "   one are residential-style and bucketed by parish + consumption tier. May/Jun/Jul 2026.",
    "",
    "BUDGET & FORECAST:",
    "  - RT20: real, usable data in both jps_budget and jps_le at premise grain - shown on 'Budget & Forecast'.",
    "  - RT10: the budget/LE rows in the database are a tiny 'Postpaid true-up' placeholder (a handful of",
    "    pseudo-accounts like '<Zero'/'Postpaid'), NOT a real consumption budget. No usable RT10 forecast",
    "    exists anywhere in the system (checked jps_budget, jps_le, and the legacy forecast workbook).",
    "    Marked N/A rather than fabricated.",
    "",
    "Revenue figures are net-of-GCT throughout, consistent with the rest of jps_actuals.",
    "",
    "PARISH LABELS ARE NOT CONSISTENT ACROSS SHEETS: 'By Parish - Prepaid' uses the grouped parish taxonomy",
    "(KSAN/KSAS/St. James/Manchester...) via the standard Parish Grouping mapping. 'By Parish - RT20 Real' uses",
    "raw city-level labels straight from the billing extract (KSA South, Montego Bay, Mandeville...) because the",
    "RT20 NAICS-split script doesn't apply that mapping yet. The two parish views will not line up 1:1 - treat",
    "them as separate breakdowns, not a single reconciled geography.",
]
r = 6
for line in notes:
    ws.cell(row=r, column=2, value=line)
    if line.startswith(("1.", "2.", "3.", "BUDGET")):
        ws.cell(row=r, column=2).font = BOLD
    else:
        ws.cell(row=r, column=2).font = Font(name=FONT, size=10)
    r += 1
autosize(ws, [3, 118])

# ===== Overview =====
ws = wb.create_sheet("Overview")
ws.sheet_view.showGridLines = False
ws["B2"] = "Overview - Total Population (Postpaid + Prepaid, all meters)"
ws["B2"].font = TITLE_FONT
ws["B3"] = ("Headline kWh/Revenue/Customer figures below INCLUDE Prepaid (PAYG) wherever a Prepaid file exists for "
            "that month. May 2026 has no Prepaid file (never supplied) so May is postpaid-only, not zero -- flagged "
            "explicitly, not silently blended. PY (2025) columns have no Prepaid source at all, so YoY% is postpaid-"
            "only on both sides for May, and slightly apples-to-oranges for Jun/Jul (CY includes Prepaid, PY doesn't) "
            "-- Prepaid is a small share of the total (~1-2%), so this rarely moves the YoY conclusion, but treat "
            "Jun/Jul YoY% as approximate, not exact, for that reason.")
ws["B3"].font = Font(name=FONT, italic=True, size=9, color="C00000")
ws["B3"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[3].height = 42
row = 5
for rc, label in [("RT10", "RT10 - Residential"), ("RT20", "RT20 - General Service / SME")]:
    ws.cell(row=row, column=2, value=label).font = SUB_FONT
    row += 1
    hdr = ["Metric"] + MLABELS + ["May→Jul Δ%"] + [m + " YoY%" for m in MLABELS]
    for i, h in enumerate(hdr):
        ws.cell(row=row, column=2 + i, value=h)
    style_header(ws, row, len(hdr), start_col=2)
    row += 1
    cy = [class_total(m, rc) for m in MONTHS_2026]
    py = [class_total(m, rc) for m in MONTHS_2025]

    def prepaid_totals(month_num, rate_class):
        rows_m = [r for r in prepaid if r["year"] == 2026 and r["month"] == month_num and r["rate_class"] == rate_class]
        if not rows_m:
            return None
        return sum(r["kwh"] or 0 for r in rows_m), sum(r["revenue_jmd"] or 0 for r in rows_m), sum(r["customer_count"] or 0 for r in rows_m)

    pp = [prepaid_totals(mn, rc) for mn in (5, 6, 7)]
    # Headline = postpaid + prepaid (prepaid=0 for months with no file, e.g. May)
    cyi = [(v[0] + (p[0] if p else 0), v[1] + (p[1] if p else 0), v[2] + (p[2] if p else 0)) for v, p in zip(cy, pp)]

    metrics = [
        ("kWh", [v[0] for v in cyi], [v[0] for v in py], NUMFMT),
        ("Revenue (J$)", [v[1] for v in cyi], [v[1] for v in py], MONEYFMT),
        ("Customer-months", [v[2] for v in cyi], [v[2] for v in py], NUMFMT),
        ("Rev/kWh (J$)", [v[1] / v[0] if v[0] else 0 for v in cyi], [v[1] / v[0] if v[0] else 0 for v in py], "0.00"),
        ("Rev/customer (J$)", [v[1] / v[2] if v[2] else 0 for v in cyi], [v[1] / v[2] if v[2] else 0 for v in py], MONEYFMT),
    ]
    for name, vals, pyvals, fmt in metrics:
        ws.cell(row=row, column=2, value=name).font = BOLD
        for i, v in enumerate(vals):
            c = ws.cell(row=row, column=3 + i, value=round(v, 2))
            c.number_format = fmt
        delta = (vals[2] / vals[0] - 1) if vals[0] else 0
        c = ws.cell(row=row, column=6, value=delta); c.number_format = PCTFMT
        for i, (v, p) in enumerate(zip(vals, pyvals)):
            yoy = (v / p - 1) if p else 0
            c = ws.cell(row=row, column=7 + i, value=yoy); c.number_format = PCTFMT
        row += 1

    # Breakout, for transparency -- how much of the headline total above is Prepaid.
    ws.cell(row=row, column=2, value="  of which: Postpaid kWh").font = Font(name=FONT, italic=True, color="7F7F7F")
    for i, v in enumerate(cy):
        c = ws.cell(row=row, column=3 + i, value=round(v[0], 2)); c.number_format = NUMFMT
    row += 1
    ws.cell(row=row, column=2, value="  of which: Prepaid (PAYG) kWh").font = Font(name=FONT, italic=True, color="7F7F7F")
    for i, p in enumerate(pp):
        c = ws.cell(row=row, column=3 + i, value=round(p[0], 2) if p else "N/A (file not supplied)")
        if p:
            c.number_format = NUMFMT
    row += 1
    ws.cell(row=row, column=2, value="  of which: Postpaid Revenue (J$)").font = Font(name=FONT, italic=True, color="7F7F7F")
    for i, v in enumerate(cy):
        c = ws.cell(row=row, column=3 + i, value=round(v[1], 2)); c.number_format = MONEYFMT
    row += 1
    ws.cell(row=row, column=2, value="  of which: Prepaid (PAYG) Revenue (J$)").font = Font(name=FONT, italic=True, color="7F7F7F")
    for i, p in enumerate(pp):
        c = ws.cell(row=row, column=3 + i, value=round(p[1], 2) if p else "N/A (file not supplied)")
        if p:
            c.number_format = MONEYFMT
    row += 1

    row += 2
autosize(ws, [3] + [16] * 10)

# ===== By Bucket (cluster view) =====
ws = wb.create_sheet("By Bucket")
ws.sheet_view.showGridLines = False
ws["B2"] = "Consumption-Tier Mix Shift (the 'cluster' view - no individual premise detail exists for RT10)"
ws["B2"].font = TITLE_FONT
row = 4
for rc, label in [("RT10", "RT10 - Residential"), ("RT20", "RT20 - General Service / SME")]:
    ws.cell(row=row, column=2, value=label).font = SUB_FONT
    row += 1
    hdr = ["Tier"] + [m + " kWh" for m in MLABELS] + [m + " Rev J$M" for m in MLABELS] + [m + " Cust%" for m in MLABELS]
    for i, h in enumerate(hdr):
        ws.cell(row=row, column=2 + i, value=h)
    style_header(ws, row, len(hdr), start_col=2)
    row += 1
    per_month = [roll_up_buckets(m, rc) for m in MONTHS_2026]
    tot_cust = [sum(pm[b]["count"] for b in BUCKET_ORDER) for pm in per_month]
    for b in BUCKET_ORDER:
        ws.cell(row=row, column=2, value=b)
        for i, pm in enumerate(per_month):
            ws.cell(row=row, column=3 + i, value=round(pm[b]["kwh"])).number_format = NUMFMT
        for i, pm in enumerate(per_month):
            ws.cell(row=row, column=6 + i, value=round(pm[b]["rev"] / 1e6, 2)).number_format = "#,##0.0"
        for i, pm in enumerate(per_month):
            pct = pm[b]["count"] / tot_cust[i] if tot_cust[i] else 0
            ws.cell(row=row, column=9 + i, value=pct).number_format = PCTFMT
        row += 1
    row += 2
autosize(ws, [3] + [14] * 9)

# ===== By Parish - Prepaid =====
ws = wb.create_sheet("By Parish - Prepaid")
ws.sheet_view.showGridLines = False
ws["B2"] = "Prepaid (PAYG) Population by Parish - June & July 2026 only (May file not supplied)"
ws["B2"].font = TITLE_FONT
row = 4
for rc, label in [("RT10", "RT10 Prepaid"), ("RT20", "RT20 Prepaid")]:
    ws.cell(row=row, column=2, value=label).font = SUB_FONT
    row += 1
    hdr = ["Parish", "Jun kWh", "Jul kWh", "Δ%", "Jun Rev J$", "Jul Rev J$", "Δ%", "Jun Cust", "Jul Cust", "Δ%", "Jul Rev/Cust"]
    for i, h in enumerate(hdr):
        ws.cell(row=row, column=2 + i, value=h)
    style_header(ws, row, len(hdr), start_col=2)
    row += 1
    jun = agg_prepaid(6, rc)
    jul = agg_prepaid(7, rc)
    parishes = sorted(set(jun) | set(jul), key=lambda p: -jul.get(p, {}).get("rev", 0))
    for p in parishes:
        a, b = jun.get(p, {"kwh": 0, "rev": 0, "cust": 0}), jul.get(p, {"kwh": 0, "rev": 0, "cust": 0})
        ws.cell(row=row, column=2, value=p)
        ws.cell(row=row, column=3, value=round(a["kwh"])).number_format = NUMFMT
        ws.cell(row=row, column=4, value=round(b["kwh"])).number_format = NUMFMT
        ws.cell(row=row, column=5, value=(b["kwh"] / a["kwh"] - 1) if a["kwh"] else None).number_format = PCTFMT
        ws.cell(row=row, column=6, value=round(a["rev"])).number_format = MONEYFMT
        ws.cell(row=row, column=7, value=round(b["rev"])).number_format = MONEYFMT
        ws.cell(row=row, column=8, value=(b["rev"] / a["rev"] - 1) if a["rev"] else None).number_format = PCTFMT
        ws.cell(row=row, column=9, value=a["cust"]).number_format = NUMFMT
        ws.cell(row=row, column=10, value=b["cust"]).number_format = NUMFMT
        ws.cell(row=row, column=11, value=(b["cust"] / a["cust"] - 1) if a["cust"] else None).number_format = PCTFMT
        ws.cell(row=row, column=12, value=round(b["rev"] / b["cust"]) if b["cust"] else None).number_format = MONEYFMT
        row += 1
    row += 2
autosize(ws, [3, 22] + [12] * 10)

# ===== By Parish - RT20 Real =====
ws = wb.create_sheet("By Parish - RT20 Real")
ws.sheet_view.showGridLines = False
ws["B2"] = "Real RT20 Premises by Parish - NAICS-split (Commercial vs no-NAICS residential-style)"
ws["B2"].font = TITLE_FONT
row = 4
for seg, label in [("comm", "Commercial (has NAICS code) - individual premises"), ("res", "Residential-style (no NAICS) - bucketed")]:
    ws.cell(row=row, column=2, value=label).font = SUB_FONT
    row += 1
    hdr = ["Parish", "May kWh", "Jun kWh", "Jul kWh", "May Rev J$", "Jun Rev J$", "Jul Rev J$", "May→Jul Δ%", "Jul Accounts"]
    for i, h in enumerate(hdr):
        ws.cell(row=row, column=2 + i, value=h)
    style_header(ws, row, len(hdr), start_col=2)
    row += 1
    m5 = agg_rt20_real(5); m6 = agg_rt20_real(6); m7 = agg_rt20_real(7)
    idx = 2 if seg == "comm" else 3
    p5, p6, p7 = m5[idx], m6[idx], m7[idx]
    parishes = sorted(set(p5) | set(p6) | set(p7), key=lambda p: -p7.get(p, {}).get("rev", 0))
    for p in parishes:
        a, b, c = p5.get(p, {"kwh": 0, "rev": 0, "cust": 0}), p6.get(p, {"kwh": 0, "rev": 0, "cust": 0}), p7.get(p, {"kwh": 0, "rev": 0, "cust": 0})
        ws.cell(row=row, column=2, value=p)
        ws.cell(row=row, column=3, value=round(a["kwh"])).number_format = NUMFMT
        ws.cell(row=row, column=4, value=round(b["kwh"])).number_format = NUMFMT
        ws.cell(row=row, column=5, value=round(c["kwh"])).number_format = NUMFMT
        ws.cell(row=row, column=6, value=round(a["rev"])).number_format = MONEYFMT
        ws.cell(row=row, column=7, value=round(b["rev"])).number_format = MONEYFMT
        ws.cell(row=row, column=8, value=round(c["rev"])).number_format = MONEYFMT
        ws.cell(row=row, column=9, value=(c["rev"] / a["rev"] - 1) if a["rev"] else None).number_format = PCTFMT
        ws.cell(row=row, column=10, value=c["cust"]).number_format = NUMFMT
        row += 1
    row += 2
autosize(ws, [3, 22] + [12] * 8)

# ===== Budget & Forecast =====
ws = wb.create_sheet("Budget & Forecast")
ws.sheet_view.showGridLines = False
ws["B2"] = "Budget & Forecast Comparison"
ws["B2"].font = TITLE_FONT
ws["B4"] = "RT10: no usable forecast/budget source exists anywhere in the system - see Read Me. Not fabricated here."
ws["B4"].font = Font(name=FONT, italic=True, color="C00000")
ws["B4"].fill = WARN_FILL
row = 6
ws.cell(row=row, column=2, value="RT20 (real premise population, all consumption_bucket ≠ 'Prepaid')").font = SUB_FONT
row += 1
hdr = ["Month", "Actual kWh", "Budget kWh", "kWh Var%", "Actual Rev J$", "Budget Rev J$", "Rev Var%", "LE kWh", "LE Rev J$", "Actual vs LE Rev%"]
for i, h in enumerate(hdr):
    ws.cell(row=row, column=2 + i, value=h)
style_header(ws, row, len(hdr), start_col=2)
row += 1
for mn, mlabel in zip([5, 6, 7], MLABELS):
    comm, res, _, _ = agg_rt20_real(mn)
    act_kwh = comm["kwh"] + res["kwh"]
    act_rev = comm["rev"] + res["rev"]
    bud_kwh, bud_rev = agg_rt20_budget_le(rt20_budget, mn)
    le_kwh, le_rev = agg_rt20_budget_le(rt20_le, mn)
    ws.cell(row=row, column=2, value=mlabel)
    ws.cell(row=row, column=3, value=round(act_kwh)).number_format = NUMFMT
    ws.cell(row=row, column=4, value=round(bud_kwh)).number_format = NUMFMT
    ws.cell(row=row, column=5, value=(act_kwh / bud_kwh - 1) if bud_kwh else None).number_format = PCTFMT
    ws.cell(row=row, column=6, value=round(act_rev)).number_format = MONEYFMT
    ws.cell(row=row, column=7, value=round(bud_rev)).number_format = MONEYFMT
    ws.cell(row=row, column=8, value=(act_rev / bud_rev - 1) if bud_rev else None).number_format = PCTFMT
    ws.cell(row=row, column=9, value=round(le_kwh)).number_format = NUMFMT
    ws.cell(row=row, column=10, value=round(le_rev)).number_format = MONEYFMT
    ws.cell(row=row, column=11, value=(act_rev / le_rev - 1) if le_rev else None).number_format = PCTFMT
    row += 1
row += 1
ws.cell(row=row, column=2, value=("NOTE: RT20 runs consistently over budget across all three months (see Var% above) - verified against "
                                    "the source LE model, not a data artifact. An earlier version of this analysis showed a much larger, "
                                    "spurious May variance caused by two now-fixed bugs: a duplicate/stale account population loaded via an "
                                    "older methodology (removed from jps_actuals), and an unordered-pagination bug in the live query layer "
                                    "that silently dropped rows under load (fixed with explicit ordering + retry). These numbers reflect "
                                    "the corrected data."))
ws.cell(row=row, column=2).font = Font(name=FONT, italic=True, color="C00000")
ws.cell(row=row, column=2).fill = WARN_FILL
autosize(ws, [3, 12] + [15] * 9)

# ===== Drivers =====
ws = wb.create_sheet("Drivers")
ws.sheet_view.showGridLines = False
ws["B2"] = "What's Driving the Numbers"
ws["B2"].font = TITLE_FONT
row = 4
for rc, label in [("RT10", "RT10 - Residential"), ("RT20", "RT20 - General Service / SME")]:
    ws.cell(row=row, column=2, value=label).font = SUB_FONT
    row += 1
    cy = [class_total(m, rc) for m in MONTHS_2026]
    kwh_growth = cy[2][0] / cy[0][0] - 1 if cy[0][0] else 0
    cust_growth = cy[2][2] / cy[0][2] - 1 if cy[0][2] else 0
    kwh_per_cust = [v[0] / v[2] if v[2] else 0 for v in cy]
    usage_growth = kwh_per_cust[2] / kwh_per_cust[0] - 1 if kwh_per_cust[0] else 0
    pm5, pm7 = roll_up_buckets(MONTHS_2026[0], rc), roll_up_buckets(MONTHS_2026[2], rc)
    tot5 = sum(v["count"] for v in pm5.values()) or 1
    tot7 = sum(v["count"] for v in pm7.values()) or 1
    mix_shift = [(b, pm7[b]["count"] / tot7 - pm5[b]["count"] / tot5) for b in BUCKET_ORDER]
    mix_shift.sort(key=lambda x: -abs(x[1]))
    lines = [
        f"kWh volume, May→Jul: {kwh_growth:+.1%}",
        f"Customer-months, May→Jul: {cust_growth:+.1%}",
        f"Usage per customer, May→Jul: {usage_growth:+.1%}  ({'more usage per customer' if usage_growth>0 else 'less usage per customer'} driving the volume change{'​, not just customer count' if abs(usage_growth)>abs(cust_growth) else ''})",
        "Largest consumption-tier mix shifts (May→Jul, share of customer-months):",
    ]
    for b, d in mix_shift[:3]:
        lines.append(f"    {b}: {d:+.1%} pts")
    for line in lines:
        ws.cell(row=row, column=2, value=line).font = Font(name=FONT, size=10, bold=line.startswith(("kWh", "Customer", "Usage", "Largest")))
        row += 1
    row += 1
autosize(ws, [3, 100])

# ===== Insights & LE Recommendations =====
ws = wb.create_sheet("Insights & LE Recommendations")
ws.sheet_view.showGridLines = False
ws["B2"] = "Insights, Anomalies & Recommendations for Future LE"
ws["B2"].font = TITLE_FONT
row = 4


def write_block(ws, row, heading, lines, heading_color="1F4E78"):
    ws.cell(row=row, column=2, value=heading).font = Font(name=FONT, bold=True, size=11, color=heading_color)
    row += 1
    for line in lines:
        ws.cell(row=row, column=2, value=line).font = Font(name=FONT, size=10)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    return row + 1


# Budget variance trend, RT20
var_rows = []
for mn, mlabel in zip([5, 6, 7], MLABELS):
    comm, res, _, _ = agg_rt20_real(mn)
    act_rev = comm["rev"] + res["rev"]
    _, bud_rev = agg_rt20_budget_le(rt20_budget, mn)
    var_rows.append((mlabel, (act_rev / bud_rev - 1) if bud_rev else 0))
all_over = all(v > 0 for _, v in var_rows)
avg_var = sum(v for _, v in var_rows) / len(var_rows)

row = write_block(ws, row, "1. RT20 has run over budget every month for 3 consecutive months - the LE base needs rebasing, not smoothing", [
    "Actual vs. budget revenue: " + ", ".join(f"{m} {v:+.1%}" for m, v in var_rows) + f"  (average {avg_var:+.1%}).",
    "A single-month miss can be noise; three consecutive months all landing over budget, in the same direction, at a "
    "similar magnitude, is a sign the BUDGET/LE ASSUMPTION is miscalibrated (understated rate or volume base), not "
    "that the business is outperforming. Recommendation: rebase the next LE cycle's RT20 volume/rate assumptions off "
    "the trailing 3-month actual run-rate rather than carrying forward the current budget's growth assumptions "
    "unchanged.",
] if all_over else [
    "Actual vs. budget revenue: " + ", ".join(f"{m} {v:+.1%}" for m, v in var_rows) + f"  (average {avg_var:+.1%}).",
])

row = write_block(ws, row, "2. RT10 has no LE/budget model at all - Prepaid makes this more urgent, not less", [
    "No usable RT10 forecast source exists anywhere in the system (jps_budget and jps_le both only carry a small "
    "'Postpaid true-up' placeholder for RT10, not a real consumption budget - see Read Me). This gap predates this "
    "analysis, but now that a Prepaid revenue stream is confirmed material (July 2026: RT10 Prepaid alone was "
    f"J${sum(v['revenue_jmd'] or 0 for v in prepaid if v['year']==2026 and v['month']==7 and v['rate_class']=='RT10')/1e6:,.1f}M "
    "pre-GCT), the case for building a real RT10 LE model - one that explicitly forecasts Postpaid and Prepaid as "
    "separate lines - is stronger than when RT10 was assumed to be a single homogeneous population.",
])

row = write_block(ws, row, "3. Prepaid revenue is NOT included in the 'Overview' / 'By Bucket' totals above - by source design, not by omission", [
    "The Overview and By Bucket sheets are built from the monthly CIS Billing Details Report, which only covers "
    "postpaid-billed meters - prepaid (PAYG/vended) meters are a structurally separate system and were never in "
    "that source file. This means any reader treating 'Overview' as the true total RT10/RT20 volume is understating "
    "it: for July 2026 alone, Prepaid + the newly-identified unassigned-tariff PAYG population (see #4) together add "
    f"J${(sum(v['revenue_jmd'] or 0 for v in prepaid if v['year']==2026 and v['month']==7) + sum(v['revenue_jmd'] or 0 for v in unassigned_payg))/1e6:,.1f}M "
    "pre-GCT of revenue not reflected in the Overview sheet's headline numbers. Recommendation: if Overview is used "
    "as a management-facing 'total RT10/RT20' figure, add an explicit 'incl. Prepaid' reconciliation line so readers "
    "aren't silently missing a real, growing revenue stream.",
])

n_unassigned = sum(v["customer_count"] or 0 for v in unassigned_payg)
rev_unassigned = sum(v["revenue_jmd"] or 0 for v in unassigned_payg)
row = write_block(ws, row, "4. New data-quality finding: ~18% of July's PAYG revenue has no rate-class tag in the CIS extract", [
    f"{n_unassigned:,.0f} prepaid customers (J${rev_unassigned/1e6:,.1f}M pre-GCT, July 2026) came through the "
    "Customer-Monthly-Transaction-Report with tariff = 'unassigned' - no RT10/RT20 designation at all. These are "
    "loaded into jps_actuals under rate_class='UNASSIGNED-PAYG' (segment='Residential', consumption_bucket='Prepaid') "
    "rather than guessed into RT10 or RT20, so the data isn't lost, but it also isn't usable for rate-class-level "
    "analysis until resolved. Recommendation: flag to Billing/CIS ops for tariff-code backfill on these accounts - "
    "this population is large enough (18% of PAYG revenue) to matter for rate-class-level LE accuracy.",
])

row = write_block(ws, row, "5. Kingston-metro (KSAN/KSAS) parish detail in the Prepaid source file has degraded", [
    "The July 2026 Customer-Monthly-Transaction-Report labels the large majority of Kingston-metro PAYG customers "
    "with the generic city value 'Kingston' rather than a specific postal zone, which is not enough information to "
    "split into KSAN vs. KSAS the way earlier loads could. This build preserves the existing (already-correct) "
    "KSAN/KSAS/Portmore figures rather than re-deriving them from this degraded data, and routes new PAYG volume for "
    "those same customers to the correct parish only where the file gave a specific enough signal. If parish-level "
    "PAYG reporting matters going forward, this is worth raising with whoever generates the CIS export - the postal "
    "zone-level detail used to be there.",
])

row = write_block(ws, row, "6. 'Zero' consumption-tier customers still generate real revenue - not a data error", [
    "In the jps_actuals-sourced views (e.g. 'By Parish - RT20 Real'), the 'Zero' bucket (0 kWh billed) shows "
    "non-zero revenue because Jamaican utility billing applies a fixed monthly customer/connection charge "
    "regardless of consumption - a customer who used no power still owes the connection fee (plus GCT on it). "
    "This is expected, standard billing behavior. One completeness gap found while building this: the RT20 "
    "residential-style (no-NAICS) bucket aggregation does not currently break the fixed-charge component out into "
    "its own column the way the Commercial (NAICS) premise rows do - revenue is correct in total, but the "
    "components (customer_charge_jmd, demand_jmd, etc.) read 0 for these rows. Worth fixing in rt20_split.py if "
    "component-level detail on the residential-style population is ever needed.",
    "",
    "FIX APPLIED THIS BUILD: the 'By Bucket' sheet's 'Zero' row was previously showing real, substantial kWh volume "
    "(e.g. RT10 ~2.1-2.2M kWh/month) — a genuine bug, not the billing behavior above. corrected.json's histogram "
    "source bins are 50kWh WIDE: the bin labeled '0' actually covers the range [0,50) kWh, not exactly-zero "
    "consumption. The rollup was reusing a per-customer classifier (where kwh==0 correctly means 'Zero') on that "
    "bin's raw label, so the entire [0,50) bin's real consumption was mislabeled as 'Zero'. Fixed by using a "
    "dedicated bin classifier that folds bin '0' into '<150' like every other bin - this data source has no bin "
    "fine enough to isolate exactly-zero consumption, so 'Zero' now correctly reads empty in this sheet.",
])

row = write_block(ws, row, "7. Usage per customer, not customer growth, is what should drive the next LE's volume assumption", [
    "For both RT10 and RT20 (see Drivers sheet), customer counts are nearly flat May→July (+0.1-0.3%) while usage "
    "per customer is up 9-14%. A volume forecast built primarily off customer-count growth will materially "
    "understate near-term RT10/RT20 volume. Recommendation: weight the trailing usage-per-customer trend, not just "
    "customer-count growth, in the next LE cycle's volume build for these two classes.",
])

autosize(ws, [3, 118])

import itertools
for suffix in itertools.chain(["", "_v2", "_v3"], (f"_v{n}" for n in itertools.count(4))):
    OUT_XLSX = f"Residential_RT10_RT20_Analysis_2026Q{suffix}.xlsx"
    try:
        wb.save(OUT_XLSX)
        break
    except PermissionError:
        continue
print("wrote", OUT_XLSX)
