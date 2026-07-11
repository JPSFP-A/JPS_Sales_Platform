import csv, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
ITALIC_RED = Font(name=FONT, italic=True, color="C00000")
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBHDR_FONT = Font(name=FONT, bold=True, color="1F4E78")
SUBHDR_FILL = PatternFill("solid", fgColor="D9E2F3")
YELLOW = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="B7C6D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUMFMT = "#,##0;(#,##0);\"-\""
PCTFMT = "0.0%"
RATIOFMT = "0.000"
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

SRC_COMM = r"D:\Projects\Sales_Platform\analysis\commercial_actuals_2024_2026.csv"
SRC_RES = r"D:\Projects\Sales_Platform\analysis\residential_actuals_2024_2026.csv"
OUT = r"D:\Projects\Sales_Platform\analysis\Commercial_RateClass_Forecast_2026.xlsx"

COMMERCIAL_RCS = [("RT40", "LV Commercial"), ("RT50", "MV Commercial"),
                  ("RT60-ST", "Street Lights"), ("RT70", "Large Industrial")]
RESIDENTIAL_RCS = [("RT10", "Residential"), ("RT20", "SME / Gen. Service")]
ALL_RCS = RESIDENTIAL_RCS + COMMERCIAL_RCS


def is_individual(ac):
    return bool(re.match(r"^[0-9]", ac))


def load_commercial():
    rows = list(csv.DictReader(open(SRC_COMM, encoding="utf-8")))
    indiv, bucket = {}, {}
    for r in rows:
        ac, rc, parish = r["jps_ac"], r["rate_class"], r["parish"]
        if rc not in dict(COMMERCIAL_RCS):
            continue
        y, m = int(r["year"]), int(r["month"])
        kwh = float(r["kwh"] or 0)
        if is_individual(ac):
            key = (ac, rc)
            d = indiv
        else:
            key = (ac, rc, parish)
            d = bucket
        if key not in d:
            d[key] = {"name": r["name"], "parish": parish, "kwh": {}}
        d[key]["kwh"][(y, m)] = kwh
        if r["name"]:
            d[key]["name"] = r["name"]
    return indiv, bucket


def load_residential():
    if not os.path.exists(SRC_RES):
        return {}
    rows = list(csv.DictReader(open(SRC_RES, encoding="utf-8")))
    bucket = {}
    for r in rows:
        rc = r["rate_class"]
        if rc not in dict(RESIDENTIAL_RCS):
            continue
        ac, parish, cb = r["jps_ac"], r["parish"], r.get("consumption_bucket", "")
        y, m = int(r["year"]), int(r["month"])
        kwh = float(r["kwh"] or 0)
        key = (ac, rc, parish, cb)
        if key not in bucket:
            bucket[key] = {"name": r["name"], "parish": parish, "bucket": cb, "kwh": {}}
        bucket[key]["kwh"][(y, m)] = kwh
    return bucket


def load_kam():
    path = r"D:\Projects\Sales_Platform\analysis\kam_assignments.csv"
    if not os.path.exists(path):
        return {}
    rows = list(csv.DictReader(open(path, encoding="utf-8")))
    return {r["jps_ac"]: r["kam"] for r in rows}


indiv_all, bucket_all = load_commercial()
res_bucket_all = load_residential()
kam_by_ac = load_kam()

wb = Workbook()
wb.remove(wb.active)


def style_header(ws, row, headers, col_start=1):
    for i, h in enumerate(headers, col_start):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 30


def add_drivers_block(ws, title):
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws["A2"] = "Drivers (edit blue cells)"
    ws["A2"].font = Font(name=FONT, bold=True, italic=True)
    headers = [f"Seasonal Ratio {m} (calc, vs 2025 Apr-Jun avg)" for m in MONTHS[6:12]]
    headers += ["Growth Adjustment % (toggle)", "Customer Count Growth % (toggle)", "Other Monthly Adjustment (kWh, toggle)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = SUBHDR_FONT
        c.fill = SUBHDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BORDER
    ws.row_dimensions[3].height = 44
    for i in range(1, 7):
        c = ws.cell(row=4, column=i, value=0)
        c.font = BLACK
        c.number_format = RATIOFMT
        c.border = BORDER
    g = ws.cell(row=4, column=7, value=0.0); g.font = BLUE; g.number_format = PCTFMT; g.fill = YELLOW; g.border = BORDER
    h = ws.cell(row=4, column=8, value=0.0); h.font = BLUE; h.number_format = PCTFMT; h.fill = YELLOW; h.border = BORDER
    o = ws.cell(row=4, column=9, value=0); o.font = BLUE; o.number_format = NUMFMT; o.fill = YELLOW; o.border = BORDER
    for i in range(1, 10):
        ws.column_dimensions[get_column_letter(i)].width = 20


def write_customer_table(ws, start_row, keys_data, use_kam_col=False, use_bucket_cols=False, show_rc_col=False):
    header_row = start_row
    cols = ["Account #" if not use_bucket_cols else "Bucket Label", "Name"]
    if use_kam_col:
        cols.append("KAM")
    if use_bucket_cols:
        cols.append("Consumption Bucket")
    if show_rc_col:
        cols.append("Rate Class")
    cols.append("Parish")
    id_cols = len(cols)
    cols += [f"2026 {m}" for m in MONTHS[:6]]
    cols += [f"2025 {m}" for m in MONTHS[3:6]]
    cols += [f"2025 {m}" for m in MONTHS[6:12]]
    cols += [f"2026 {m} (proj)" for m in MONTHS[6:12]]
    cols += ["FY2026 Total"]
    style_header(ws, header_row, cols)

    COL_JANJUN = list(range(id_cols + 1, id_cols + 7))
    COL_AMJ25 = list(range(id_cols + 7, id_cols + 10))
    COL_JD25 = list(range(id_cols + 10, id_cols + 16))
    COL_PROJ = list(range(id_cols + 16, id_cols + 22))
    COL_TOTAL = id_cols + 22

    r = header_row + 1
    for key, d in keys_data:
        col = 1
        ws.cell(row=r, column=col, value=key[0]).font = BLACK; col += 1
        ws.cell(row=r, column=col, value=d.get("name", "")).font = BLACK; col += 1
        if use_kam_col:
            ws.cell(row=r, column=col, value=kam_by_ac.get(key[0].split("-")[0], "-")).font = BLACK; col += 1
        if use_bucket_cols:
            ws.cell(row=r, column=col, value=d.get("bucket", "")).font = BLACK; col += 1
        if show_rc_col:
            ws.cell(row=r, column=col, value=key[1]).font = BLACK; col += 1
        ws.cell(row=r, column=col, value=d.get("parish", "")).font = BLACK; col += 1

        for j, m in enumerate(range(1, 7)):
            v = d["kwh"].get((2026, m), 0)
            c = ws.cell(row=r, column=COL_JANJUN[j], value=round(v)); c.font = BLACK; c.number_format = NUMFMT; c.border = BORDER
        for j, m in enumerate(range(4, 7)):
            v = d["kwh"].get((2025, m), 0)
            c = ws.cell(row=r, column=COL_AMJ25[j], value=round(v)); c.font = BLACK; c.number_format = NUMFMT
        for j, m in enumerate(range(7, 13)):
            v = d["kwh"].get((2025, m), 0)
            c = ws.cell(row=r, column=COL_JD25[j], value=round(v)); c.font = BLACK; c.number_format = NUMFMT
        r += 1
    last_row = r - 1
    return {"header_row": header_row, "first_row": header_row + 1, "last_row": last_row,
            "COL_JANJUN": COL_JANJUN, "COL_AMJ25": COL_AMJ25, "COL_JD25": COL_JD25,
            "COL_PROJ": COL_PROJ, "COL_TOTAL": COL_TOTAL, "id_cols": id_cols,
            "parish_col": (4 if show_rc_col and use_bucket_cols else (3 if (show_rc_col or use_bucket_cols or use_kam_col) else 3))}


def fill_projection_formulas(ws, tbl):
    if tbl["first_row"] > tbl["last_row"]:
        return
    apr_col = get_column_letter(tbl["COL_JANJUN"][3])
    jun_col = get_column_letter(tbl["COL_JANJUN"][5])
    for r in range(tbl["first_row"], tbl["last_row"] + 1):
        for j in range(6):
            proj_col = tbl["COL_PROJ"][j]
            ratio_ref = f"${get_column_letter(j+1)}$4"
            f = (f"=AVERAGE({apr_col}{r}:{jun_col}{r})*{ratio_ref}*(1+$G$4)"
                 f"+$I$4*IFERROR({jun_col}{r}/SUM(${jun_col}${tbl['first_row']}:${jun_col}${tbl['last_row']}),0)")
            c = ws.cell(row=r, column=proj_col, value=f); c.font = BLACK; c.number_format = NUMFMT; c.border = BORDER
        fa = get_column_letter(tbl["COL_JANJUN"][0])
        la = get_column_letter(tbl["COL_JANJUN"][-1])
        fp = get_column_letter(tbl["COL_PROJ"][0])
        lp = get_column_letter(tbl["COL_PROJ"][-1])
        tot = ws.cell(row=r, column=tbl["COL_TOTAL"], value=f"=SUM({fa}{r}:{la}{r})+SUM({fp}{r}:{lp}{r})")
        tot.font = BOLD; tot.number_format = NUMFMT; tot.border = BORDER


def fill_seasonal_ratios(ws, table_ranges):
    amj_parts = []
    for tbl in table_ranges:
        if tbl["first_row"] > tbl["last_row"]:
            continue
        for j in range(3):
            col = get_column_letter(tbl["COL_AMJ25"][j])
            amj_parts.append(f"SUM({col}{tbl['first_row']}:{col}{tbl['last_row']})")
    if not amj_parts:
        for i in range(1, 7):
            ws.cell(row=4, column=i, value=0)
        return
    amj_total_expr = "(" + "+".join(amj_parts) + ")/3"
    for j in range(6):
        month_parts = []
        for tbl in table_ranges:
            if tbl["first_row"] > tbl["last_row"]:
                continue
            col = get_column_letter(tbl["COL_JD25"][j])
            month_parts.append(f"SUM({col}{tbl['first_row']}:{col}{tbl['last_row']})")
        month_expr = "+".join(month_parts) if month_parts else "0"
        f = f"=IFERROR(({month_expr})/({amj_total_expr}),0)"
        c = ws.cell(row=4, column=j + 1, value=f)
        c.font = BLACK
        c.number_format = RATIOFMT


sheet_meta = {}

for rc, desc in ALL_RCS:
    title = rc.replace("-", "_")[:31]
    ws = wb.create_sheet(title)
    add_drivers_block(ws, f"{rc} - {desc}")

    tables = []
    row_cursor = 6
    if rc in dict(COMMERCIAL_RCS):
        indiv_keys = sorted(k for k in indiv_all if k[1] == rc)
        indiv_data = [(k, indiv_all[k]) for k in indiv_keys]
        ws.cell(row=row_cursor - 1, column=1, value=f"Individual Commercial Customers ({len(indiv_data)} accounts)").font = BOLD
        tbl1 = write_customer_table(ws, row_cursor, indiv_data, use_kam_col=True)
        tables.append(tbl1)
        row_cursor = tbl1["last_row"] + 3
    else:
        bkeys = sorted(k for k in res_bucket_all if k[1] == rc)
        bdata = [(k, res_bucket_all[k]) for k in bkeys]
        ws.cell(row=row_cursor - 1, column=1,
                value=f"Parish x Consumption-Bucket rows ({len(bdata)}) - residential classes have no individual account detail in jps_actuals").font = BOLD
        tbl1 = write_customer_table(ws, row_cursor, bdata, use_bucket_cols=True)
        tables.append(tbl1)
        row_cursor = tbl1["last_row"] + 3

    for tbl in tables:
        fill_projection_formulas(ws, tbl)
    fill_seasonal_ratios(ws, tables)

    for i in range(1, 30):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 26
    ws.freeze_panes = "A6"

    sheet_meta[rc] = {"title": title, "tables": tables}

# ── UNIDENTIFIED / AGGREGATE ROWS (ambiguous bucket rows found inside commercial classes) ──
ws = wb.create_sheet("Unidentified Rows")
add_drivers_block(ws, "Unidentified / Aggregate Rows (ambiguous - see Read Me)")
bkeys = sorted(bucket_all.keys())
bdata = [(k, bucket_all[k]) for k in bkeys]
ws.cell(row=5, column=1,
        value=f"{len(bdata)} rows found tagged as RT40/RT50 in jps_actuals but not individually-identified customers (labels like Postpaid/Zero/Zero-Postpaid). "
              f"Small per-parish customer_count (single digits to low 20s) - NOT the same as the confirmed June RT20 mass-mistagging. "
              f"Nature unconfirmed: could be real small aggregated postpaid commercial accounts, or misclassified residential. "
              f"Excluded from every rate class tab and from the Total tab until confirmed.").font = ITALIC_RED
ws.row_dimensions[5].height = 45
tbl_u = write_customer_table(ws, 7, bdata, use_bucket_cols=False, show_rc_col=True)
fill_projection_formulas(ws, tbl_u)
fill_seasonal_ratios(ws, [tbl_u])
for i in range(1, 30):
    ws.column_dimensions[get_column_letter(i)].width = 13
ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 20
ws.freeze_panes = "A7"

LAST_ROW = {rc: sheet_meta[rc]["tables"][0]["last_row"] for rc in sheet_meta}

# ── TOTAL SHEET ──────────────────────────────────────────────────────
tot = wb.create_sheet("Total", 0)
tot.column_dimensions["A"].width = 26
for i in range(2, 15):
    tot.column_dimensions[get_column_letter(i)].width = 12
style_header(tot, 1, ["Rate Class"] + [f"2026 {m}" for m in MONTHS])

row = 2
for rc, desc in ALL_RCS:
    tot.cell(row=row, column=1, value=f"{rc} - {desc}").font = BOLD
    meta = sheet_meta[rc]
    sheet_title = meta["title"]
    for j in range(6):
        parts = []
        for tbl in meta["tables"]:
            if tbl["first_row"] > tbl["last_row"]:
                continue
            col = get_column_letter(tbl["COL_JANJUN"][j])
            parts.append(f"SUM('{sheet_title}'!{col}{tbl['first_row']}:{col}{tbl['last_row']})")
        f = "=" + "+".join(parts) if parts else "=0"
        c = tot.cell(row=row, column=2 + j, value=f); c.font = GREEN; c.number_format = NUMFMT; c.border = BORDER
    for j in range(6):
        parts = []
        for tbl in meta["tables"]:
            if tbl["first_row"] > tbl["last_row"]:
                continue
            col = get_column_letter(tbl["COL_PROJ"][j])
            parts.append(f"SUM('{sheet_title}'!{col}{tbl['first_row']}:{col}{tbl['last_row']})")
        f = "=" + "+".join(parts) if parts else "=0"
        c = tot.cell(row=row, column=8 + j, value=f); c.font = GREEN; c.number_format = NUMFMT; c.border = BORDER
    row += 1

cust_growth_row = row + 1
tot.cell(row=cust_growth_row, column=1, value="Customer Count Growth Add-On (Jul-Dec, kWh)").font = Font(name=FONT, italic=True)
for j in range(6):
    parts = []
    for rc, desc in ALL_RCS:
        meta = sheet_meta[rc]
        sheet_title = meta["title"]
        for tbl in meta["tables"]:
            if tbl["first_row"] > tbl["last_row"]:
                continue
            jun_col = get_column_letter(tbl["COL_JANJUN"][5])
            n = tbl["last_row"] - tbl["first_row"] + 1
            parts.append(
                f"'{sheet_title}'!$H$4*{n}*IFERROR(SUM('{sheet_title}'!{jun_col}{tbl['first_row']}:{jun_col}{tbl['last_row']})/{n},0)"
            )
    f = "=" + "+".join(parts) if parts else "=0"
    c = tot.cell(row=cust_growth_row, column=8 + j, value=f); c.font = BLACK; c.number_format = NUMFMT

total_row = cust_growth_row + 1
tot.cell(row=total_row, column=1, value="TOTAL (incl. Count Growth)").font = BOLD
for j in range(6):
    col = get_column_letter(2 + j)
    c = tot.cell(row=total_row, column=2 + j, value=f"=SUM({col}2:{col}{cust_growth_row-2})")
    c.font = BOLD; c.number_format = NUMFMT
for j in range(6):
    col = get_column_letter(8 + j)
    c = tot.cell(row=total_row, column=8 + j, value=f"=SUM({col}2:{col}{cust_growth_row})")
    c.font = BOLD; c.number_format = NUMFMT

note_row = total_row + 2
tot.cell(row=note_row, column=1, value="Note: 'Unidentified Rows' tab (ambiguous Postpaid/Zero rows found in RT40/RT50) is intentionally EXCLUDED from this total. See Read Me.").font = ITALIC_RED

# ── BY PARISH ────────────────────────────────────────────────────────
parish_start = note_row + 2
tot.cell(row=parish_start, column=1, value="BY PARISH (all rate classes, Jan-Jun actual + Jul-Dec projected)").font = Font(name=FONT, bold=True, size=12)
hdr_row = parish_start + 1
style_header(tot, hdr_row, ["Parish"] + [f"2026 {m}" for m in MONTHS])

all_parishes = sorted(set(d["parish"] for k, d in indiv_all.items() if d.get("parish")) |
                       set(d["parish"] for k, d in res_bucket_all.items() if d.get("parish")))

pr = hdr_row + 1
for parish in all_parishes:
    tot.cell(row=pr, column=1, value=parish).font = BLACK
    for j in range(6):
        parts = []
        for rc, desc in ALL_RCS:
            meta = sheet_meta[rc]
            sheet_title = meta["title"]
            for tbl in meta["tables"]:
                if tbl["first_row"] > tbl["last_row"]:
                    continue
                col = get_column_letter(tbl["COL_JANJUN"][j])
                parish_col_letter = get_column_letter(tbl["parish_col"])
                parts.append(f"SUMIF('{sheet_title}'!{parish_col_letter}{tbl['first_row']}:{parish_col_letter}{tbl['last_row']},\"{parish}\",'{sheet_title}'!{col}{tbl['first_row']}:{col}{tbl['last_row']})")
        f = "=" + "+".join(parts) if parts else "=0"
        c = tot.cell(row=pr, column=2 + j, value=f); c.font = GREEN; c.number_format = NUMFMT; c.border = BORDER
    for j in range(6):
        parts = []
        for rc, desc in ALL_RCS:
            meta = sheet_meta[rc]
            sheet_title = meta["title"]
            for tbl in meta["tables"]:
                if tbl["first_row"] > tbl["last_row"]:
                    continue
                col = get_column_letter(tbl["COL_PROJ"][j])
                parish_col_letter = get_column_letter(tbl["parish_col"])
                parts.append(f"SUMIF('{sheet_title}'!{parish_col_letter}{tbl['first_row']}:{parish_col_letter}{tbl['last_row']},\"{parish}\",'{sheet_title}'!{col}{tbl['first_row']}:{col}{tbl['last_row']})")
        f = "=" + "+".join(parts) if parts else "=0"
        c = tot.cell(row=pr, column=8 + j, value=f); c.font = GREEN; c.number_format = NUMFMT; c.border = BORDER
    pr += 1

parish_total_row = pr
tot.cell(row=parish_total_row, column=1, value="TOTAL").font = BOLD
for j in range(6):
    col = get_column_letter(2 + j)
    c = tot.cell(row=parish_total_row, column=2 + j, value=f"=SUM({col}{hdr_row+1}:{col}{pr-1})")
    c.font = BOLD; c.number_format = NUMFMT
for j in range(6):
    col = get_column_letter(8 + j)
    c = tot.cell(row=parish_total_row, column=8 + j, value=f"=SUM({col}{hdr_row+1}:{col}{pr-1})")
    c.font = BOLD; c.number_format = NUMFMT

# ── READ ME ──────────────────────────────────────────────────────────
rm = wb.create_sheet("Read Me", 0)
rm.column_dimensions["A"].width = 112
lines = [
    ("Rate Class Forecast - 2026", BOLD),
    ("", None),
    ("Source: jps_actuals + jps_kam tables, Supabase project bhrswnbenkvflpdjhfpa. Pulled 2026-07-06.", BLACK),
    ("Scope: all rate classes - RT10 (Residential), RT20 (SME/Gen. Service), RT40 (LV Commercial), RT50 (MV Commercial), RT60-ST (Street Lights), RT70 (Large Industrial).", BLACK),
    ("Jan-Jun 2026: real actuals. Jul-Dec 2026: formula-driven projection (see Methodology). Each rate class has its own tab with a local Drivers block at the top.", BLACK),
    ("", None),
    ("METHODOLOGY CHANGE (v3) - why this replaced the prior version:", Font(name=FONT, bold=True, color="C00000")),
    ("The first version projected Jul-Dec from each row's OWN 2024/2025 same-month history, scaled by a class-wide growth rate.", BLACK),
    ("This broke badly for any account whose 2025 volume was tiny (new, dormant, or just onboarded) but whose 2026 volume has since", BLACK),
    ("ramped up - e.g. National Water Commission (100012-101367): real Apr-Jun 2026 average is ~23,446 kWh/month, but the old formula", BLACK),
    ("anchored on a ~76 kWh historical Jul figure and projected July at ~72 kWh - a fake 'decline' driven entirely by stale history, not reality.", BLACK),
    ("Fixed by anchoring every row's projection on ITS OWN real Apr-Jun 2026 average, then applying a class-wide seasonal shape ratio", BLACK),
    ("(that month's 2025 total / the class's 2025 Apr-Jun average) on top. Seasonality still comes from real history; growth now comes", BLACK),
    ("from each row's own actual recent trend, not a multi-year-old absolute level.", BLACK),
    ("", None),
    ("DATA QUALITY FLAGS - NOT FIXED, DISCLOSED HERE:", Font(name=FONT, bold=True, color="C00000")),
    ("1) jps_actuals 2026-06 contains 23,780 rows tagged rate_class='RT20', segment='Commercial' that are really residential", ITALIC_RED),
    ("   Prepaid/Zero bucket data (customer_count in the THOUSANDS per parish, e.g. 5,832 in KSAS - genuinely residential scale).", ITALIC_RED),
    ("   Inserted 2026-07-03 14:45-14:48 UTC, no matching sync_runs entry. EXCLUDED from the RT20 tab entirely.", ITALIC_RED),
    ("2) Separately, RT40 and RT50 contain real 'Postpaid'/'Zero'/'Zero-Postpaid' rows by parish, present in every month except June 2026.", ITALIC_RED),
    ("   Unlike #1, customer_count here is small (single digits to low 20s per parish) - NOT clearly the same mass-residential pattern.", ITALIC_RED),
    ("   Could be real small aggregated postpaid commercial accounts, or a smaller-scale version of the same mistagging - unconfirmed.", ITALIC_RED),
    ("   Moved to the 'Unidentified Rows' tab, kept fully visible with real numbers, but EXCLUDED from every rate class tab and the Total.", ITALIC_RED),
    ("   Needs a real answer from whoever manages the CIS extract before it's folded back into RT40/RT50 or RT10/RT20.", ITALIC_RED),
    ("3) KAM column on RT40/50/60-ST/70 comes from jps_kam (jps_ac -> kam). Only accounts with an assigned KAM show a name; the rest show '-'.", ITALIC_RED),
    ("", None),
    ("Methodology (Jul-Dec 2026 projection, per row):", BOLD),
    ("  projected kWh (month m) = AVERAGE(that row's Apr,May,Jun 2026 actual) x seasonal ratio for month m x (1 + Growth Adjustment toggle)", BLACK),
    ("                            + (Other Monthly Adjustment toggle x row's share of Jun-2026 kWh within its table)", BLACK),
    ("  seasonal ratio (month m) = (rate class's real 2025 total for month m) / (rate class's real 2025 Apr-Jun monthly average) - calculated live, per tab", BLACK),
    ("  Customer Count Growth toggle applies only on the Total tab, as a class-level add-on (new accounts), not spread across existing rows", BLACK),
    ("", None),
    ("Tabs:", BOLD),
    ("  Total - all rate classes, Jan-Jun actual + Jul-Dec projected, Customer Count Growth add-on, By-Parish breakdown.", BLACK),
    ("  RT10 / RT20 / RT40 / RT50 / RT60_ST / RT70 - one tab per rate class, Drivers block + detail rows + projection.", BLACK),
    ("  Unidentified Rows - flagged ambiguous data (see flag #2 above), excluded from all totals.", BLACK),
]
r = 1
for text, font in lines:
    c = rm.cell(row=r, column=1, value=text)
    if font:
        c.font = font
    r += 1

wb.move_sheet("Read Me", offset=-100)
wb.move_sheet("Total", offset=-99)
wb.active = 0
wb.save(OUT)
print("saved", OUT)
print("sheets:", wb.sheetnames)
