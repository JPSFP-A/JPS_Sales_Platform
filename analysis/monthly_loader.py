# -*- coding: utf-8 -*-
# Monthly data load: "Check Consumption Detail/Report" Excel files -> Supabase.
# Single pass per file, produces two things and upserts both directly:
#   1) jps_demand_actuals   (kVA by rate_class/sub_class/demand_type, from SCAT_CODE)
#   2) jps_special_account_adjustments  (BILLED_CONSUMP_ADJ != 0 on the 40 Special
#      Accounts, cross-referenced with who/dept from special_accounts_who.json)
#
# Run from this folder (D:\Projects\Sales_Platform\analysis) with the month's new
# "Check Consumption..." file(s) present alongside the prior months' files (or just
# the new one — reprocessing old files is safe, upserts are idempotent).
#
# Requires a local .env in this folder (gitignored) with:
#   SUPABASE_URL=https://bhrswnbenkvflpdjhfpa.supabase.co
#   SUPABASE_SERVICE_ROLE_KEY=<service_role key from Supabase dashboard -> Settings -> API>
# The service_role key bypasses RLS — never commit it, never share it outside this file.

import csv, json, glob, os, sys, time, zipfile
import openpyxl
import requests

HERE = os.path.dirname(os.path.abspath(sys.argv[0]))
os.chdir(HERE)

DEMAND_SCAT = {"KVA": "standard", "KVAP": "on_peak", "KVAO": "off_peak", "KVAL": "partial_peak"}
MON = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}


def norm(s):
    return str(s).strip().upper()


def parse_month_year(v):
    if hasattr(v, 'year'):
        return v.year, v.month
    s = str(v).strip().upper()
    if '-' in s:
        mo, yr = s.split('-')
        return int(yr), MON[mo[:3]]
    return None, None


def is_real_xlsx(fp):
    try:
        with zipfile.ZipFile(fp):
            return True
    except Exception:
        return False


def load_env(path):
    env = {}
    if os.path.exists(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def load_rcmap():
    rcmap = {}
    csv_path = r"C:\Users\jwilson\Downloads\Rate categorry Data mapping.csv"
    with open(csv_path, encoding="utf-8-sig") as f:
        for row in list(csv.reader(f))[1:]:
            if len(row) >= 3:
                rcmap[norm(row[2])] = row[0]
    rcmap['RT76'] = 'RT70'  # Jamalco-only wholesale tariff variant (Special Accounts Directory: "RT 76 consists of Jamalco only")
    return rcmap


def supabase_upsert(base_url, key, table, on_conflict, rows, batch=500):
    if not rows:
        print(f"  {table}: 0 rows, skipping")
        return
    url = f"{base_url}/rest/v1/{table}?on_conflict={on_conflict}"
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    total = 0
    for i in range(0, len(rows), batch):
        chunk = rows[i:i + batch]
        r = requests.post(url, headers=headers, data=json.dumps(chunk), timeout=60)
        if r.status_code >= 300:
            print(f"  ERROR upserting {table} batch {i}-{i+len(chunk)}: {r.status_code} {r.text[:500]}")
            sys.exit(1)
        total += len(chunk)
        print(f"  {table}: upserted {total}/{len(rows)}")


def main():
    env = load_env(os.path.join(HERE, ".env"))
    base_url = env.get("SUPABASE_URL") or os.environ.get("SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not base_url or not key:
        print("Missing SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY.")
        print(f"Create a .env file at {os.path.join(HERE, '.env')} with those two lines")
        print("(get the service_role key from Supabase dashboard -> Settings -> API).")
        sys.exit(1)

    rcmap = load_rcmap()
    who_map = {}
    who_path = os.path.join(HERE, "special_accounts_who.json")
    if os.path.exists(who_path):
        who_map = json.load(open(who_path))
    special = set(who_map.keys())

    files = sorted(set(glob.glob("Check Consumption*.xlsx")) | set(glob.glob("Check Consumption*.xls")))
    print("files:", len(files))
    if not files:
        print("No 'Check Consumption*' files found in", HERE)
        sys.exit(1)

    demand_agg = {}       # (year,month,rate_class,sub_class,demand_type) -> kva sum
    adj_agg = {}          # (year,month,jps_ac,scat_code) -> row dict (last value wins, matches DB unique key)
    unmapped_srat = {}

    t0 = time.time()
    for fp in files:
        tf = time.time()
        n = kept_d = kept_a = 0
        if is_real_xlsx(fp):
            wb = openpyxl.load_workbook(fp, read_only=True)
            sn = [s for s in wb.sheetnames if "CONSUMP" in s.upper()]
            sn = sn[0] if sn else wb.sheetnames[0]
            rows_iter = wb[sn].iter_rows(values_only=True)
        else:
            fh = open(fp, encoding="iso-8859-1")
            rows_iter = csv.reader(fh, delimiter="\t")

        hdr = idx = None
        for r in rows_iter:
            if r and r[0] == "CUST_CODE":
                hdr = list(r)
                idx = {name: i for i, name in enumerate(hdr)}
                continue
            if hdr is None:
                continue
            if len(r) <= idx.get("MONTH_YEAR", 11):
                continue
            n += 1
            serv_rate = norm(r[idx["SERV_RATE"]])
            rc = rcmap.get(serv_rate)
            scat = r[idx["SCAT_CODE"]]
            year, month = parse_month_year(r[idx["MONTH_YEAR"]])

            # --- demand (jps_demand_actuals) ---
            if rc and rc != "EV" and scat in DEMAND_SCAT and year:
                val = r[idx["NET_BILLED_CONSUMP"]]
                if val not in (None, ''):
                    sub_class = r[idx["STYP_CODE"]]
                    dkey = (year, month, rc, sub_class, DEMAND_SCAT[scat])
                    demand_agg[dkey] = demand_agg.get(dkey, 0.0) + float(val)
                    kept_d += 1
            elif not rc and scat in DEMAND_SCAT:
                unmapped_srat[serv_rate] = unmapped_srat.get(serv_rate, 0) + 1

            # --- special-account billing adjustments (jps_special_account_adjustments) ---
            ac = f"{r[idx['CUST_CODE']]}-{r[idx['PREM_CODE']]}"
            if ac in special and year:
                try:
                    adj = float(r[idx["BILLED_CONSUMP_ADJ"]])
                except (TypeError, ValueError):
                    adj = 0
                if adj != 0:
                    who = who_map[ac]
                    akey = (year, month, ac, scat)
                    adj_agg[akey] = {
                        "year": year, "month": month, "jps_ac": ac,
                        "account_name": who["name"], "rate_class": rc or r[idx["SERV_RATE"]],
                        "scat_code": scat, "adj_value": adj,
                        "who": who["who"], "dept": who["dept"],
                    }
                    kept_a += 1

        if is_real_xlsx(fp):
            wb.close()
        else:
            fh.close()
        print(f"{fp}: {n} rows scanned, {kept_d} demand rows, {kept_a} adj rows, {time.time()-tf:.1f}s")

    print("TOTAL parse time", round(time.time() - t0, 1), "s")
    print("aggregated demand keys:", len(demand_agg))
    print("aggregated adjustment keys:", len(adj_agg))
    if unmapped_srat:
        print("WARNING unmapped SERV_RATE codes (not loaded):", unmapped_srat)

    demand_rows = [
        {"year": y, "month": m, "rate_class": rc, "sub_class": sc, "demand_type": dt, "demand_kva": round(v, 4)}
        for (y, m, rc, sc, dt), v in demand_agg.items()
    ]
    adj_rows = list(adj_agg.values())

    print("\nUpserting to Supabase...")
    supabase_upsert(base_url, key, "jps_demand_actuals", "year,month,rate_class,sub_class,demand_type", demand_rows)
    supabase_upsert(base_url, key, "jps_special_account_adjustments", "year,month,jps_ac,scat_code", adj_rows)
    print("\nDone.")


if __name__ == "__main__":
    main()
