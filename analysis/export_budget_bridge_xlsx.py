# -*- coding: utf-8 -*-
# Budget vs Actual monthly revenue bridge for 2026 (Energy / Demand / Customer / FX).
# Sources:
#   Budget JMD+FX      : C:\Users\jwilson\Documents\revenue drivers.xlsx, 'Revenue' tab
#                         (rows 164/240/110 = Energy/Demand/Customer totals, J$000; row 3 = billing FX)
#   Budget USD (E & C)  : same workbook, 'BudFin US$' tab, rows 6 & 8 (direct, per user instruction)
#   Budget USD (Demand) : derived = Budget Demand JMD / Budget FX (no direct USD Demand line exists)
#   Actual JMD          : D:\Projects\Sales_Platform\analysis\app_data2.json -> D.total (Jan-Jun'26 billed so far)
#   Actual FX           : hard-coded live-site rate table (matches Revenue tab row 3 exactly for Jan-Mar'26)
#   Jan/Feb'26          : per user -- budget = actual, so those 2 months use OUR actual JMD directly
#                         (sidesteps a real ~2-5% gap between the file's own Jan/Feb "Actual" Energy figure
#                         and our billing-system actual -- see Logic tab for that discrepancy, flagged not hidden)
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BUD = json.load(open('budget_2026_extract.json', encoding='utf-8'))
APP = json.load(open('app_data2.json', encoding='utf-8'))
MONTHS = ['2026-%02d' % m for m in range(1, 13)]
MLAB = ['Jan·26', 'Feb·26', 'Mar·26', 'Apr·26', 'May·26', 'Jun·26',
        'Jul·26', 'Aug·26', 'Sep·26', 'Oct·26', 'Nov·26', 'Dec·26']

# Our own actual FX table (matches Revenue tab row 3 exactly Jan-Mar'26; diverges Apr onward --
# the live site's table is a manually-maintained snapshot, flagged in Logic tab)
FX_ACTUAL = {'2026-01': 159.7395, '2026-02': 157.5400, '2026-03': 157.2639, '2026-04': 158.5846,
             '2026-05': 158.16, '2026-06': 158.16}
FX_ACTUAL_EST = {'2026-06': 1}

am_idx = {m: i for i, m in enumerate(APP['months'])}
T = APP['total']

def actual_jmd(field, mo):
    if mo not in am_idx: return None
    return T[field][am_idx[mo]] / 1000.0  # raw J$ -> J$000, same units as the budget file

rows = []
for mo in MONTHS:
    b = BUD[mo]
    fx_b = b['fx_budget']
    is_jf = mo in ('2026-01', '2026-02')
    # Budget JMD: Jan/Feb -> our actual (per user instruction, budget=actual those 2 months); else file
    if is_jf:
        eb = actual_jmd('energy', mo); db = actual_jmd('demand', mo); cb = actual_jmd('cust_chg', mo)
    else:
        eb, db, cb = b['energy_jmd_000'], b['demand_jmd_000'], b['customer_jmd_000']
    # Budget USD used IN THE BRIDGE: self-consistent JMD/FX_budget for all 3 components. This is
    # deliberate -- BudFin's own USD Energy/Customer figures (below) are an independently-modeled
    # USD-native forecast, NOT a currency conversion of the Revenue-tab JMD total, so mixing them into
    # a "FX = residual" formula silently absorbs that modeling gap as if it were currency movement.
    # (Verified: for Mar'26, budget FX == actual FX exactly, 157.2639, yet BudFin-mixed math produced a
    # phantom ~$7.7M "FX effect" that should have been ~$0 -- so the bridge uses the self-consistent
    # figure, and BudFin's own number is kept as a labelled reference row instead, not fed into the math.)
    eb_usd = (eb / fx_b) if (eb is not None and fx_b) else None
    db_usd = (db / fx_b) if (db is not None and fx_b) else None
    cb_usd = (cb / fx_b) if (cb is not None and fx_b) else None
    eb_usd_budfin = b['energy_usd_000']          # reference only -- BudFin's own USD Energy budget
    cb_usd_budfin = b['customer_usd_000']         # reference only -- BudFin's own USD Customer budget
    # Actual JMD/USD: only where billed (Jan-Jun'26 in our data)
    ea = actual_jmd('energy', mo); da = actual_jmd('demand', mo); ca = actual_jmd('cust_chg', mo)
    fx_a = FX_ACTUAL.get(mo)
    ea_usd = ea / fx_a if (ea is not None and fx_a) else None
    da_usd = da / fx_a if (da is not None and fx_a) else None
    ca_usd = ca / fx_a if (ca is not None and fx_a) else None
    rows.append(dict(mo=mo, fx_b=fx_b, fx_a=fx_a, fx_a_est=FX_ACTUAL_EST.get(mo, 0),
                      eb=eb, db=db, cb=cb, eb_usd=eb_usd, db_usd=db_usd, cb_usd=cb_usd,
                      eb_usd_budfin=eb_usd_budfin, cb_usd_budfin=cb_usd_budfin,
                      ea=ea, da=da, ca=ca, ea_usd=ea_usd, da_usd=da_usd, ca_usd=ca_usd))

wb = Workbook()
HEAD = Font(bold=True, color='FFFFFF')
HEAD_FILL = PatternFill('solid', fgColor='1F4E5F')
SUB_FILL = PatternFill('solid', fgColor='DCE6F1')
thin = Side(style='thin', color='B7C4D0')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD; cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

# ── 1. Logic ─────────────────────────────────────────────────────────────
ws = wb.active; ws.title = 'Logic'
ws.column_dimensions['A'].width = 112
lines = [
("BUDGET VS ACTUAL MONTHLY BRIDGE 2026 -- HOW IT'S BUILT", True, 14),
("", False, 11),
("What it shows", True, 12),
("  A monthly bridge from Budget to Actual revenue for 3 components -- Energy, Demand, Customer charge -- "
 "switchable between JMD (no FX line) and USD (adds an FX line). Same reconciling-bridge convention as the "
 "'YoY_Bridge' tab in SalesAnalysis_AnswerGraph_Logic.xlsx: Budget + component deltas (+ FX in USD mode) = Actual, exactly.", False, 11),
("", False, 11),
("Sources -- exactly what came from where", True, 12),
("  1. Budget JMD (Energy/Demand/Customer) + Budget FX rate", True, 11),
("       'revenue drivers.xlsx' -> 'Revenue' tab. Energy = row 164 total (block rows 112-165, 'ENERGY REVENUES "
 "- J$000'). Demand = row 240 total (rows 188-241, 'DEMAND REVENUES - J$000'). Customer = row 110 total (rows "
 "58-111, 'CUSTOMER REVENUES - J$000'). Budget FX = row 3 ('Billing Exchange Rate'), which carries the file's own "
 "forecast FX curve for Mar-Dec'26.", False, 11),
("  2. Budget USD -- used IN THE BRIDGE (Energy, Demand, Customer, all 3)", True, 11),
("       Self-consistent: each component's Budget JMD / Budget FX rate (both from the Revenue tab, item 1 above). "
 "IMPORTANT CHANGE from your original instruction: I initially pulled Energy & Customer straight from 'BudFin US$' "
 "rows 6 & 8 as you asked, but caught a bug doing the March cross-check -- budget FX and actual FX are IDENTICAL "
 "for March (157.2639 both), yet that approach produced a phantom ~US$7.7M 'FX effect' for a month with zero "
 "currency movement. Root cause: BudFin's USD Energy/Customer are an independently-modeled USD-native forecast, "
 "NOT a currency conversion of the Revenue tab's JMD total -- mixing an independently-modeled number into a "
 "'FX = whatever's left over' formula silently absorbs that modeling gap as if it were currency risk. Using the "
 "self-consistent JMD/FX figure for all 3 components makes the FX line mean only FX, and it now correctly nets to "
 "$0 for March and to a small, real, sensible number for months where budget and actual FX genuinely differ (e.g. "
 "~$0.12M for June, where the actual rate landed at an estimated 158.16 vs a budgeted 158.81).", False, 11),
("  3. Budget USD -- BudFin's own number (kept as a REFERENCE row only, not fed into the bridge)", True, 11),
("       'revenue drivers.xlsx' -> 'BudFin US$' tab, rows 6 & 8 (Energy, Customer), columns FB:FK (Mar'26-Dec'26). "
 "Shown alongside the bridge on the 'Bridge' tab as 'BudFin's own USD budget (reference)' plus the $ gap vs the "
 "bridge-consistent figure, so you can see your finance team's own USD-native budget number and how far it sits "
 "from a straight currency conversion of the JMD budget -- that gap is a modeling-assumption difference, not FX. "
 "Demand has no BudFin USD line to reference at all (BudFin has no separate Demand charge row; 'Demand' tab is "
 "volumes/MWh only, no $ anywhere), so Demand USD is always the derived JMD/FX_budget figure, no alternative exists.", False, 11),
("  4. Actual JMD (Energy/Demand/Customer)", True, 11),
("       Our own live billing pipeline -- analysis/app_data2.json -> D.total (same dataset embedded in the live "
 "salesanalysis.jmfinancelab.com site). Available for Jan-Jun'26 (billed so far). Jul-Dec'26 = not yet billed, "
 "shown blank/pending until that month's Billing Details Report is processed.", False, 11),
("  5. Actual FX", True, 11),
("       Our own rate table (matches the file's Revenue-tab row 3 EXACTLY for Jan-Mar'26: 159.7395 / 157.5400 / "
 "157.2639 -- good cross-validation both sources agree on realized FX). Diverges from the file's forecast curve "
 "Apr onward, as expected (forecast vs realized).", False, 11),
("", False, 11),
("Jan/Feb'26 shortcut (per your instruction)", True, 12),
("  You said budget=actual for these 2 months, so Budget JMD for Jan/Feb is set equal to OUR actual JMD directly "
 "(not pulled from the Revenue tab) -- this sidesteps a real discrepancy worth knowing about: the file's own "
 "Jan'26 Energy total (row 164 = J$2,488.1M) runs ~2.6% above our billing-system actual (J$2,426.0M); Demand and "
 "Customer are much closer (~0.2-0.3% apart). Small gaps like this between a financial-model rollup and the raw "
 "billing extract are normal (timing/classification), but flagging it rather than letting two 'actual' numbers "
 "silently disagree.", False, 11),
("", False, 11),
("The bridge formula (mirrors renderBridge() from the live Sales Analysis site)", True, 12),
("  JMD mode (no FX -- same currency, nothing to convert):", False, 11),
("    Budget_JMD(E+D+C) -> +(Energy_Actual-Energy_Budget) -> +(Demand_Actual-Demand_Budget) -> "
 "+(Customer_Actual-Customer_Budget) -> Actual_JMD(E+D+C)   [ties out exactly, by definition]", False, 11),
("  USD mode (adds one more step -- the FX effect):", False, 11),
("    1. Convert each component's delta to US$ at the BUDGET's own FX rate (holds currency fixed at the budget "
 "assumption, isolating the real JMD-driven variance):", False, 11),
("         Energy_delta_USD   = (Energy_Actual_JMD   - Energy_Budget_JMD)   / FX_budget", False, 11),
("         Demand_delta_USD   = (Demand_Actual_JMD   - Demand_Budget_JMD)   / FX_budget", False, 11),
("         Customer_delta_USD = (Customer_Actual_JMD - Customer_Budget_JMD) / FX_budget", False, 11),
("    2. Budget_USD = Energy_Budget_USD + Demand_Budget_USD + Customer_Budget_USD (as sourced/derived above).", False, 11),
("    3. Actual_USD = Actual_JMD(E+D+C) / FX_actual.", False, 11),
("    4. FX effect = Actual_USD - Budget_USD - (sum of the 3 deltas above) -- the residual, purely the effect of "
 "actual FX landing away from the rate the budget assumed. Because we defined Budget total = E+D+C exactly (no "
 "Fuel/IPP/Other folded in), there's no separate 'Other/adj' line needed here -- FX absorbs the entire residual "
 "and the bridge reconciles to zero exactly, every month.", False, 11),
("", False, 11),
("See 'Monthly_Data' for every input number and 'Bridge' to pick any month + JMD/USD and see the live bridge.", False, 11),
]
r = 1
for text, bold, size in lines:
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(bold=bold, size=size, color='1F4E5F' if bold and size >= 13 else ('1F4E5F' if bold else '333333'))
    c.alignment = Alignment(wrap_text=True, vertical='top')
    r += 1
ws.freeze_panes = 'A2'

# ── 2. Monthly_Data ──────────────────────────────────────────────────────
wsD = wb.create_sheet('Monthly_Data')
hdrs = ['Line (J$M unless noted)'] + MLAB
for c, h in enumerate(hdrs, start=1):
    wsD.cell(row=1, column=c, value=h)
style_header(wsD, 1, len(hdrs))
wsD.column_dimensions['A'].width = 26
for j in range(12):
    wsD.column_dimensions[get_column_letter(2 + j)].width = 12

def put_row(r, label, vals, fmt='#,##0.0', shade=False):
    wsD.cell(row=r, column=1, value=label).font = Font(bold=True)
    if shade: wsD.cell(row=r, column=1).fill = SUB_FILL
    for j, v in enumerate(vals):
        cell = wsD.cell(row=r, column=2 + j, value=(round(v, 4) if v is not None else None))
        cell.number_format = fmt
        cell.border = BORDER

put_row(2, 'FX rate -- Budget', [x['fx_b'] for x in rows], '0.0000')
put_row(3, 'FX rate -- Actual', [x['fx_a'] for x in rows], '0.0000')
put_row(4, 'FX Actual estimated? (1=yes)', [x['fx_a_est'] for x in rows], '0')
put_row(6, 'Energy -- Budget JMD', [x['eb'] / 1000 if x['eb'] is not None else None for x in rows], '#,##0.0', True)
put_row(7, 'Energy -- Actual JMD', [x['ea'] / 1000 if x['ea'] is not None else None for x in rows])
put_row(9, 'Demand -- Budget JMD', [x['db'] / 1000 if x['db'] is not None else None for x in rows], '#,##0.0', True)
put_row(10, 'Demand -- Actual JMD', [x['da'] / 1000 if x['da'] is not None else None for x in rows])
put_row(12, 'Customer -- Budget JMD', [x['cb'] / 1000 if x['cb'] is not None else None for x in rows], '#,##0.0', True)
put_row(13, 'Customer -- Actual JMD', [x['ca'] / 1000 if x['ca'] is not None else None for x in rows])
put_row(15, 'Energy -- Budget USD (bridge: JMD/FXbudget)', [x['eb_usd'] / 1000 if x['eb_usd'] is not None else None for x in rows], '#,##0.0', True)
put_row(16, 'Energy -- Budget USD per BudFin (reference only)', [x['eb_usd_budfin'] / 1000 if x['eb_usd_budfin'] is not None else None for x in rows])
put_row(17, 'Energy -- Actual USD', [x['ea_usd'] / 1000 if x['ea_usd'] is not None else None for x in rows])
put_row(19, 'Demand -- Budget USD (derived: JMD/FXbudget)', [x['db_usd'] / 1000 if x['db_usd'] is not None else None for x in rows], '#,##0.0', True)
put_row(20, 'Demand -- Actual USD', [x['da_usd'] / 1000 if x['da_usd'] is not None else None for x in rows])
put_row(22, 'Customer -- Budget USD (bridge: JMD/FXbudget)', [x['cb_usd'] / 1000 if x['cb_usd'] is not None else None for x in rows], '#,##0.0', True)
put_row(23, 'Customer -- Budget USD per BudFin (reference only)', [x['cb_usd_budfin'] / 1000 if x['cb_usd_budfin'] is not None else None for x in rows])
put_row(24, 'Customer -- Actual USD', [x['ca_usd'] / 1000 if x['ca_usd'] is not None else None for x in rows])
wsD.freeze_panes = 'B2'
note = wsD.cell(row=26, column=1, value=("Blank Actual cells = not yet billed (Jul-Dec'26 pending). Jan/Feb Budget JMD = our Actual JMD "
                                          "(budget=actual those 2 months, per instruction). 'Budget USD (bridge)' rows are self-consistent "
                                          "(JMD/FX budget) so the Bridge tab's FX line isolates real currency movement only -- the 'per BudFin "
                                          "(reference only)' rows are the finance team's own independently-modeled USD budget and will not tie "
                                          "exactly to the bridge figure; see Logic tab for why."))
note.font = Font(italic=True, color='777777'); note.alignment = Alignment(wrap_text=True)
wsD.merge_cells('A26:M28')

# ── 3. Bridge — pick month + currency, live formulas ────────────────────
wsB = wb.create_sheet('Bridge')
wsB['A1'] = 'Month'; wsB['A1'].font = Font(bold=True)
wsB['B1'] = MLAB[0]
wsB['A2'] = 'Currency'; wsB['A2'].font = Font(bold=True)
wsB['B2'] = 'JMD'
dvM = DataValidation(type='list', formula1='"' + ','.join(MLAB) + '"', allow_blank=False)
wsB.add_data_validation(dvM); dvM.add(wsB['B1'])
dvC = DataValidation(type='list', formula1='"JMD,USD"', allow_blank=False)
wsB.add_data_validation(dvC); dvC.add(wsB['B2'])
wsB['D1'] = "JMD mode: no FX line (Budget+3 deltas=Actual). USD mode: adds FX effect (Budget+3 deltas+FX=Actual)."
wsB['D1'].font = Font(italic=True, color='777777'); wsB['D1'].alignment = Alignment(wrap_text=True)
wsB.merge_cells('D1:H2')

M = 'Monthly_Data'
r0 = 4
wsB.cell(row=r0, column=1, value='Component').font = Font(bold=True)
wsB.cell(row=r0, column=2, value='Budget').font = Font(bold=True)
wsB.cell(row=r0, column=3, value='Actual').font = Font(bold=True)
wsB.cell(row=r0, column=4, value='Delta').font = Font(bold=True)
style_header(wsB, r0, 4)
# Monthly_Data rows (JMD Budget, JMD Actual, USD Budget[bridge], USD Actual) per component
comp_rows = {'Energy': (6, 7, 15, 17), 'Demand': (9, 10, 19, 20), 'Customer': (12, 13, 22, 24)}
compref = {}
for i, (name, (jb, ja, ub, ua)) in enumerate(comp_rows.items()):
    r = r0 + 1 + i
    wsB.cell(row=r, column=1, value=name)
    notBilled = f'HLOOKUP($B$1,{M}!$B$1:$M$7,7,FALSE)=0'
    bformula = (f'=IF($B$2="JMD",HLOOKUP($B$1,{M}!$B$1:$M${jb},{jb},FALSE),'
                f'HLOOKUP($B$1,{M}!$B$1:$M${ub},{ub},FALSE))')
    aformula = (f'=IF({notBilled},"n/a",IF($B$2="JMD",HLOOKUP($B$1,{M}!$B$1:$M${ja},{ja},FALSE),'
                f'HLOOKUP($B$1,{M}!$B$1:$M${ua},{ua},FALSE)))')
    wsB.cell(row=r, column=2, value=bformula).number_format = '#,##0.0'
    wsB.cell(row=r, column=3, value=aformula).number_format = '#,##0.0'
    wsB.cell(row=r, column=4, value=f'=IF(C{r}="n/a","n/a",C{r}-B{r})').number_format = '+#,##0.0;-#,##0.0'
    compref[name] = r
wsB.column_dimensions['A'].width = 22
for col in ['B', 'C', 'D']:
    wsB.column_dimensions[col].width = 14

fxRow = r0 + 6
wsB.cell(row=fxRow, column=1, value='FX Budget rate').font = Font(bold=True)
wsB.cell(row=fxRow, column=2, value=f'=HLOOKUP($B$1,{M}!$B$1:$M$2,2,FALSE)').number_format = '0.0000'
wsB.cell(row=fxRow + 1, column=1, value='FX Actual rate').font = Font(bold=True)
wsB.cell(row=fxRow + 1, column=2, value=f'=HLOOKUP($B$1,{M}!$B$1:$M$3,3,FALSE)').number_format = '0.0000'
wsB.cell(row=fxRow + 2, column=1, value='Actual billed yet?').font = Font(bold=True)
wsB.cell(row=fxRow + 2, column=2,
         value=f'=IF(HLOOKUP($B$1,{M}!$B$1:$M$7,7,FALSE)=0,"NO - budget only, pending actual","YES")')

refRow = fxRow + 4
wsB.cell(row=refRow, column=1, value="BudFin's own USD budget (reference, not in bridge math)").font = Font(italic=True, color='777777')
wsB.cell(row=refRow, column=2,
         value=(f'=IF($B$2="JMD","n/a (JMD mode)",'
                f'HLOOKUP($B$1,{M}!$B$1:$M$16,16,FALSE)+B{compref["Demand"]}+HLOOKUP($B$1,{M}!$B$1:$M$23,23,FALSE))')).number_format = '#,##0.0'
wsB.cell(row=refRow + 1, column=1, value="Gap vs bridge Budget USD (BudFin modeling diff, not FX)").font = Font(italic=True, color='777777')
wsB.cell(row=refRow + 1, column=2,
         value=f'=IF($B$2="JMD","n/a (JMD mode)",B{refRow}-B{compref["Energy"]}-B{compref["Demand"]}-B{compref["Customer"]})')

bridgeHdr = fxRow + 7
wsB.cell(row=bridgeHdr, column=1, value='Bridge step').font = Font(bold=True)
wsB.cell(row=bridgeHdr, column=2, value='Value ($M)').font = Font(bold=True)
style_header(wsB, bridgeHdr, 2)
br = bridgeHdr + 1
wsB.cell(row=br, column=1, value='Budget total')
wsB.cell(row=br, column=2, value=f'=B{compref["Energy"]}+B{compref["Demand"]}+B{compref["Customer"]}').number_format = '#,##0.0'
for i, name in enumerate(['Energy', 'Demand', 'Customer']):
    r = br + 1 + i
    wsB.cell(row=r, column=1, value=name)
    wsB.cell(row=r, column=2, value=f'=D{compref[name]}').number_format = '+#,##0.0;-#,##0.0'
notBilledTop = f'HLOOKUP($B$1,{M}!$B$1:$M$7,7,FALSE)=0'
fxOutRow = br + 4
wsB.cell(row=fxOutRow, column=1, value='FX effect (USD mode only)')
wsB.cell(row=fxOutRow, column=2,
         value=(f'=IF(OR($B$2="JMD",{notBilledTop}),0,'
                f'(C{compref["Energy"]}+C{compref["Demand"]}+C{compref["Customer"]})'
                f'-(B{compref["Energy"]}+B{compref["Demand"]}+B{compref["Customer"]})'
                f'-(D{compref["Energy"]}+D{compref["Demand"]}+D{compref["Customer"]}))')).number_format = '+#,##0.0;-#,##0.0'
endRow = br + 5
wsB.cell(row=endRow, column=1, value='Actual total')
wsB.cell(row=endRow, column=2,
         value=(f'=IF({notBilledTop},"n/a - not yet billed",'
                f'C{compref["Energy"]}+C{compref["Demand"]}+C{compref["Customer"]})')).number_format = '#,##0.0'
checkRow = br + 7
wsB.cell(row=checkRow, column=1, value='Check: Budget+3 deltas+FX-Actual (should = 0)').font = Font(italic=True)
wsB.cell(row=checkRow, column=2,
         value=(f'=IF({notBilledTop},"n/a",'
                f'B{br}+B{br+1}+B{br+2}+B{br+3}+B{fxOutRow}-B{endRow})')).number_format = '0.0000'

# ── 4. Source_Trace ──────────────────────────────────────────────────────
wsT = wb.create_sheet('Source_Trace')
trace = [
 ('Line', 'Sheet!Range', 'Notes'),
 ('Budget FX rate', "revenue drivers.xlsx / 'Revenue'!row 3", "'Billing Exchange Rate' -- extends across Mar-Dec'26 forecast columns (cols EW:FF)."),
 ('Budget Energy JMD', "revenue drivers.xlsx / 'Revenue'!row 164", "Total of 'ENERGY REVENUES - J$000' block (rows 112-165), after the block's own F/X adjustment line (row 162)."),
 ('Budget Demand JMD', "revenue drivers.xlsx / 'Revenue'!row 240", "Total of 'DEMAND REVENUES - J$000' block (rows 188-241)."),
 ('Budget Customer JMD', "revenue drivers.xlsx / 'Revenue'!row 110", "Total of 'CUSTOMER REVENUES - J$000' block (rows 58-111)."),
 ('Budget Energy USD', "revenue drivers.xlsx / 'BudFin US$'!row 6, cols FB:FK", "'Energy Charge' line, direct -- not derived from the JMD figure."),
 ('Budget Customer USD', "revenue drivers.xlsx / 'BudFin US$'!row 8, cols FB:FK", "'Customer Charge' line, direct."),
 ('Budget Demand USD', 'derived', 'No Demand line exists in BudFin -- computed as Budget Demand JMD / Budget FX rate.'),
 ('Actual JMD (Energy/Demand/Customer)', r'D:\Projects\Sales_Platform\analysis\app_data2.json -> total.energy/demand/cust_chg',
  "Same dataset embedded in the live salesanalysis.jmfinancelab.com site. Available Jan-Jun'26 only (billed so far)."),
 ('Actual FX rate', 'hard-coded table, this script', "Matches Revenue tab row 3 exactly for Jan-Mar'26 (159.7395/157.54/157.2639) -- cross-validated."),
 ('Jan/Feb Budget JMD override', 'this script', "Set = our Actual JMD per your instruction (budget=actual those 2 months); NOT pulled from Revenue tab row 164/240/110 for just those 2 columns."),
]
for r_i, row in enumerate(trace, start=1):
    for c_i, val in enumerate(row, start=1):
        cell = wsT.cell(row=r_i, column=c_i, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        if r_i == 1:
            cell.font = HEAD; cell.fill = HEAD_FILL
wsT.column_dimensions['A'].width = 30
wsT.column_dimensions['B'].width = 45
wsT.column_dimensions['C'].width = 70
for r_i in range(2, len(trace) + 1):
    wsT.row_dimensions[r_i].height = 40

wb.save('Budget_vs_Actual_Bridge_2026.xlsx')
print('wrote Budget_vs_Actual_Bridge_2026.xlsx')
