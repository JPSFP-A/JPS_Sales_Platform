# -*- coding: utf-8 -*-
# Windows-native LibreOffice recalc — the xlsx skill's bundled recalc.py assumes
# Linux/macOS (AF_UNIX shim crashes on Windows Python; its macro dir is
# ~/.config/libreoffice, but Windows LO profile lives under %APPDATA%). Same
# StarBasic macro approach, just pointed at the right places for this machine.
import json, os, subprocess, sys
from pathlib import Path
from openpyxl import load_workbook

SOFFICE = r'C:\Program Files\LibreOffice\program\soffice.exe'
MACRO_DIR = Path(os.environ['APPDATA']) / 'LibreOffice' / '4' / 'user' / 'basic' / 'Standard'
MACRO_FILE = MACRO_DIR / 'Module1.xba'
MACRO = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''


def setup_macro():
    if MACRO_FILE.exists() and 'RecalculateAndSave' in MACRO_FILE.read_text():
        return True
    if not MACRO_DIR.exists():
        subprocess.run([SOFFICE, '--headless', '--terminate_after_init'], capture_output=True, timeout=30)
        MACRO_DIR.mkdir(parents=True, exist_ok=True)
    MACRO_FILE.write_text(MACRO)
    return True


def recalc(filename, timeout=60):
    abs_path = str(Path(filename).absolute())
    setup_macro()
    cmd = [SOFFICE, '--headless', '--norestore',
           'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application',
           abs_path]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    if result.returncode != 0:
        return {'error': result.stderr or 'soffice failed', 'returncode': result.returncode}

    wb = load_workbook(filename, data_only=True)
    errs = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
    detail = {e: [] for e in errs}
    total = 0
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for e in errs:
                        if e in cell.value:
                            detail[e].append(f'{sn}!{cell.coordinate}')
                            total += 1
                            break
    wb.close()
    out = {'status': 'success' if total == 0 else 'errors_found', 'total_errors': total,
           'error_summary': {k: {'count': len(v), 'locations': v[:20]} for k, v in detail.items() if v}}
    return out


if __name__ == '__main__':
    r = recalc(sys.argv[1], int(sys.argv[2]) if len(sys.argv) > 2 else 60)
    print(json.dumps(r, indent=2))
