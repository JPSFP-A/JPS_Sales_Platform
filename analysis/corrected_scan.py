# -*- coding: utf-8 -*-
# Corrected meter-level scan. Grain = (Srat_Code, billing rate_class) x Parish, so the
# class report is derived by joining the editable mapping. Title resolves by rate_class
# FIRST (so Jamalco/Carib/JPS separate out even when they share a Srat_Code), Srat fallback.
#   mapping cols: Title, Rate_Class, Srat_Code   [Rate categorry Data mapping.csv]
#   parish:       Old Parish -> Parish Grouping  [Parish Grouping.csv]
# Keeps demand split KVAP/KVAL/KVAO. Account grain for non-RT10/RT20; RT10/RT20 -> kWh bins.
import openpyxl, json, time, os, csv
DL=r'C:\Users\jwilson\Downloads'
# Billing files are AUTO-DISCOVERED from Downloads and this folder — drop the new
# month's "Billing Details Report ... <Mon> <YYYY>.xlsx" in either place and rerun.
# (Used to be a hardcoded dict, which is why new months silently never appeared.)
# The scan below is incremental: months already in corrected.json are skipped.
import re as _re, glob as _glob
_MONNUM={'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}
def _discover_billing_files():
    files={}
    cands=sorted(set(_glob.glob(os.path.join(DL,'Billing Details Report*.xls*'))+_glob.glob('Billing Details Report*.xls*')))
    for fp in cands:
        base=os.path.basename(fp)
        m=_re.search(r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*[ _-]*(\d{4})', base.upper())
        if not m:
            print('  ?? cannot parse month/year from filename, SKIPPING:', base, flush=True)
            continue
        key='%s-%02d'%(m.group(2), _MONNUM[m.group(1)])
        if key in files and files[key]!=base:
            print('  ?? duplicate files for', key, ':', files[key], 'vs', base, '- keeping first', flush=True)
            continue
        files[key]=base
    return dict(sorted(files.items()))
FILES=_discover_billing_files()
print('discovered %d billing files:'%len(FILES), flush=True)
for _k,_v in FILES.items(): print('  ',_k,'->',_v, flush=True)
def NUM(v):
    if isinstance(v,(int,float)): return float(v)
    if isinstance(v,str):
        v=v.strip()
        if v=='': return 0.0
        try: return float(v)
        except ValueError: return 0.0
    return 0.0
def norm(s): return str(s).strip().upper()
SEP='||'
RMAP={}; RCMAP={}
for r in list(csv.reader(open(DL+r'\Rate categorry Data mapping.csv')))[1:]:
    if len(r)>=3:
        RMAP[norm(r[2])]=r[0]                       # srat -> title
        if norm(r[1]) not in RCMAP: RCMAP[norm(r[1])]=r[0]   # rate_class -> title
PMAP={}
for r in list(csv.reader(open(DL+r'\Parish Grouping.csv')))[1:]:
    if len(r)>=2: PMAP[norm(r[0])]=r[1]
def title_of(rc,srat): return RCMAP.get(norm(rc)) or RMAP.get(norm(srat))
BIN=50; CAP=5000
def binof(k):
    if k<0: return -1
    b=int(k//BIN)*BIN
    return b if b<CAP else CAP
def _is_zip_xlsx(path):
    with open(path,'rb') as f: return f.read(2)==b'PK'
def _rows_xlsx(path):
    wb=openpyxl.load_workbook(path, read_only=True)
    detail=None
    for sn in wb.sheetnames:
        ws=wb[sn]
        for i,r in enumerate(ws.iter_rows(values_only=True)):
            if r and r[0]=='Cust_Code': detail=sn; break
            if i>8: break
        if detail: break
    ws=wb[detail]
    for r in ws.iter_rows(values_only=True): yield r
    wb.close()
def _rows_tsv(path):
    # Raw CIS export: plain tab-delimited text (4-line report header, blank line, then Cust_Code header row)
    with open(path, encoding='latin-1', errors='replace') as f:
        for line in f:
            yield tuple(line.rstrip('\r\n').split('\t'))
def proc(path):
    rows=_rows_xlsx(path) if _is_zip_xlsx(path) else _rows_tsv(path)
    hdr=None
    for r in rows:
        if r and r[0]=='Cust_Code': hdr=list(r); break
    I={n:k for k,n in enumerate(hdr) if n}
    nc=len(hdr)
    g=lambda r,n: NUM(r[I[n]]) if n in I and I[n]<len(r) else 0.0
    gv=lambda r,n: (r[I[n]] if n in I and I[n]<len(r) else None)
    it=rows
    srt={}; acc={}; buck={}; uns={}; unp={}; n=0; skipped_unbilled=0
    for r in it:
        code=r[I['Cust_Code']]
        if code in (None,'','Cust_Code'): continue
        # cust_billed=0 means the row wasn't actually billed this cycle (verified: 5,719
        # of 724,563 June rows are 0, and over half of those still carry nonzero
        # net_billed_revenue/Cust_Charge — deposits/administrative charges, not real
        # monthly billing) — exclude entirely rather than let them inflate totals.
        cb=gv(r,'cust_billed')
        if cb is not None and str(cb).strip() in ('0','0.0'):
            skipped_unbilled+=1; continue
        n+=1
        srat=str(gv(r,'Srat_Code')); rc=str(gv(r,'rate_class')); title=title_of(rc,srat)
        if title is None: uns[srat+' / '+rc]=uns.get(srat+' / '+rc,0)+1; title='UNMAPPED'
        praw=str(gv(r,'Parish')); pg=PMAP.get(norm(praw))
        if pg is None: unp[praw]=unp.get(praw,0)+1; pg='UNMAPPED'
        kwh=g(r,'net_kwh_billed_consump'); rev=g(r,'net_revenue')
        kvap=g(r,'KVAP_KVA_Demand'); kval=g(r,'KVAL_Demand'); kvao=g(r,'KVAO_Demand'); kva=g(r,'kva_billed_consump')
        en=g(r,'KWHP_KWH_Energy')+g(r,'KWHL_Energy')+g(r,'KWHO_Energy')
        fu=g(r,'fuel')+g(r,'FuelOffPeak')+g(r,'FuelPartialPeak')+g(r,'FuelOnPeak')
        ip=g(r,'IPP_Charge'); cc=g(r,'Cust_Charge'); radj=g(r,'revenue_adj'); nba=g(r,'Net_Billing_Adj')
        key=srat+SEP+rc   # (srat, rate_class)
        cp=srt.setdefault(key,{}).setdefault(pg,[0.0]*13)
        cp[0]+=1;cp[1]+=kwh;cp[2]+=rev;cp[3]+=kvap;cp[4]+=kval;cp[5]+=kvao;cp[6]+=kva;cp[7]+=en;cp[8]+=fu;cp[9]+=ip;cp[10]+=cc;cp[11]+=radj;cp[12]+=nba
        if title in ('RT10','RT20'):   # RT10/RT20 mass-market -> consumption buckets; RT40+ -> by customer
            bk=buck.setdefault(title,{}).setdefault(binof(kwh),[0.0]*7)
            bk[0]+=1;bk[1]+=kwh;bk[2]+=rev;bk[3]+=en;bk[4]+=fu;bk[5]+=ip;bk[6]+=cc
        else:
            # Premise-level grain: the real account number is "cust_code-prem_code"
            # (verified against jps_actuals.jps_ac). One row per PHYSICAL PREMISE, not
            # per customer — a customer with several premises now gets several account
            # rows, each carrying its own true number, instead of one rolled-up row with
            # no single real account number (was cust_code-only; ~54% of accounts here
            # span multiple premises, one as high as 1,190).
            prem=str(gv(r,'Prem_Code') or '').strip()
            full_ac=str(code)+'-'+prem if prem else str(code)
            akey=(full_ac,title)   # split each premise by rate class
            b=acc.get(akey)
            if b is None:
                b=acc[akey]={'name':str(gv(r,'Name') or '').strip(),'srat':srat,'rc':rc,'title':title,'pg':pg,
                             'key':1 if str(gv(r,'key_acct_ind') or '').strip() not in ('','0','N','None') else 0,
                             'v':[0.0]*11}
            v=b['v']
            v[0]+=kwh;v[1]+=rev;v[2]+=kvap;v[3]+=kval;v[4]+=kvao;v[5]+=kva;v[6]+=en;v[7]+=fu;v[8]+=ip;v[9]+=cc;v[10]+=radj
    return srt,acc,buck,uns,unp,n,skipped_unbilled
M=json.load(open('corrected.json')) if os.path.exists('corrected.json') else {
  'months':[], 'bin':BIN,'cap':CAP,'sep':SEP,
  'srat_legend':['count','kwh','rev','kvap','kval','kvao','kva','energy','fuel','ipp','cust','rev_adj','net_bill_adj'],
  'acct_legend':['kwh','rev','kvap','kval','kvao','kva','energy','fuel','ipp','cust','rev_adj'],
  'bucket_legend':['count','kwh','rev','energy','fuel','ipp','cust'],
  'srat':{}, 'acct':{}, 'bucket':{}, 'unmapped_srat':{}, 'unmapped_parish':{}}
for mo,f in FILES.items():
    if mo in M['months']: print('skip',mo,'(done)',flush=True); continue
    t=time.time(); print('processing',mo,flush=True)
    srt,acc,buck,uns,unp,n,skipped_unbilled=proc(os.path.join(DL,f) if not os.path.exists(f) else f)
    M['srat'][mo]={k:{pg:[round(x,2) for x in v] for pg,v in d.items()} for k,d in srt.items()}
    M['bucket'][mo]={tt:{str(b):[round(x,2) for x in v] for b,v in d.items()} for tt,d in buck.items()}
    for (code,title),b in acc.items():
        akey=str(code)+'~'+title   # one record per (premise, rate class) — code is now "cust_code-prem_code"
        rec=M['acct'].setdefault(akey,{'name':b['name'],'code':str(code),'srat':b['srat'],'rc':b['rc'],'title':title,'pg':b['pg'],'key':b['key'],'m':{}})
        rec.update(name=b['name'],code=str(code),srat=b['srat'],rc=b['rc'],title=title,pg=b['pg'],key=b['key'])
        rec['m'][mo]=[round(x) for x in b['v']]
    for s,c in uns.items(): M['unmapped_srat'][s]=M['unmapped_srat'].get(s,0)+c
    for p,c in unp.items(): M['unmapped_parish'][p]=M['unmapped_parish'].get(p,0)+c
    M['months']=sorted(set(M['months']+[mo]))
    json.dump(M,open('corrected.json','w'))
    tr={}
    for k,d in srt.items():
        srat,rc=k.split(SEP,1); tt=title_of(rc,srat) or 'UNMAPPED'; tr[tt]=tr.get(tt,0)+round(sum(v[2] for v in d.values()))
    print('  done',mo,'rows',n,'(skipped',skipped_unbilled,'cust_billed=0)','in',round(time.time()-t),'s | title rev:',{k:round(v/1e6) for k,v in tr.items()},'M',flush=True)
print('WROTE corrected.json months',M['months'],'| unmapped(srat/rc):',M['unmapped_srat'],flush=True)
