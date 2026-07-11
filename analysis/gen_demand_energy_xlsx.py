# -*- coding: utf-8 -*-
# Excel: monthly Demand vs Energy (J$M) for Actual (meter scan) and Budget (Feb 2026 LE).
# 3 line charts: Actuals, Budget, Combined.
import openpyxl, json, csv, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
DL=r'C:\Users\jwilson\Downloads'
MONTHS=['2026-%02d'%m for m in range(1,13)]
MLAB={'2026-01':'Jan-26','2026-02':'Feb-26','2026-03':'Mar-26','2026-04':'Apr-26','2026-05':'May-26','2026-06':'Jun-26',
      '2026-07':'Jul-26','2026-08':'Aug-26','2026-09':'Sep-26','2026-10':'Oct-26','2026-11':'Nov-26','2026-12':'Dec-26'}
def norm(s): return str(s).strip().upper()
# ---- BUDGET monthly (Revenue sheet, J$000 -> J$M) ----
wb=openpyxl.load_workbook('FEB2026_LE.xlsm', read_only=True, data_only=True)
ws=wb['Revenue']; rows=list(ws.iter_rows(values_only=True))
hdr=rows[4]; col26={}
for c,v in enumerate(hdr):
    if isinstance(v,datetime.datetime) and v.year==2026: col26[v.strftime('%Y-%m')]=c
def cls_of(lbl):
    u=lbl.upper()
    for k,t in [('MT10','RT10'),('RATE 10','RT10'),('MT20','RT20'),('RATE 20','RT20'),('MT40','RT40'),('RATE 40','RT40'),
                ('MT60','RT60'),('RATE 60','RT60'),('STREET','RT60'),('TRAFFIC','RT60'),('CEMENT','RT50'),('MT50','RT50'),
                ('RATE 50','RT50'),('PORT AUTH','RT50'),('MT70','RT70'),('RATE 70','RT70')]:
        if k in u: return t
    return None
def section_monthly(a,b):
    res={m:0.0 for m in MONTHS}
    for i in range(a,b+1):
        r=rows[i]; lbl=' '.join(str(r[c]).strip() for c in range(3,7) if r[c] is not None and str(r[c]).strip())
        if cls_of(lbl) is None: continue
        for m,c in col26.items():
            if isinstance(r[c],(int,float)): res[m]+=r[c]
    return {m:v/1000.0 for m,v in res.items()}   # J$000 -> J$M
bud_energy=section_monthly(112,164)
bud_demand=section_monthly(188,240)
wb.close()
# ---- ACTUAL monthly (scan, J$ -> J$M) ----
RMAP={};RCMAP={}
for r in list(csv.reader(open(DL+r'\Rate categorry Data mapping.csv')))[1:]:
    if len(r)>=3:
        RMAP[norm(r[2])]=r[0]
        if norm(r[1]) not in RCMAP: RCMAP[norm(r[1])]=r[0]
d=json.load(open('corrected.json'))
L=d['srat_legend']; ix={n:i for i,n in enumerate(L)}; SEP=d.get('sep','||')
act_energy={}; act_demand={}
for mo in MONTHS:
    if mo not in d['months']: continue
    en=dm=0.0
    for key,par in d['srat'][mo].items():
        for pg,v in par.items():
            en+=v[ix['energy']]; dm+=v[ix['kvap']]+v[ix['kval']]+v[ix['kvao']]
    act_energy[mo]=en/1e6; act_demand[mo]=dm/1e6   # J$ -> J$M
# ---- BUILD WORKBOOK ----
W=Workbook(); ws=W.active; ws.title='Data'
ARIAL=Font(name='Arial',size=10); BOLD=Font(name='Arial',size=10,bold=True)
HEADf=Font(name='Arial',size=10,bold=True,color='FFFFFF'); HEADfill=PatternFill('solid',fgColor='1F3864')
TITLE=Font(name='Arial',size=13,bold=True,color='1F3864')
ws['A1']='JPS — Demand vs Energy Revenue: Actual vs Budget (Feb 2026 LE)'; ws['A1'].font=TITLE
ws['A2']='Monthly, J$ millions · Actual = meter billing detail · Budget = Feb 2026 LE'; ws['A2'].font=Font(name='Arial',size=9,italic=True,color='808080')
hdr=['Month','Actual Demand','Actual Energy','Budget Demand','Budget Energy']
r0=4
for j,h in enumerate(hdr,1):
    c=ws.cell(r0,j,h); c.font=HEADf; c.fill=HEADfill; c.alignment=Alignment(horizontal='center')
for i,mo in enumerate(MONTHS):
    r=r0+1+i
    ws.cell(r,1,MLAB[mo]).font=ARIAL
    ws.cell(r,2, round(act_demand[mo],1) if mo in act_demand else None).font=ARIAL
    ws.cell(r,3, round(act_energy[mo],1) if mo in act_energy else None).font=ARIAL
    ws.cell(r,4, round(bud_demand[mo],1)).font=ARIAL
    ws.cell(r,5, round(bud_energy[mo],1)).font=ARIAL
    for col in range(2,6): ws.cell(r,col).number_format='#,##0.0'
last=r0+len(MONTHS)
ws.column_dimensions['A'].width=10
for col in 'BCDE': ws.column_dimensions[col].width=15
nmon=len([m for m in MONTHS if m in act_demand])
ws.cell(last+2,1,'Actuals through %s; budget full-year 2026.'%(MLAB[[m for m in MONTHS if m in act_demand][-1]] if nmon else 'n/a')).font=Font(name='Arial',size=9,italic=True,color='808080')
# ---- CHARTS ----
chS=W.create_sheet('Charts')
cats=Reference(ws,min_col=1,min_row=r0+1,max_row=last)
def mkchart(title,series,anchor):
    ch=LineChart(); ch.title=title; ch.style=2; ch.height=8.5; ch.width=18
    ch.y_axis.title='J$ millions'; ch.x_axis.title='Month'; ch.x_axis.delete=False; ch.y_axis.delete=False
    for col,name,color,dsh in series:
        ref=Reference(ws,min_col=col,min_row=r0,max_row=last)
        ch.add_data(ref,titles_from_data=True)
    ch.set_categories(cats)
    ch.legend.position='b'
    # color/style series
    colors=[s[2] for s in series]; dash=[s[3] for s in series]
    for k,ser in enumerate(ch.series):
        ser.graphicalProperties.line.solidFill=colors[k]
        ser.graphicalProperties.line.width=28000
        if dash[k]: ser.graphicalProperties.line.dashStyle=dash[k]
        ser.smooth=False
    chS.add_chart(ch,anchor)
# Actuals: demand(2) vs energy(3)
mkchart('Actuals — Demand vs Energy (J$M)',
        [(2,'Actual Demand','C00000',None),(3,'Actual Energy','1F3864',None)],'A1')
# Budget: demand(4) vs energy(5)
mkchart('Budget (Feb LE) — Demand vs Energy (J$M)',
        [(4,'Budget Demand','C00000','sysDash'),(5,'Budget Energy','1F3864','sysDash')],'A19')
# Combined: all 4
mkchart('Combined — Demand & Energy: Actual (solid) vs Budget (dashed)',
        [(2,'Actual Demand','C00000',None),(4,'Budget Demand','C00000','sysDash'),
         (3,'Actual Energy','1F3864',None),(5,'Budget Energy','1F3864','sysDash')],'A37')
out='JPS_Demand_vs_Energy_Actual_Budget.xlsx'
W.save(out)
print('wrote',out,'| actual months:',[MLAB[m] for m in MONTHS if m in act_demand])
print('budget demand YTD JanMay J\$M:',round(sum(bud_demand[m] for m in MONTHS[:5]),0),'| energy:',round(sum(bud_energy[m] for m in MONTHS[:5]),0))
