# -*- coding: utf-8 -*-
# Re-scans the raw monthly Billing Details Report files for RT10 rows only, with a
# finer bin scheme that isolates TRUE zero-kWh consumption (kwh==0 exactly) from the
# [1,50) range -- corrected.json's existing 50kWh-wide "0" bin conflates the two,
# which caused the Zero-tier bug (2.1M+ kWh of real consumption mislabeled as "Zero").
# Country-wide aggregate only (parish='ALL'), matching RT10's existing jps_actuals grain.
import csv, glob, json, os, re, sys

DL = r'C:\Users\jwilson\Downloads'
HERE = os.path.dirname(os.path.abspath(__file__))

MONNUM = {'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6, 'JUL': 7, 'AUG': 8,
          'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12}


def discover_billing_files():
    files = {}
    cands = sorted(set(glob.glob(os.path.join(DL, 'Billing Details Report*.xls*')) + glob.glob(os.path.join(HERE, 'Billing Details Report*.xls*'))))
    for fp in cands:
        base = os.path.basename(fp)
        m = re.search(r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*[ _-]*(\d{4})', base.upper())
        if not m:
            continue
        key = '%s-%02d' % (m.group(2), MONNUM[m.group(1)])
        if key not in files:
            files[key] = fp
    csv_cands = sorted(set(glob.glob(os.path.join(DL, '*.csv')) + glob.glob(os.path.join(HERE, '*.csv'))))
    for fp in csv_cands:
        base = os.path.basename(fp)
        m = re.match(r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s+(\d{2})\.CSV$', base.upper())
        if not m:
            continue
        key = '20%s-%02d' % (m.group(2), MONNUM[m.group(1)])
        if key not in files:
            files[key] = fp
    return dict(sorted(files.items()))


def NUM(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        v = v.strip()
        if v == '':
            return 0.0
        try:
            return float(v)
        except ValueError:
            return 0.0
    return 0.0


def norm(s):
    return str(s).strip().upper()


RMAP = {}
RCMAP = {}
for r in list(csv.reader(open(DL + r'\Rate categorry Data mapping.csv')))[1:]:
    if len(r) >= 3:
        RMAP[norm(r[2])] = r[0]
        if norm(r[1]) not in RCMAP:
            RCMAP[norm(r[1])] = r[0]


def title_of(rc, srat):
    return RCMAP.get(norm(rc)) or RMAP.get(norm(srat))


def _is_zip_xlsx(path):
    with open(path, 'rb') as f:
        return f.read(2) == b'PK'


def _rows_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    detail = None
    for sn in wb.sheetnames:
        ws = wb[sn]
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if r and r[0] == 'Cust_Code':
                detail = sn
                break
            if i > 8:
                break
        if detail:
            break
    ws = wb[detail]
    for r in ws.iter_rows(values_only=True):
        yield r
    wb.close()


def _rows_tsv(path):
    with open(path, encoding='latin-1', errors='replace') as f:
        for line in f:
            yield tuple(line.rstrip('\r\n').split('\t'))


def _rows_csv(path):
    with open(path, encoding='utf-8', errors='replace', newline='') as f:
        for r in csv.reader(f):
            yield tuple(r)


def proc_rt10(path):
    if path.lower().endswith('.csv'):
        rows = _rows_csv(path)
    else:
        rows = _rows_xlsx(path) if _is_zip_xlsx(path) else _rows_tsv(path)
    hdr = None
    for r in rows:
        if r and r[0] == 'Cust_Code':
            hdr = list(r)
            break
    I = {n: k for k, n in enumerate(hdr) if n}
    g = lambda r, n: NUM(r[I[n]]) if n in I and I[n] < len(r) else 0.0
    gv = lambda r, n: (r[I[n]] if n in I and I[n] < len(r) else None)

    # [count, kwh, rev, energy, fuel, ipp, cust_charge] per tier
    tiers = {'TrueZero': [0.0] * 7, '<150': [0.0] * 7}
    n = 0
    for r in rows:
        code = r[I['Cust_Code']] if I.get('Cust_Code', -1) < len(r) else None
        if code in (None, '', 'Cust_Code'):
            continue
        cb = gv(r, 'cust_billed')
        if cb is not None and str(cb).strip() in ('0', '0.0'):
            continue
        srat = str(gv(r, 'Srat_Code'))
        rc = str(gv(r, 'rate_class'))
        title = title_of(rc, srat)
        if title != 'RT10':
            continue
        kwh = g(r, 'net_kwh_billed_consump')
        if not (0 <= kwh < 150):
            continue  # only re-deriving the Zero/<150 split; other tiers unaffected
        n += 1
        rev = g(r, 'net_revenue')
        en = g(r, 'KWHP_KWH_Energy') + g(r, 'KWHL_Energy') + g(r, 'KWHO_Energy')
        fu = g(r, 'fuel') + g(r, 'FuelOffPeak') + g(r, 'FuelPartialPeak') + g(r, 'FuelOnPeak')
        ip = g(r, 'IPP_Charge')
        cc = g(r, 'Cust_Charge')
        tier = tiers['TrueZero'] if kwh == 0 else tiers['<150']
        tier[0] += 1; tier[1] += kwh; tier[2] += rev; tier[3] += en; tier[4] += fu; tier[5] += ip; tier[6] += cc
    return tiers, n


if __name__ == '__main__':
    files = discover_billing_files()
    need = ['2025-%02d' % m for m in range(1, 13)] + ['2026-%02d' % m for m in range(1, 8)]
    missing = [mo for mo in need if mo not in files]
    if missing:
        print('MISSING raw files for:', missing)
    out = {}
    for mo in need:
        if mo not in files:
            continue
        print('processing', mo, files[mo], flush=True)
        tiers, n = proc_rt10(files[mo])
        out[mo] = tiers
        print('  ', mo, 'TrueZero: count=%d kwh=%.2f rev=%.2f | <150: count=%d kwh=%.2f rev=%.2f | rows in [0,150)=%d' % (
            tiers['TrueZero'][0], tiers['TrueZero'][1], tiers['TrueZero'][2],
            tiers['<150'][0], tiers['<150'][1], tiers['<150'][2], n), flush=True)
    json.dump(out, open('rt10_zero_fix_result.json', 'w'))
    print('WROTE rt10_zero_fix_result.json')
