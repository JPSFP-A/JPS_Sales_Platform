# -*- coding: utf-8 -*-
# Pull Bay Roc Hotel demand-charge breakout (3 TOU columns + kVA qty) for Apr & May 2026.
import openpyxl, json, time
FILES={'Apr':'Billing Details Report_Apr 2026.xlsx','May':'Billing Details Report_May 2026.xlsx'}
NEEDLE='BAY ROC'
NUM=lambda v: float(v) if isinstance(v,(int,float)) else 0.0
COLS=['KVAP_KVA_Demand','KVAL_Demand','KVAO_Demand','kva_billed_consump',
      'net_kwh_billed_consump','net_revenue','IPP_Charge','Cust_Charge',
      'KWHP_KWH_Energy','KWHL_Energy','KWHO_Energy','fuel','FuelOffPeak','FuelPartialPeak','FuelOnPeak']
def proc(path):
    wb=openpyxl.load_workbook(path, read_only=True)
    detail=None
    for sn in wb.sheetnames:
        ws=wb[sn]
        for i,r in enumerate(ws.iter_rows(values_only=True)):
            if r and r[0]=='Cust_Code': detail=sn; break
            if i>8: break
        if detail: break
    ws=wb[detail]; it=ws.iter_rows(values_only=True)
    hdr=None
    for r in it:
        if r and r[0]=='Cust_Code': hdr=list(r); break
    I={n:k for k,n in enumerate(hdr) if n}
    g=lambda r,n: NUM(r[I[n]]) if n in I else 0.0
    present={n:(n in I) for n in COLS}
    agg={n:0.0 for n in COLS}; rows=[]; meters=set()
    for r in it:
        code=r[I['Cust_Code']]
        if code in (None,'','Cust_Code'): continue
        nm=str(r[I['Name']] or '').upper()
        if NEEDLE not in nm: continue
        meters.add(code)
        srat=r[I['Srat_Code']]
        for n in COLS: agg[n]+=g(r,n)
        rows.append({'code':code,'name':str(r[I['Name']]).strip(),'srat':srat,
                     'KVAP':g(r,'KVAP_KVA_Demand'),'KVAL':g(r,'KVAL_Demand'),'KVAO':g(r,'KVAO_Demand'),
                     'kva_qty':g(r,'kva_billed_consump'),'kwh':g(r,'net_kwh_billed_consump'),'rev':g(r,'net_revenue')})
    wb.close()
    return present,agg,rows,len(meters)
OUT={}
for mo,f in FILES.items():
    t=time.time(); print('processing',mo,flush=True)
    present,agg,rows,nm=proc(f)
    dem=agg['KVAP_KVA_Demand']+agg['KVAL_Demand']+agg['KVAO_Demand']
    en=agg['KWHP_KWH_Energy']+agg['KWHL_Energy']+agg['KWHO_Energy']
    fu=agg['fuel']+agg['FuelOffPeak']+agg['FuelPartialPeak']+agg['FuelOnPeak']
    OUT[mo]={'present':present,'agg':{k:round(v,2) for k,v in agg.items()},
             'demand_total':round(dem,2),'energy_total':round(en,2),'fuel_total':round(fu,2),
             'meters':nm,'rows':rows}
    json.dump(OUT,open('bayroc.json','w'),indent=1)
    print('  done',mo,'meters',nm,'demand J$',round(dem),'kVA',round(agg['kva_billed_consump']),
          'kWh',round(agg['net_kwh_billed_consump']),'rev',round(agg['net_revenue']),'in',round(time.time()-t),'s',flush=True)
print('WROTE bayroc.json',flush=True)
