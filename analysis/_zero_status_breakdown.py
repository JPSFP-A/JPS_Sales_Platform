# -*- coding: utf-8 -*-
# Pulls Account_Status / is_suspended from the raw CIS billing files for every
# flagged zero-consumption streak account, at its streak's most recent month, and
# joins it onto _zero_streak_result.json.
import csv, json, glob, os, re

HERE = os.path.dirname(os.path.abspath(__file__))
DL = r'C:\Users\jwilson\Downloads'
MONNUM = {'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6, 'JUL': 7, 'AUG': 8,
          'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12}


def discover_billing_files():
    files = {}
    cands = sorted(set(glob.glob(os.path.join(DL, 'Billing Details Report*.xls*')) + glob.glob(os.path.join(HERE, 'Billing Details Report*.xls*'))))
    for fp in cands:
        base = os.path.basename(fp)
        m = re.search(r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*[ _-]*(\d{4})', base.upper())
        if not m:
            continue
        key = '%s-%02d' % (m.group(2), MONNUM[m.group(1)])
        if key not in files:
            files[key] = fp
    csv_cands = sorted(set(glob.glob(os.path.join(DL, '*.csv')) + glob.glob(os.path.join(HERE, '*.csv'))))
    for fp in csv_cands:
        base = os.path.basename(fp)
        m = re.match(r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s+(\d{2})\.CSV$', base.upper())
        if not m:
            continue
        key = '20%s-%02d' % (m.group(2), MONNUM[m.group(1)])
        if key not in files:
            files[key] = fp
    return dict(sorted(files.items()))


def _is_zip_xlsx(path):
    with open(path, 'rb') as f:
        return f.read(2) == b'PK'


def _rows_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    detail = None
    for sn in wb.sheetnames:
        ws = wb[sn]
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if r and r[0] == 'Cust_Code':
                detail = sn
                break
            if i > 8:
                break
        if detail:
            break
    ws = wb[detail]
    for r in ws.iter_rows(values_only=True):
        yield r
    wb.close()


def _rows_tsv(path):
    with open(path, encoding='latin-1', errors='replace') as f:
        for line in f:
            yield tuple(line.rstrip('\r\n').split('\t'))


def _rows_csv(path):
    with open(path, encoding='utf-8', errors='replace', newline='') as f:
        for r in csv.reader(f):
            yield tuple(r)


def load_status_lookup(path):
    if path.lower().endswith('.csv'):
        rows = _rows_csv(path)
    else:
        rows = _rows_xlsx(path) if _is_zip_xlsx(path) else _rows_tsv(path)
    hdr = None
    for r in rows:
        if r and r[0] == 'Cust_Code':
            hdr = list(r)
            break
    I = {n: k for k, n in enumerate(hdr) if n}
    lookup = {}
    for r in rows:
        code = r[I['Cust_Code']] if I.get('Cust_Code', -1) < len(r) else None
        if code in (None, '', 'Cust_Code'):
            continue
        prem = r[I['Prem_Code']] if I.get('Prem_Code', -1) < len(r) else ''
        status = r[I['Account_Status']] if I.get('Account_Status', -1) < len(r) else ''
        susp = r[I['is_suspended']] if I.get('is_suspended', -1) < len(r) else ''
        lookup[(str(code), str(prem))] = (str(status or '').strip(), str(susp or '').strip())
    return lookup


d = json.load(open('_zero_streak_result.json'))
streaks = d['streaks']

files = discover_billing_files()

# group accounts by their streak_end month so each raw file is only opened once
by_month = {}
for s in streaks:
    by_month.setdefault(s['streak_end'], []).append(s)

STATUS_LABEL = {'A': 'Active', 'I': 'Inactive', 'N': 'New', 'F': 'Final/Closed', '': 'Unknown'}

missing_months = []
for mo, accts in sorted(by_month.items()):
    if mo not in files:
        missing_months.append(mo)
        for s in accts:
            s['account_status'] = None
            s['account_status_label'] = 'No raw file for this month'
            s['is_suspended'] = None
        continue
    print('loading', mo, files[mo], 'for', len(accts), 'accounts', flush=True)
    lookup = load_status_lookup(files[mo])
    for s in accts:
        code, prem = s['jps_ac'].split('-', 1) if '-' in s['jps_ac'] else (s['jps_ac'], '')
        v = lookup.get((code, prem))
        if v:
            s['account_status'] = v[0]
            s['account_status_label'] = STATUS_LABEL.get(v[0], v[0])
            s['is_suspended'] = v[1]
        else:
            s['account_status'] = None
            s['account_status_label'] = 'Not found in raw file'
            s['is_suspended'] = None

if missing_months:
    print('WARNING: no raw file for months:', missing_months)

from collections import Counter
status_counts = Counter(s['account_status_label'] for s in streaks)
print('status breakdown (all flagged accounts):', status_counts.most_common())
active_zero_rev = [s for s in streaks if s['account_status'] == 'A' and s['current_monthly_revenue'] == 0]
print('ACTIVE status + $0 revenue (the real anomalies):', len(active_zero_rev))
for s in active_zero_rev:
    print(' ', s['jps_ac'], s['name'], s['rate_class'], s['streak_len'])

json.dump(d, open('_zero_streak_result.json', 'w'))
print('UPDATED _zero_streak_result.json with status')
