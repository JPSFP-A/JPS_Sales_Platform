# -*- coding: utf-8 -*-
# Customer x rate-class, Apr vs May 2026 — ONE ROW PER (customer, rate class).
# Multi-meter customers (e.g. Bay Roc) appear as separate rows: Bay Roc/RT40, Bay Roc/RT50.
# Exact requested format: Customer | Class | April kWh | May kWh | Var | April Demand | May Demand | Var | Δ Dem | Δ kVA
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
C=json.load(open('custclass.json'))
rows=[]
for key,r in C.items():
    apr=r.get('Apr',[0,0,0]); may=r.get('May',[0,0,0])
    aK,mK=apr[0]/1e6,may[0]/1e6        # GWh
    aD,mD=apr[1]/1e6,may[1]/1e6        # J$M (demand charge)
    if max(abs(aK),abs(mK),abs(aD),abs(mD))<1e-9: continue
    rows.append((r['name'],r['title'],aK,mK,aD,mD))
rows.sort(key=lambda x:-x[5])          # by May demand desc
NAVY='1F3864'
W=Workbook(); ws=W.active; ws.title='Customers'
ws['A1']='JPS — Customers by Rate Class · April vs May 2026'; ws['A1'].font=Font(name='Arial',size=13,bold=True,color=NAVY)
ws['A2']='One row per customer-meter (multi-meter customers split by class, e.g. Bay Roc Hotel RT40 + RT50) · kWh in GWh · Demand = demand charge J$M · Δ Dem = demand %  · Δ kVA = kWh %'
ws['A2'].font=Font(name='Arial',size=9,italic=True,color='808080')
hdr=['Customer','Class','April kWh','May kWh','Var','April Demand','May Demand','Var','Δ Dem','Δ kVA']
HF=Font(name='Arial',size=10,bold=True,color='FFFFFF'); HFill=PatternFill('solid',fgColor=NAVY)
for j,h in enumerate(hdr,1):
    c=ws.cell(4,j,h); c.font=HF; c.fill=HFill; c.alignment=Alignment(horizontal='center')
AR=Font(name='Arial',size=10)
for i,(nm,cls,aK,mK,aD,mD) in enumerate(rows):
    r=5+i
    ws.cell(r,1,nm).font=AR; ws.cell(r,2,cls).font=AR
    ws.cell(r,3,round(aK,3)); ws.cell(r,4,round(mK,3)); ws.cell(r,5,f'=D{r}-C{r}')
    ws.cell(r,6,round(aD,3)); ws.cell(r,7,round(mD,3)); ws.cell(r,8,f'=G{r}-F{r}')
    ws.cell(r,9,f'=IFERROR(G{r}/F{r}-1,"")'); ws.cell(r,10,f'=IFERROR(D{r}/C{r}-1,"")')
    for col in (3,4,5,6,7,8): ws.cell(r,col).number_format='#,##0.000;(#,##0.000)'
    for col in (9,10): ws.cell(r,col).number_format='0.0%'
    for col in range(3,11): ws.cell(r,col).font=AR
last=4+len(rows)
ws.cell(last+1,1,'TOTAL').font=Font(name='Arial',size=10,bold=True)
for col in (3,4,5,6,7,8):
    L=get_column_letter(col); cc=ws.cell(last+1,col,f'=SUM({L}5:{L}{last})'); cc.font=Font(name='Arial',size=10,bold=True); cc.number_format='#,##0.000;(#,##0.000)'
ws.column_dimensions['A'].width=32; ws.column_dimensions['B'].width=10
for col in 'CDEFGHIJ': ws.column_dimensions[col].width=12
ws.freeze_panes='C5'
W.save('JPS_Customers_by_RateClass_AprMay2026.xlsx')
print('wrote workbook | rows (customer x class):',len(rows))
br=[r for r in rows if 'bay roc' in r[0].lower()]
print('Bay Roc rows:',[(r[0],r[1],'AprkWh=%.3f'%r[2],'AprDem=%.3f'%r[4]) for r in br])
