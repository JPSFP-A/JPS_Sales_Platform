# Multi-month meter-level scan. Per month: rate-class aggregates (all classes) + industrial
# account-level (RT40/50/70). Full charge components. Saves incrementally to multi.json.
import openpyxl, json, time, os
FILES={
 '2025-05':'Billing Details Report May 2025.xlsx',
 '2025-10':'Billing Details Report - October-2025.xlsx',
 '2025-11':'Billing Details Report - November-2025.xlsx',
 '2025-12':'Billing Details Report - December-2025.xlsx',
 '2026-01':'Billing Details Report_Jan 2026.xlsx',
 '2026-02':'Billing Details Report_Feb 2026.xlsx',
 '2026-03':'Billing Details Report_Mar 2026.xlsx',
 '2026-04':'Billing Details Report_Apr 2026.xlsx',
 '2026-05':'Billing Details Report_May 2026.xlsx',
}
NUM=lambda v: float(v) if isinstance(v,(int,float)) else 0.0
def grp(c):
    c=str(c).upper()
    if c[:3] in ('RT1','NB1') or c[:2]=='FL': return 'RT10'
    if c[:3] in ('RT2','NB2','RT3'): return 'RT20'
    if c[:3] in ('RT4','NB4'): return 'RT40'
    if c[:3]=='RT5': return 'RT50'
    if c[:3]=='RT6' or c[:1]=='L': return 'RT60'
    if c[:3]=='RT7': return 'RT70'
    if c[:2]=='EV': return 'EV'
    return 'Other'
IND={'RT40','RT50','RT70'}
def proc(path):
    wb=openpyxl.load_workbook(path, read_only=True)
    detail=None
    for sn in wb.sheetnames:
        ws=wb[sn]
        for i,r in enumerate(ws.iter_rows(values_only=True)):
            if r and r[0]=='Cust_Code': detail=sn; break
            if i>8: break
        if detail: break
    ws=wb[detail]; it=ws.iter_rows(values_only=True)
    hdr=None
    for r in it:
        if r and r[0]=='Cust_Code': hdr=list(r); break
    I={n:k for k,n in enumerate(hdr) if n}
    g=lambda r,n: NUM(r[I[n]]) if n in I else 0.0
    en=lambda r: g(r,'KWHP_KWH_Energy')+g(r,'KWHL_Energy')+g(r,'KWHO_Energy')
    dm=lambda r: g(r,'KVAP_KVA_Demand')+g(r,'KVAL_Demand')+g(r,'KVAO_Demand')
    fu=lambda r: g(r,'fuel')+g(r,'FuelOffPeak')+g(r,'FuelPartialPeak')+g(r,'FuelOnPeak')
    cls={}; acc={}; n=0
    for r in it:
        code=r[I['Cust_Code']]
        if code in (None,'','Cust_Code'): continue
        n+=1
        gr=grp(r[I['Srat_Code']])
        kwh=g(r,'net_kwh_billed_consump'); rev=g(r,'net_revenue'); e=en(r); d=dm(r); f=fu(r)
        ip=g(r,'IPP_Charge'); cc=g(r,'Cust_Charge'); kv=g(r,'kva_billed_consump')
        a=cls.get(gr)
        if a is None: a=cls[gr]=[0]*9
        a[0]+=1; a[1]+=kwh; a[2]+=rev; a[3]+=e; a[4]+=d; a[5]+=f; a[6]+=ip; a[7]+=cc; a[8]+=kv
        if gr in IND:
            b=acc.get(code)
            if b is None: b=acc[code]={'name':(r[I['Name']] or '').strip(),'cls':{},'m':[0]*6}
            b['cls'][gr]=b['cls'].get(gr,0)+kwh
            mm=b['m']; mm[0]+=kwh; mm[1]+=rev; mm[2]+=d; mm[3]+=e; mm[4]+=f; mm[5]+=kv
    wb.close()
    return cls, acc, n

# load/init
M=json.load(open('multi.json')) if os.path.exists('multi.json') else {
  'months':[], 'legend_class':['count','kwh','rev','energy','demand','fuel','ipp','cust_chg','kva'],
  'legend_acct':['kwh','rev','demand','energy','fuel','kva'], 'class':{}, 'acct':{}}
for mo,f in FILES.items():
    if mo in M['months']:
        print('skip',mo,'(done)',flush=True); continue
    t=time.time(); print('processing',mo,flush=True)
    cls,acc,n=proc(f)
    M['class'][mo]={k:[round(x,2) for x in v] for k,v in cls.items()}
    for code,b in acc.items():
        dom=max(b['cls'],key=b['cls'].get)
        rec=M['acct'].setdefault(str(code),{'name':b['name'],'cls':dom,'m':{}})
        rec['name']=b['name']; rec['cls']=dom
        rec['m'][mo]=[round(x) for x in b['m']]
    M['months'].append(mo); M['months']=sorted(set(M['months']))
    json.dump(M, open('multi.json','w'))
    print('  done',mo,'rows',n,'classes',len(cls),'accts',len(acc),'in',round(time.time()-t),'s',flush=True)
print('WROTE multi.json · months',M['months'],flush=True)
