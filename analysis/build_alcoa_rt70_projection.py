# -*- coding: utf-8 -*-
# Alcoa (100185-607213) RT70 — 3-year revenue projection, sensitivity model.
# History pulled from jps_actuals for Jan 2025-Jul 2026 (19 months, via Supabase MCP)
# plus 12 months of 2024 (2026-08-12) sourced directly from raw "<Mon> 24.csv" CIS
# export files the user placed in this folder -- jps_actuals has zero 2024 rows for
# this account. 31 months total, hardcoded here as the model's source data (blue
# inputs, sourced). Multiplier (48000x) and current billed kVA (21,025.01) pulled
# from the raw July 2026 Billing Details Report (multilpier / kva_billed_consump
# columns) -- jps_actuals' kwh/revenue already have the multiplier applied; the
# Drivers-tab multiplier is an ADDITIONAL override for scenario testing, not a
# re-application.
#
# jps_actuals.fuel_jmd was $0 in all 19 2025/2026 months for this account (unlike
# every other RT40/50/70 account, all with fuel ~60-70% of revenue) -- Alcoa's own
# raw extracts never populate the Fuel/FuelOffPeak/FuelPartialPeak/FuelOnPeak
# columns, but the true fuel charge is embedded in Revenue: the Aug-2025
# reversal-of-July row itemizes real TOU fuel that matches July's unexplained
# residual almost exactly. Recomputed and backfilled fuel_jmd (Revenue-Demand-
# Energy-IPP-CustCharge-GCT-FEX) for the 13 of 19 2025/2026 months with a raw file
# available, plus all 12 months of 2024 using the same method -- 25 of 31 months
# now reconcile to ~96-99.98% of revenue via 6 named components. 2025 Jan/Feb/Mar/
# Apr/Jun/Sep have no raw file anywhere on this machine to verify, so fuel stays $0
# there and "Other/Base Tariff" (still tracked as its own component, so Total
# Revenue always reconciles) carries the full unexplained residual for those 6
# months only.
#
# Demand/Energy/IPP/Fuel/Other are all modeled as $/kWh rates derived from
# history and escalated by driver %, since no per-month kVA history is
# available for this specific account (jps_demand_actuals is rate-class-level,
# not per-account) -- documented as an assumption on the Drivers tab.
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

FONT = 'Arial'
BLUE = Font(name=FONT, color='0000FF', size=10)
BLACK = Font(name=FONT, color='000000', size=10)
GREEN = Font(name=FONT, color='008000', size=10)
BOLD = Font(name=FONT, color='000000', size=10, bold=True)
HDR = Font(name=FONT, color='FFFFFF', size=10, bold=True)
TITLE = Font(name=FONT, color='FFFFFF', size=13, bold=True)
YELLOW = PatternFill('solid', start_color='FFFF00')
NAVY = PatternFill('solid', start_color='0C3547')
GREY = PatternFill('solid', start_color='F0F4F8')
THIN = Side(style='thin', color='CBD5E1')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '$#,##0;($#,##0);"-"'
KWHFMT = '#,##0;(#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'
RATE = '$#,##0.0000;($#,##0.0000);"-"'

wb = Workbook()

# ============================================================ HISTORY ============
hs = wb.active
hs.title = 'History'
hs.sheet_view.showGridLines = False
for c, w in zip('ABCDEFGHIJKLMNOPQ', [6, 10, 12, 14, 15, 13, 13, 13, 14, 11, 15, 15, 15, 15, 15, 12, 13]):
    hs.column_dimensions[c].width = w
hs['A1'] = 'ALCOA MINS OF JA LTD — RT70 — HISTORICAL ACTUALS'
hs['A1'].font = TITLE
hs.merge_cells('A1:Q1')
hs['A1'].fill = NAVY
hs['A1'].alignment = Alignment(vertical='center', indent=1)
hs.row_dimensions[1].height = 22
hs['A2'] = '31 months, Jan-2024 to Jul-2026 · Account 100185-607213 · 2024: raw "<Mon> 24.csv" CIS export files (not yet in jps_actuals) · 2025-2026: jps_actuals (Supabase bhrswnbenkvflpdjhfpa) · GCT-exempt throughout · Fuel corrected/reclassified for 25 of 31 months (see note below) — 2025 Jan/Feb/Mar/Apr/Jun/Sep still show $0, raw file unavailable to verify · Nov/Dec-2025 flagged Hurricane, excluded from all baseline-rate averages'
hs['A2'].font = Font(name=FONT, italic=True, size=9, color='6B7A99')
hs.merge_cells('A2:Q2')

HDRS = ['Year', 'Month#', 'Period', 'kWh', 'Revenue $', 'Demand $', 'Fuel $', 'Energy $', 'IPP $',
        'Cust Chg $', 'GCT $', 'Demand $/kWh', 'Energy $/kWh', 'Other/Base $', 'Other $/kWh', 'Flag', 'KVA (true meter)']
r0 = 4
for j, h in enumerate(HDRS):
    c = hs.cell(r0, j + 1, h)
    c.font = HDR
    c.fill = NAVY
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = BORDER
hs.row_dimensions[r0].height = 28

# (year, month, kwh, revenue, demand, fuel, energy, ipp, custchg, gct) — None = null in DB
# 2024 rows sourced 2026-08-12 directly from the raw "<Mon> 24.csv" CIS export files
# (jps_actuals has ZERO 2024 rows for this account -- confirmed via SQL -- so these are
# NOT yet in the DB; sourced here straight from the raw file the same way Jul/Aug-2025
# were). Fuel = Revenue-Demand-Energy-IPP-CustChg-GCT (the same reclassification applied
# to the 13 already-corrected 2025/2026 months below), leaving only FEX/small adjustments
# in "Other/Base" -- confirmed for Jan-2024: the residual after that is exactly $54,814,
# matching FEX to the dollar.
ROWS = [
    (2024, 1, 10253257, 736497100, 157282591, 467074484, 88388289, 23678789, 18133, 0),
    (2024, 2, 9018952, 362135089, 100543308, 201366813, 44102677, 14230742, 9067, 0),
    (2024, 3, 3381218, 142539162, 39086809, 73419552, 16534155, 5271784, 9067, 0),
    (2024, 4, 4933429, 237984024, 97304304, 100637025, 24124469, 16098603, 9067, 0),
    (2024, 5, 5832221, 245723216, 70741601, 136054054, 28519561, 9412187, 9067, 0),
    (2024, 6, 5854692, 239081247, 61751914, 138188307, 28629446, 10120162, 9067, 0),
    (2024, 7, 6916870, 268805864, 61579422, 161591922, 33823495, 11163652, 9067, 0),
    (2024, 8, 6755643, 345715674, 88353918, 208648041, 33035095, 14085308, 9067, 0),
    (2024, 9, 5794081, 216034708, 59567023, 118471583, 28333058, 8279078, 9067, 0),
    (2024, 10, 8160455, 333407112, 92062482, 184238585, 39904623, 14670290, 9067, 0),
    (2024, 11, 8790584, 343026033, 74832510, 211554196, 42985956, 11128687, 9067, 0),
    (2024, 12, 9658757, 415877709, 99821400, 252933864, 47231321, 14168787, 9067, 0),
    (2025, 1, 7283162, 266092728, 47668312, 0, 35614660, 6356502, 9067, 0),
    (2025, 2, 7558914, 338431883, 98780063, 0, 36963087, 15370340, 9067, 0),
    (2025, 3, 9220246, 347997726, 62116062, 0, 45087002, 6840302, 9067, 0),
    (2025, 4, 6776457, 306339672, 99821400, 0, 33136874, 14980429, 9067, 0),
    (2025, 5, 6360004, 304760808, 96646281, 159584943, 31100421, 12528312, 9067, 0),
    (2025, 6, 6112234, 254803373, 58048083, 0, 29888823, 7104539, 9067, 0),
    (2025, 7, 7807153, 285797656, 56361244, 178924341, 38176979, 9708956, 9067, 0),
    (2025, 8, 6836301, 257804420, 57679201, 154623450, 33429511, 9124644, 9066, 0),
    (2025, 9, 6921867, 273820535, 55312236, 0, 33847930, 7948863, 9067, 0),
    (2025, 10, 7047477, 322020733, 84597439, 187744778, 34462161, 11382804, 9067, 0),
    (2025, 11, 3374336, 201961923, 81023034, 89564989, 16500501, 11610645, 9067, 0),
    (2025, 12, 1864266, 92527465, 26065364, 53396300, 9116260, 2813669, 9067, 0),
    (2026, 1, 3618112, 154329498, 36146527, 91292198, 17692567, 7872014, 9067, 0),
    (2026, 2, 6153753, 252683564, 38373999, 175634278, 30091855, 7675975, 9067, 0),
    (2026, 3, 8480800, 401358542, 75273321, 268739604, 41471114, 14501754, 9067, 0),
    (2026, 4, 8347849, 313688892, 51555757, 215182515, 40820984, 4411433, 9067, 0),
    (2026, 5, 9056910, 344070135, 50765228, 240814189, 44288292, 6643840, 9067, None),
    (2026, 6, 8590490, 388387586, 99821400, 228756157, 42007496, 15274424, 9067, 0),
    (2026, 7, 9004421, 347336115, 59964170, 225515413, 44031617, 9457813, 9067, None),
]
# True metered KVA (kva_billed_consump), pulled 2026-08-12 from the same raw files used
# for the Fuel fix -- NOT the same as demand_jmd / $2,852.04: that division only matches
# the true reading in about 60% of months (likely a demand ratchet/minimum-demand clause
# decouples billed $ from the current month's raw meter reading in the rest). Blank where
# no raw file exists (same 6-month 2025 gap as the Fuel fix). Jul/Aug-2025 netted the same
# way as their kwh/revenue (real bill + reversal).
KVA_TRUE = {
    (2024, 1): 55147.40, (2024, 2): 35000.00, (2024, 3): 11363.52, (2024, 4): 34117.44,
    (2024, 5): 24803.86, (2024, 6): 21651.84, (2024, 7): 21591.36, (2024, 8): 30979.20,
    (2024, 9): 20885.76, (2024, 10): 32279.52, (2024, 11): 37758.24, (2024, 12): 46520.00,
    (2025, 5): 49617.60, (2025, 7): 36082.00, (2025, 8): 41823.57,
    (2025, 10): 51262.08, (2025, 11): 28408.80, (2025, 12): 9139.20,
    (2026, 1): 12673.92, (2026, 2): 13454.93, (2026, 3): 26392.80, (2026, 4): 18076.80,
    (2026, 5): 17799.62, (2026, 6): 35000.00, (2026, 7): 21025.01,
}
MNAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
r = r0 + 1
first_data_row = r
for (y, m, kwh, rev, dem, fuel, en, ipp, cc, gct) in ROWS:
    vals = [y, m, MNAMES[m - 1] + "'" + str(y)[2:], kwh, rev, dem, fuel, en, ipp, cc, gct]
    for j, v in enumerate(vals):
        cell = hs.cell(r, j + 1, v)
        cell.font = BLUE
        cell.border = BORDER
        if j == 3:
            cell.number_format = KWHFMT
        elif j >= 4:
            cell.number_format = MONEY
        if j == 1:
            cell.alignment = Alignment(horizontal='center')
    dcell = hs.cell(r, 12)
    dcell.value = f'=IF(OR(F{r}="",D{r}=0),"",F{r}/D{r})'
    dcell.font = BLACK; dcell.number_format = RATE; dcell.border = BORDER
    ecell = hs.cell(r, 13)
    ecell.value = f'=IF(OR(H{r}="",D{r}=0),"",H{r}/D{r})'
    ecell.font = BLACK; ecell.number_format = RATE; ecell.border = BORDER
    # Other/Base $ = Revenue - (Demand+Fuel+Energy+IPP+CustChg+GCT). Blank only when the
    # 5 CORE components are missing (the Jul-2025 pre-migration gap) — GCT null is treated
    # as 0 (matches every other month; this account is GCT-exempt), not as missing data,
    # so a lone-null GCT month (May/Jul-2026) doesn't wrongly blank out Other too.
    ocell = hs.cell(r, 14)
    ocell.value = f'=IF(OR(F{r}="",G{r}="",H{r}="",I{r}="",J{r}=""),"",E{r}-F{r}-G{r}-H{r}-I{r}-J{r}-IF(K{r}="",0,K{r}))'
    ocell.font = BLACK; ocell.number_format = MONEY; ocell.border = BORDER
    oratecell = hs.cell(r, 15)
    oratecell.value = f'=IF(OR(N{r}="",D{r}=0),"",N{r}/D{r})'
    oratecell.font = BLACK; oratecell.number_format = RATE; oratecell.border = BORDER
    flag = 'Hurricane' if (y, m) in ((2025, 11), (2025, 12)) else 'Normal'
    fcell2 = hs.cell(r, 16, flag)
    fcell2.font = BLUE if flag == 'Hurricane' else BLACK
    fcell2.border = BORDER
    fcell2.alignment = Alignment(horizontal='center')
    kva_true = KVA_TRUE.get((y, m))
    kcell = hs.cell(r, 17, kva_true if kva_true is not None else '')
    kcell.font = BLUE if kva_true is not None else BLACK
    kcell.number_format = '#,##0.00'
    kcell.border = BORDER
    if flag == 'Hurricane':
        hcfill = PatternFill('solid', start_color='FDE7D0')
        for col in range(1, 18):
            hs.cell(r, col).fill = hcfill
    r += 1
last_data_row = r - 1
hs.cell(r, 3, 'AVERAGE (excl. Hurricane)').font = BOLD
for col in range(4, 16):
    cl = get_column_letter(col)
    fcell = hs.cell(r, col)
    fcell.value = f'=AVERAGEIF($P${first_data_row}:$P${last_data_row},"Normal",{cl}{first_data_row}:{cl}{last_data_row})'
    fcell.font = BOLD
    fcell.number_format = KWHFMT if col == 4 else (RATE if col in (12, 13, 15) else MONEY)
    fcell.border = BORDER
hs.freeze_panes = 'A5'
hs['A' + str(r + 2)] = ('Note: Jul/Aug-2025 corrected 2026-08-11. Jul-2025 (7,807,153 kWh / $285,797,656) was a normal, correctly-billed month, including its component breakout (Demand $56,361,244 / Energy $38,176,979 / IPP $9,708,956 / Cust Chg $9,067) -- '
                         'the raw file actually had TWO Jul rows for this account: the real bill (comma-formatted Cust_Code, cust_billed=1) and an unrelated all-zero row (clean code, cust_billed=0); jps_actuals had picked up kWh/revenue from the real row but lost the component detail to the same comma-parsing gap. '
                         'The Aug-2025 raw extract separately carried BOTH a positive 14,643,454 kWh / $543,602,076 charge AND an exact reversal of the Jul bill (-7,807,153 kWh / -$285,797,656, comma-formatted Cust_Code) -- '
                         'the reversal was silently dropped by the same bug, so jps_actuals kept Aug\'s raw positive figure un-netted, double-counting Jul\'s consumption on top of the real Jul bill. '
                         'Aug-2025 is now the net of both rows: 6,836,301 kWh / $257,804,420 (demand/energy/IPP/customer-charge components netted the same way; the customer-charge net of 9,066 vs. every other month\'s ~9,067 corroborates this reading). '
                         'FUEL FIX (2026-08-12): jps_actuals.fuel_jmd was $0 in all 19 months, unlike every other RT40/50/70 account (all showing fuel as 60-70% of revenue). Alcoa\'s own raw-file extracts NEVER populate the Fuel/FuelOffPeak/FuelPartialPeak/FuelOnPeak columns (confirmed across all 13 raw files checked) -- '
                         'but the true fuel charge IS embedded in Revenue: the Aug-2025 reversal-of-July row itemizes real TOU fuel (-$178,924,339) that exactly matches July\'s unexplained residual (off by $2). Recomputed fuel for every month with a locally available raw file as Revenue-Demand-Energy-IPP-CustCharge-GCT-FEX(-revenue_adj where present), '
                         'and updated jps_actuals + this model for 13 of 19 2025/2026 months (2025 May/Jul/Aug/Oct/Nov/Dec + all of 2026). 2025 Jan/Feb/Mar/Apr/Jun/Sep have no raw file available anywhere on this machine -- fuel is unverified and left at $0 for those 6 months; "Other/Base $" there still carries the full unexplained residual. '
                         '"Other/Base $" = Revenue - (Demand+Fuel+Energy+IPP+CustChg+GCT) — for the 25 corrected months this is now small (~0.02-3.5% of revenue, matching FEX/revenue_adj); for the 6 unverified 2025 months it remains large (~50-60% of revenue). '
                         '2024 ADDED (2026-08-12): all 12 months pulled directly from the raw "<Mon> 24.csv" CIS export files the user placed in this folder — jps_actuals has zero 2024 rows for this account, so these are sourced straight from the raw files, not the DB, using the same Fuel reclassification as the 2025/2026 fix above. Each month is a single clean row (no comma-parsing duplicates like Jul/Aug-2025). '
                         'HURRICANE FLAG (2026-08-12, user-confirmed): Nov/Dec-2025 (shaded) were storm-impacted months — kWh/revenue are left as actually billed (real numbers, not fabricated), but the AVERAGE row above and every rate/seasonality calc that feeds the projection now excludes these 2 rows via AVERAGEIF on the Flag column. With 2024 now providing real Nov/Dec observations, Seasonality no longer needs to interpolate those two calendar months. '
                         'KVA ADDED (2026-08-12): pulled from the true kva_billed_consump field in the raw files (NOT the same as Demand $ / $2,852.04 — that division matches the true meter reading in only ~60% of months, likely a demand ratchet/minimum-demand clause; e.g. Aug-2025\'s true net KVA is 41,823.57 vs 20,223.84 implied by simple division). Blank for the same 6 unverified 2025 months as the Fuel fix.')
hs['A' + str(r + 2)].font = Font(name=FONT, italic=True, size=8.5, color='B87800')
hs['A' + str(r + 2)].alignment = Alignment(wrap_text=True, vertical='top')
hs.merge_cells(f'A{r+2}:Q{r+2}')
hs.row_dimensions[r + 2].height = 130

AVG_ROW = r

# ============================================================ SEASONALITY ========
ss = wb.create_sheet('Seasonality')
ss.sheet_view.showGridLines = False
for c, w in zip('ABCDE', [6, 10, 16, 18, 12]):
    ss.column_dimensions[c].width = w
ss['A1'] = 'MONTHLY SEASONALITY INDEX (derived from History, all available years per calendar month, excl. Hurricane-flagged months)'
ss['A1'].font = TITLE
ss.merge_cells('A1:E1')
ss['A1'].fill = NAVY
ss.row_dimensions[1].height = 22
hdrs2 = ['Month#', 'Month', 'Avg Historical kWh', 'Avg Historical Revenue $', 'Seasonality Index']
for j, h in enumerate(hdrs2):
    c = ss.cell(3, j + 1, h); c.font = HDR; c.fill = NAVY; c.border = BORDER
    c.alignment = Alignment(horizontal='center', wrap_text=True)
ss.row_dimensions[3].height = 28
for i in range(12):
    rr = 4 + i
    monthnum = i + 1
    ss.cell(rr, 1, monthnum).font = BLACK
    ss.cell(rr, 1).alignment = Alignment(horizontal='center')
    ss.cell(rr, 2, MNAMES[i]).font = BLACK
    # 2024 now supplies a non-hurricane observation for every calendar month (including
    # Nov/Dec, whose only prior observation was the storm-hit 2025 row), so a plain
    # AVERAGEIFS excluding Hurricane-flagged rows works uniformly -- no interpolation needed.
    kc = ss.cell(rr, 3)
    kc.value = f'=AVERAGEIFS(History!$D${first_data_row}:$D${last_data_row},History!$B${first_data_row}:$B${last_data_row},A{rr},History!$P${first_data_row}:$P${last_data_row},"Normal")'
    kc.font = GREEN; kc.number_format = KWHFMT
    rc = ss.cell(rr, 4)
    rc.value = f'=AVERAGEIFS(History!$E${first_data_row}:$E${last_data_row},History!$B${first_data_row}:$B${last_data_row},A{rr},History!$P${first_data_row}:$P${last_data_row},"Normal")'
    rc.font = GREEN; rc.number_format = MONEY
    ic = ss.cell(rr, 5)
    ic.value = f'=C{rr}/AVERAGE($C$4:$C$15)'
    ic.font = BLACK; ic.number_format = '0.00"x"'
    for col in range(1, 6):
        ss.cell(rr, col).border = BORDER
ss.cell(16, 2, 'AVERAGE (=1.00x baseline)').font = BOLD
ss.cell(16, 3).value = '=AVERAGE(C4:C15)'; ss.cell(16, 3).font = BOLD; ss.cell(16, 3).number_format = KWHFMT
ss.cell(16, 4).value = '=AVERAGE(D4:D15)'; ss.cell(16, 4).font = BOLD; ss.cell(16, 4).number_format = MONEY
ss.cell(16, 5).value = '=AVERAGE(E4:E15)'; ss.cell(16, 5).font = BOLD; ss.cell(16, 5).number_format = '0.00"x"'
for col in range(1, 6):
    ss.cell(16, col).border = BORDER
ss['A18'] = 'All 12 months now average across 2-3 years (2024-2026) excluding Hurricane-flagged rows — Nov/Dec no longer need interpolation now that 2024 supplies a real, non-hurricane observation for both.'
ss['A18'].font = Font(name=FONT, italic=True, size=8.5, color='B87800')
ss.merge_cells('A18:E18')
ss.row_dimensions[18].height = 28
ss['A18'].alignment = Alignment(wrap_text=True, vertical='top')
ss.freeze_panes = 'A4'

# ============================================================ DRIVERS ============
ds = wb.create_sheet('Drivers')
ds.sheet_view.showGridLines = False
for c, w in zip('ABCDE', [42, 15, 42, 15, 40]):
    ds.column_dimensions[c].width = w
ds['A1'] = 'DRIVER / SENSITIVITY PANEL — toggle any blue cell to flex the projection'
ds['A1'].font = TITLE
ds.merge_cells('A1:E1')
ds['A1'].fill = NAVY
ds.row_dimensions[1].height = 22

def section(row, text):
    ds.cell(row, 1, text).font = Font(name=FONT, bold=True, size=10.5, color='0C3547')
    ds.merge_cells(f'A{row}:E{row}')
    for col in range(1, 6):
        ds.cell(row, col).fill = GREY

def drow(row, label, value, fmt, note='', input_cell=True):
    ds.cell(row, 1, label).font = BLACK
    c = ds.cell(row, 2, value)
    c.font = BLUE if input_cell else BLACK
    c.number_format = fmt
    c.fill = YELLOW if input_cell else PatternFill()
    c.border = BORDER
    ds.cell(row, 1).border = BORDER
    if note:
        ds.cell(row, 3, note).font = Font(name=FONT, italic=True, size=8.5, color='6B7A99')
        ds.merge_cells(f'C{row}:E{row}')

section(3, 'VOLUME SCENARIO')
drow(4, 'Volume adjustment vs. seasonal baseline (%)', -0.50, PCT, 'User scenario: 50% volume reduction, applied flat across all 36 projected months')
drow(5, 'Apply seasonality? (1 = Yes, 0 = No / use flat monthly average)', 1, '0', 'Set to 0 to test a flat (non-seasonal) run-rate instead')
drow(6, 'Meter multiplier override (x)', 1.00, '0.00"x"', 'jps_actuals kWh already reflects the current 48,000x CT/PT multiplier (see reference below) — this is an ADDITIONAL scenario multiplier on top, 1.00x = no change')

section(8, 'DEMAND & IPP-FIXED — driven directly by the given KVA schedule (Projection tab), user-confirmed 2026-08-12')
drow(9, 'Demand rate ($/KVA)', 2852.04, RATE, 'Applied to the given monthly KVA schedule, NOT derived from kWh/history')
drow(10, 'IPP Fixed rate ($/KVA, additional to IPP Variable below)', 424.14, RATE, 'Adds to — does not replace — the kWh-driven IPP Variable component')
drow(11, 'Demand & IPP-Fixed rate escalation (%/yr)', 0.00, PCT, 'Compounds on the two $/KVA rates above, from Year 1; the given schedule itself does not extend past the 36 projected months')

section(13, 'RATE ESCALATION — kWh-driven components (% per year, compounding from Year 1)')
drow(14, 'Energy rate escalation (%/yr)', 0.00, PCT)
drow(15, 'IPP Variable rate escalation (%/yr)', 0.00, PCT)
drow(16, 'Other/base tariff rate escalation (%/yr)', 0.00, PCT, 'Covers the residual not broken into the named components — see modeling note below')
drow(17, 'Fuel rate escalation (%/yr)', 0.00, PCT, 'Applied on top of the $25.628/kWh rate below')
drow(18, 'Fuel rate override ($/kWh, replaces history if >0)', 25.628, RATE, "User-provided 2-yr average (2026-08-12) — overrides the history-derived rate (Reference row 34)")
drow(19, 'Customer charge escalation (%/yr)', 0.00, PCT, 'Applied to the flat monthly customer charge, not volume-driven')

section(21, 'OTHER')
drow(22, 'GCT rate applied (%)', 0.00, PCT, 'Account has been GCT-exempt in all 19 historical months')
drow(23, 'Projection start (first row of Projection tab)', '2024-01', '@', 'Format YYYY-MM — Jan-2024, per the user\'s 2026-08-12 request that Projection show actuals from 2024 onward')

section(25, 'REFERENCE (informational, not scenario inputs)')
drow(26, 'Current meter multiplier', 48000, '#,##0"x"', 'Source: Billing Details Report_Jul 2026.xls, multilpier field, account 100185-607213', input_cell=False)
ds.cell(26, 2).font = BLACK
drow(27, 'Current billed demand (kVA, Jul-2026)', 21025.01, '#,##0.00', 'Source: Billing Details Report_Jul 2026.xls, kva_billed_consump field', input_cell=False)
ds.cell(27, 2).font = BLACK

section(29, 'HISTORICAL BASELINE UNIT RATES (computed from History tab averages, excl. Hurricane — feed the kWh-driven components only)')
def rate_row(row, label, num_col, den_col='D'):
    ds.cell(row, 1, label).font = BLACK
    ds.cell(row, 2).value = f'=History!{num_col}{AVG_ROW}/History!{den_col}{AVG_ROW}'
    ds.cell(row, 2).font = GREEN; ds.cell(row, 2).number_format = RATE
    ds.cell(row, 2).border = BORDER; ds.cell(row, 1).border = BORDER

rate_row(30, 'Avg Demand $/kWh (History col F/D) — reference only, NOT used (Demand now driven by KVA schedule)', 'F')
rate_row(31, 'Avg Energy $/kWh (History col H/D)', 'H')
rate_row(32, 'Avg IPP $/kWh (History col I/D) — reference only, NOT used (IPP Variable derivation below is separate)', 'I')
rate_row(33, 'Avg Other/Base $/kWh (History col N/D)', 'N')
rate_row(34, 'Avg Fuel $/kWh (History col G/D) — reference only, overridden by row 18 above', 'G')
ds.cell(35, 1, 'Avg Customer Charge $/mo (History col J avg)').font = BLACK
ds.cell(35, 2).value = f'=History!J{AVG_ROW}'
ds.cell(35, 2).font = GREEN; ds.cell(35, 2).number_format = MONEY
ds.cell(35, 2).border = BORDER; ds.cell(35, 1).border = BORDER
ds.cell(36, 1, 'IPP Variable $/kWh (= Avg IPP $/kWh above, kept as its own driven line)').font = BLACK
ds.cell(36, 2).value = '=B32'
ds.cell(36, 2).font = BLACK; ds.cell(36, 2).number_format = RATE
ds.cell(36, 2).border = BORDER; ds.cell(36, 1).border = BORDER

ds['A38'] = ('MODELING NOTES: (1) Demand and IPP-Fixed are driven DIRECTLY off the given KVA schedule (Projection tab) at the $/KVA rates above — not derived from kWh or history, per the 2026-08-12 update. '
             '(2) IPP Variable (kWh-based, derived from history) still applies on top of IPP Fixed — the two are additive, not alternatives. '
             '(3) Fuel uses the user-provided $25.628/kWh 2-yr average (row 18) rather than the history-derived rate (row 34). '
             '(4) Nov/Dec-2025 (hurricane) are excluded from every average feeding this tab and Seasonality — see History/Seasonality tab notes for how those two months were normalized. '
             '(5) "Other/Base" still carries fuel that never made it into jps_actuals\' Fuel column for 6 of 31 months (2025 Jan/Feb/Mar/Apr/Jun/Sep, no raw file to verify) — see History tab note; it is NOT part of the KVA-driven Demand/IPP-Fixed calculation.')
ds['A38'].font = Font(name=FONT, italic=True, size=8.5, color='B87800')
ds.merge_cells('A38:E38')
ds.row_dimensions[38].height = 72
ds['A38'].alignment = Alignment(wrap_text=True, vertical='top')

# ============================================================ PROJECTION =========
# 60 months, Jan-2024 to Dec-2028, per the user's 2026-08-12 request: one continuous
# timeline showing ACTUALS where we have actuals (Jan-2024 to Jul-2026, 31 months,
# pulled straight from History row-for-row, not re-derived) and the model-projected
# figures for the rest (Aug-2026 to Dec-2028, 29 months, using the given KVA schedule
# + kWh-seasonality drivers). An A/P flag column marks which is which.
#
# Given future KVA schedule covers Aug-2025 to Jul-2028 (user-provided 2026-08-12);
# Aug-2025 to Jul-2026 is now superseded by actuals (shown via History, not this
# schedule). Aug-2028 to Dec-2028 (5 months) has no user-provided figure at all --
# extended flat at 25 KVA, matching the established Jan-2027/Jul-2028 pattern, and
# clearly flagged below as an assumption, not given data.
PROJECTED_KVA = [29556.91, 29556.91, 29556.91, 29556.91, 15263.1] + [25] * 24  # Aug-2026 to Dec-2028

ACTUAL_ROWS = 31   # Jan-2024 to Jul-2026
PROJ_ROWS = 29     # Aug-2026 to Dec-2028
TOTAL_ROWS = ACTUAL_ROWS + PROJ_ROWS  # 60

ps = wb.create_sheet('Projection')
ps.sheet_view.showGridLines = False
PW = [6, 6, 10, 10, 10, 13, 15, 14, 15, 13, 13, 13, 13, 13, 15, 14, 15, 13, 16]
for c, w in zip('ABCDEFGHIJKLMNOPQRS', PW):
    ps.column_dimensions[c].width = w
ps['A1'] = 'PROJECTION — ALCOA RT70 (60 months, Jan-2024 to Dec-2028; Actual through Jul-2026, then modeled)'
ps['A1'].font = TITLE
ps.merge_cells('A1:S1')
ps['A1'].fill = NAVY
ps.row_dimensions[1].height = 22
hdrs3 = ['Per#', 'CY', 'Month', 'Cal Mo#', 'A/P', 'KVA', 'Seasonal kWh', 'Volume-Adj kWh', 'Final kWh\n(x multiplier)',
         'Demand $', 'Energy $', 'IPP Variable $', 'IPP Fixed $', 'Fuel $', 'Other/Base $', 'Cust Chg $', 'Subtotal $', 'GCT $', 'Total Revenue $']
for j, h in enumerate(hdrs3):
    c = ps.cell(3, j + 1, h); c.font = HDR; c.fill = NAVY; c.border = BORDER
    c.alignment = Alignment(horizontal='center', wrap_text=True)
ps.row_dimensions[3].height = 30

pr0 = 4
for i in range(TOTAL_ROWS):
    rr = pr0 + i
    is_actual = i < ACTUAL_ROWS
    ps.cell(rr, 1, i + 1).font = BLACK
    ps.cell(rr, 2).value = f'=YEAR(C{rr})'
    ps.cell(rr, 2).font = BLACK
    ps.cell(rr, 2).alignment = Alignment(horizontal='center')
    ps.cell(rr, 3).value = f'=EDATE(DATEVALUE(Drivers!$B$23&"-01"),{i})'
    ps.cell(rr, 3).number_format = 'mmm-yy'
    ps.cell(rr, 3).font = BLACK
    ps.cell(rr, 4).value = f'=MONTH(C{rr})'
    ps.cell(rr, 4).font = BLACK
    ps.cell(rr, 4).alignment = Alignment(horizontal='center')
    apcell = ps.cell(rr, 5, 'Actual' if is_actual else 'Projected')
    apcell.font = BLACK
    apcell.alignment = Alignment(horizontal='center')

    if is_actual:
        hr = first_data_row + i  # History row for this same month, 1:1 aligned
        ps.cell(rr, 6).value = f'=History!Q{hr}'
        ps.cell(rr, 6).font = GREEN; ps.cell(rr, 6).number_format = '#,##0.00'
        ps.cell(rr, 7).value = ''  # Seasonal kWh n/a for actual months
        ps.cell(rr, 8).value = ''  # Volume-Adj kWh n/a for actual months
        ps.cell(rr, 9).value = f'=History!D{hr}'
        ps.cell(rr, 9).font = GREEN; ps.cell(rr, 9).number_format = KWHFMT
        ps.cell(rr, 10).value = f'=History!F{hr}'
        ps.cell(rr, 10).font = GREEN; ps.cell(rr, 10).number_format = MONEY
        ps.cell(rr, 11).value = f'=History!H{hr}'
        ps.cell(rr, 11).font = GREEN; ps.cell(rr, 11).number_format = MONEY
        ps.cell(rr, 12).value = f'=History!I{hr}'
        ps.cell(rr, 12).font = GREEN; ps.cell(rr, 12).number_format = MONEY
        ps.cell(rr, 13).value = 0  # IPP Fixed not itemized historically
        ps.cell(rr, 13).font = BLACK; ps.cell(rr, 13).number_format = MONEY
        ps.cell(rr, 14).value = f'=History!G{hr}'
        ps.cell(rr, 14).font = GREEN; ps.cell(rr, 14).number_format = MONEY
        ps.cell(rr, 15).value = f'=History!N{hr}'
        ps.cell(rr, 15).font = GREEN; ps.cell(rr, 15).number_format = MONEY
        ps.cell(rr, 16).value = f'=History!J{hr}'
        ps.cell(rr, 16).font = GREEN; ps.cell(rr, 16).number_format = MONEY
        ps.cell(rr, 17).value = f'=SUM(J{rr}:P{rr})'
        ps.cell(rr, 17).font = BOLD; ps.cell(rr, 17).number_format = MONEY
        ps.cell(rr, 18).value = f'=History!K{hr}'
        ps.cell(rr, 18).font = GREEN; ps.cell(rr, 18).number_format = MONEY
        ps.cell(rr, 19).value = f'=History!E{hr}'
        ps.cell(rr, 19).font = GREEN; ps.cell(rr, 19).number_format = MONEY
    else:
        pk = i - ACTUAL_ROWS
        kcell = ps.cell(rr, 6, PROJECTED_KVA[pk])
        kcell.font = BLUE; kcell.number_format = '#,##0.00'
        ps.cell(rr, 7).value = f'=IF(Drivers!$B$5=1,INDEX(Seasonality!$C$4:$C$15,MATCH(D{rr},Seasonality!$A$4:$A$15,0)),Seasonality!$C$16)'
        ps.cell(rr, 7).font = GREEN; ps.cell(rr, 7).number_format = KWHFMT
        ps.cell(rr, 8).value = f'=G{rr}*(1+Drivers!$B$4)'
        ps.cell(rr, 8).font = BLACK; ps.cell(rr, 8).number_format = KWHFMT
        ps.cell(rr, 9).value = f'=H{rr}*Drivers!$B$6'
        ps.cell(rr, 9).font = BLACK; ps.cell(rr, 9).number_format = KWHFMT
        yexp = f'(ROUNDUP((A{rr}-{ACTUAL_ROWS+1})/12,0)-1)'
        ps.cell(rr, 10).value = f'=$F{rr}*Drivers!$B$9*(1+Drivers!$B$11)^{yexp}'
        ps.cell(rr, 10).font = BLACK; ps.cell(rr, 10).number_format = MONEY
        ps.cell(rr, 11).value = f'=$I{rr}*Drivers!$B$31*(1+Drivers!$B$14)^{yexp}'
        ps.cell(rr, 11).font = BLACK; ps.cell(rr, 11).number_format = MONEY
        ps.cell(rr, 12).value = f'=$I{rr}*Drivers!$B$36*(1+Drivers!$B$15)^{yexp}'
        ps.cell(rr, 12).font = BLACK; ps.cell(rr, 12).number_format = MONEY
        ps.cell(rr, 13).value = f'=$F{rr}*Drivers!$B$10*(1+Drivers!$B$11)^{yexp}'
        ps.cell(rr, 13).font = BLACK; ps.cell(rr, 13).number_format = MONEY
        ps.cell(rr, 14).value = f'=$I{rr}*IF(Drivers!$B$18>0,Drivers!$B$18,Drivers!$B$34)*(1+Drivers!$B$17)^{yexp}'
        ps.cell(rr, 14).font = BLACK; ps.cell(rr, 14).number_format = MONEY
        ps.cell(rr, 15).value = f'=$I{rr}*Drivers!$B$33*(1+Drivers!$B$16)^{yexp}'
        ps.cell(rr, 15).font = BLACK; ps.cell(rr, 15).number_format = MONEY
        ps.cell(rr, 16).value = f'=Drivers!$B$35*(1+Drivers!$B$19)^{yexp}'
        ps.cell(rr, 16).font = BLACK; ps.cell(rr, 16).number_format = MONEY
        ps.cell(rr, 17).value = f'=SUM(J{rr}:P{rr})'
        ps.cell(rr, 17).font = BOLD; ps.cell(rr, 17).number_format = MONEY
        ps.cell(rr, 18).value = f'=Q{rr}*Drivers!$B$22'
        ps.cell(rr, 18).font = BLACK; ps.cell(rr, 18).number_format = MONEY
        ps.cell(rr, 19).value = f'=Q{rr}+R{rr}'
        ps.cell(rr, 19).font = BOLD; ps.cell(rr, 19).number_format = MONEY

    for col in range(1, 20):
        ps.cell(rr, col).border = BORDER
    band = None
    y_i, m_i = (2024 + i // 12, i % 12 + 1)
    if (y_i, m_i) in ((2025, 11), (2025, 12)):
        band = PatternFill('solid', start_color='FDE7D0')
    elif not is_actual and y_i % 2 == 0:
        band = PatternFill('solid', start_color='F8FAFC')
    elif is_actual and y_i % 2 == 1:
        band = PatternFill('solid', start_color='EFF6FF')
    if band:
        for col in range(1, 20):
            ps.cell(rr, col).fill = band
ps.freeze_panes = 'C4'
last_proj_row = pr0 + TOTAL_ROWS - 1
ps['A62'] = ('Jan-2024 to Jul-2026 (31 rows, "Actual") are pulled directly from the History tab, row for row — not re-derived through the seasonality/KVA-rate machinery, so they exactly match actual billing. IPP Fixed is $0 for these rows since it wasn\'t itemized historically. '
             'Aug-2026 to Dec-2028 (29 rows, "Projected") use the given/extended KVA schedule for Demand $ and IPP Fixed $, and the Drivers-tab rates for everything kWh-driven. '
             'Aug-2028 to Dec-2028 (5 months) has no user-provided KVA figure — extended flat at 25 KVA to match the established 2027/Jan-Jul-2028 pattern; treat as an assumption, not given data. '
             'Nov/Dec-2025 (shaded, still Actual) show real billed figures; their KVA is unaffected by the hurricane, but note the Seasonality tab (which only feeds the Projected rows) excludes these two months from its baseline averages.')
ps['A62'].font = Font(name=FONT, italic=True, size=8.5, color='B87800')
ps.merge_cells('A62:S62')
ps.row_dimensions[62].height = 56
ps['A62'].alignment = Alignment(wrap_text=True, vertical='top')

# ============================================================ SUMMARY ============
sm = wb.create_sheet('Summary')
sm.sheet_view.showGridLines = False
for c, w in zip('ABCDEFGHIJKL', [28, 13, 13, 14, 14, 14, 14, 14, 14, 14, 12, 16]):
    sm.column_dimensions[c].width = w
sm['A1'] = 'SUMMARY — ANNUAL ROLLUP & ACTUAL COMPARISON'
sm['A1'].font = TITLE
sm.merge_cells('A1:L1')
sm['A1'].fill = NAVY
sm.row_dimensions[1].height = 22
sm['A2'] = 'Alcoa Mins Of Ja Ltd · Account 100185-607213 · RT70 · Scenario: see Drivers tab'
sm['A2'].font = Font(name=FONT, italic=True, size=9, color='6B7A99')
sm.merge_cells('A2:L2')

hdrs4 = ['', 'Avg KVA', 'kWh', 'Demand $', 'Energy $', 'IPP Var $', 'IPP Fixed $', 'Fuel $', 'Other/Base $', 'Cust Chg $', 'GCT $', 'Total Revenue $']
for j, h in enumerate(hdrs4):
    c = sm.cell(4, j + 1, h); c.font = HDR; c.fill = NAVY; c.border = BORDER
    c.alignment = Alignment(horizontal='center', wrap_text=True)
sm.row_dimensions[4].height = 26

# Calendar-year rollup — one row per year, SUMIF straight off the Projection tab, which
# now spans Jan-2024 to Dec-2028 as one continuous timeline (Actual through Jul-2026,
# Projected after). A year straddling the Actual/Projected boundary (2026) blends both
# automatically -- that's intentional, it's the real full-year total either way.
year_labels = {
    2024: 'CY2024 (all Actual)', 2025: 'CY2025 (all Actual)',
    2026: 'CY2026 (Jan-Jul Actual + Aug-Dec Projected, blended)',
    2027: 'CY2027 (all Projected)', 2028: 'CY2028 (all Projected)',
}
for k, yr in enumerate([2024, 2025, 2026, 2027, 2028]):
    rr = 5 + k
    sm.cell(rr, 1, year_labels[yr]).font = BOLD
    sm.cell(rr, 2).value = f'=AVERAGEIF(Projection!$B${pr0}:$B${last_proj_row},{yr},Projection!$F${pr0}:$F${last_proj_row})'
    sm.cell(rr, 3).value = f'=SUMIF(Projection!$B${pr0}:$B${last_proj_row},{yr},Projection!$I${pr0}:$I${last_proj_row})'
    sm.cell(rr, 4).value = f'=SUMIF(Projection!$B${pr0}:$B${last_proj_row},{yr},Projection!$J${pr0}:$J${last_proj_row})'
    sm.cell(rr, 5).value = f'=SUMIF(Projection!$B${pr0}:$B${last_proj_row},{yr},Projection!$K${pr0}:$K${last_proj_row})'
    sm.cell(rr, 6).value = f'=SUMIF(Projection!$B${pr0}:$B${last_proj_row},{yr},Projection!$L${pr0}:$L${last_proj_row})'
    sm.cell(rr, 7).value = f'=SUMIF(Projection!$B${pr0}:$B${last_proj_row},{yr},Projection!$M${pr0}:$M${last_proj_row})'
    sm.cell(rr, 8).value = f'=SUMIF(Projection!$B${pr0}:$B${last_proj_row},{yr},Projection!$N${pr0}:$N${last_proj_row})'
    sm.cell(rr, 9).value = f'=SUMIF(Projection!$B${pr0}:$B${last_proj_row},{yr},Projection!$O${pr0}:$O${last_proj_row})'
    sm.cell(rr, 10).value = f'=SUMIF(Projection!$B${pr0}:$B${last_proj_row},{yr},Projection!$P${pr0}:$P${last_proj_row})'
    sm.cell(rr, 11).value = f'=SUMIF(Projection!$B${pr0}:$B${last_proj_row},{yr},Projection!$R${pr0}:$R${last_proj_row})'
    sm.cell(rr, 12).value = f'=SUMIF(Projection!$B${pr0}:$B${last_proj_row},{yr},Projection!$S${pr0}:$S${last_proj_row})'
    for col in range(2, 13):
        sm.cell(rr, col).font = GREEN
        sm.cell(rr, col).number_format = KWHFMT if col in (2, 3) else MONEY

for row in range(5, 10):
    for col in range(1, 13):
        sm.cell(row, col).border = BORDER

sm['A12'] = 'CY2026\'s Avg KVA blends 7 real billed months (Jan-Jul) with 5 given/assumed months (Aug-Dec) — not a single consistent basis, shown for reference only. Every $ column reconciles exactly to Projection: Actual rows (Jan-2024/Jul-2026) are pulled straight from History, not re-derived; Projected rows (Aug-2026/Dec-2028) use the given KVA schedule and the $2,852.04/KVA, $424.14/KVA and $25.628/kWh rates (confirmed with the user 2026-08-12).'
sm['A12'].font = Font(name=FONT, italic=True, size=8.5, color='B87800')
sm.merge_cells('A12:L12')
sm.row_dimensions[12].height = 40
sm['A12'].alignment = Alignment(wrap_text=True, vertical='top')

sm['A14'] = 'kWh & Total Revenue by month — full Jan-2024 to Dec-2028 timeline (see chart)'
sm['A14'].font = BOLD

chart1 = LineChart()
chart1.title = 'Total Revenue by Month, Jan-2024 to Dec-2028 (J$) — Actual through Jul-2026, then Projected'
chart1.y_axis.title = 'J$'
chart1.x_axis.title = 'Period'
chart1.height, chart1.width = 9, 24
data = Reference(ps, min_col=19, min_row=3, max_row=last_proj_row)
cats = Reference(ps, min_col=3, min_row=pr0, max_row=last_proj_row)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
sm.add_chart(chart1, 'A16')

chart2 = LineChart()
chart2.title = 'KVA vs Final kWh by Month — Actual through Jul-2026, then Projected'
chart2.y_axis.title = 'KVA / kWh'
chart2.x_axis.title = 'Period'
chart2.height, chart2.width = 9, 24
data2 = Reference(ps, min_col=6, min_row=3, max_row=last_proj_row)
data2b = Reference(ps, min_col=9, min_row=3, max_row=last_proj_row)
chart2.add_data(data2, titles_from_data=True)
chart2.add_data(data2b, titles_from_data=True)
chart2.set_categories(cats)
sm.add_chart(chart2, 'A33')

wb._sheets = [wb['Drivers'], wb['History'], wb['Seasonality'], wb['Projection'], wb['Summary']]

import sys
OUT = sys.argv[1] if len(sys.argv) > 1 else 'Alcoa_RT70_3Year_Projection.xlsx'
wb.save(OUT)
print('WROTE', OUT)
