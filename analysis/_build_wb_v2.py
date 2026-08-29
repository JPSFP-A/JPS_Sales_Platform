# -*- coding: utf-8 -*-
import openpyxl, numpy as np
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

src = r'C:\Users\jwilson\Downloads\July LE Sales Gen.xlsx'
sv = openpyxl.load_workbook(src, data_only=True)
vs, dv = sv['Summary'], sv['Demand']
CLS = ['RT10','RT20','RT40','RT50','RT60-ST','RT70']

def cell(ws, r, c):
    v = ws.cell(r, c).value
    return float(v) if isinstance(v, (int, float)) else 0.0

# FY2026 monthly by class, read straight off the Driver Forecast engine after the
# submitted true-up overlay was applied. Previously this came from the July LE
# workbook and was then pinned to 3,281.97 here -- which meant the workbook tied to
# the submitted total while the app did not, and their class splits never matched.
# The engine now ties on its own, so there is nothing left to pin.
m26 = {
    'RT10': [
        84427346,
        76489141,
        88277732,
        87175653,
        99081407,
        101299278,
        113612986,
        112884872,
        110316970,
        113999942,
        111375592,
        110279292
    ],
    'RT20': [
        44388714,
        41872696,
        49678165,
        47362312,
        55463130,
        55124575,
        60644232,
        61560143,
        59678391,
        61560968,
        60377505,
        59663474
    ],
    'RT40': [
        56013982,
        52824365,
        61689786,
        59597985,
        65913616,
        64392244,
        70139337,
        68428481,
        65769716,
        67836252,
        67332299,
        67057448
    ],
    'RT50': [
        26723901,
        26712470,
        29779870,
        28362025,
        30151229,
        28210033,
        28679193,
        29650671,
        28115024,
        28821339,
        28917253,
        28812031
    ],
    'RT60-ST': [
        2224345,
        2309941,
        3098902,
        3413243,
        2798277,
        3409919,
        3409214,
        3300094,
        3304540,
        3307233,
        3282753,
        3279533
    ],
    'RT70': [
        20371321,
        21088523,
        26193990,
        24727332,
        27726514,
        26848147,
        28562426,
        19772739,
        18432359,
        18982266,
        17937598,
        17664223
    ]
}
l26 = [cell(vs, 15, c) for c in range(2, 14)]

# Rate-class totals read straight off the Driver Forecast engine (29 Aug), after the
# storm normalisation was scaled into the engine and Petrojam was re-solved. These are
# the engine's own numbers -- do not hand-edit them here, change the assumptions and
# re-read, or the workbook and the app start telling different stories.
T = {2027: {'RT10':1262.50,'RT20':693.81,'RT40':827.11,'RT50':393.83,'RT60-ST':39.18,'RT70':188.91},
     2028: {'RT10':1298.15,'RT20':703.06,'RT40':819.28,'RT50':425.81,'RT60-ST':38.20,'RT70':179.01}}

def shape(rc):
    v = m26[rc]
    if rc == 'RT70':
        avg = sum(v[7:]) / 5.0
        h1 = sum(v[:7]) / 7.0
        b = [avg * (v[i] / h1 if i < 7 else 1.0) for i in range(12)]
        s = sum(b)
        return [x / s for x in b]
    s = sum(v)
    return [x / s for x in v]

series = {rc: {2026: m26[rc],
               2027: [T[2027][rc] * 1e6 * s for s in shape(rc)],
               2028: [T[2028][rc] * 1e6 * s for s in shape(rc)]} for rc in CLS}
sales = np.array([sum(series[rc][y][m] for rc in CLS) for y in (2026, 2027, 2028) for m in range(12)])

# FY2026 is the submitted number of record. Jan-Jul are closed actuals and are never
# touched; the small residual against the submitted total is absorbed across the
# Aug-Dec forecast months only, pro rata to their size.
TARGET_2026 = 3281.97e6
cur26 = sum(sales[:12])
gap = TARGET_2026 - cur26
# The engine already carries the submitted true-up, so this should now be ~zero. It is
# kept as an assertion rather than an adjustment: if the engine ever drifts off the
# submitted total again, the build should fail loudly, not quietly paper over it.
assert abs(gap) < 5e3, 'FY2026 is %.3f GWh, off the submitted 3,281.97 by %.1f MWh' % (cur26/1e6, gap/1e3)
print('FY2026 ties to the submitted total: %.3f GWh (gap %.1f MWh)' % (cur26/1e6, gap/1e3))

roll = np.array([26.38,26.62,26.81,26.83,26.70,26.71,26.56,26.55,26.58,27.18,27.08,27.10] * 2) / 100.0
TG, n = 0.2710, 24
A = np.zeros((24, n)); b = np.zeros(24)
for k in range(24):
    t = 12 + k; r = roll[k]
    b[k] = r * sales[t-11:t+1].sum() / (1 - r)
    for i in range(t-11, t+1):
        if i < 12: b[k] -= l26[i]
        else: A[k, i-12] = sales[i]
sc = abs(b).mean(); A /= sc; b /= sc
H = np.zeros((2, n)); hb = np.zeros(2)
for j, off in enumerate((12, 24)):
    S = sales[off:off+12].sum()
    hb[j] = (TG * S / (1 - TG)) / sc
    for i in range(off, off+12): H[j, i-12] = sales[i] / sc
D = np.zeros((n, n)); x0 = l26[11] / sales[11]
for i in range(n):
    D[i, i] = 1.0
    if i > 0: D[i, i-1] = -1.0
d0 = np.zeros(n); d0[0] = x0
x, *_ = np.linalg.lstsq(np.vstack([A, 1e4*H, 1.0*D]), np.concatenate([b, 1e4*hb, 1.0*d0]), rcond=None)
full = np.concatenate([np.array(l26), x * sales[12:]])
pct = [full[t] / (sales[t] + full[t]) * 100 for t in range(36)]
rchk = [full[t-11:t+1].sum() / (sales[t-11:t+1].sum() + full[t-11:t+1].sum()) * 100 for t in range(12, 36)]
ng = [sales[t] + full[t] for t in range(36)]

dst = r'C:\Projects\Sales_Platform\JPS_LE_Sales_Gen_FY2026-28.xlsx'
wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Sales Wide'
HDR = PatternFill('solid', fgColor='0B3D66'); W = Font(bold=True, color='FFFFFF')
YR = PatternFill('solid', fgColor='E8EEF4')
months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

ws.cell(1,1).value = 'JPS Sales Forecast - FY2026 to FY2028, monthly kWh'
ws.cell(1,1).font = Font(bold=True, size=13)
ws.cell(2,1).value = ('Actual Jan-Jul 2026; forecast thereafter. Storm recovery now runs THROUGH the driver engine '
                      '(normalisation on RT50/RT40/RT20, cement excluded, industrial nil) rather than as a separate overlay, so these '
                      'figures equal what the platform shows. FY2028 carries no storm uplift: recovery completes by end-2027. '
                      'Losses solved so each fiscal year closes at 27.10%. FY2026 is pinned to the 3,281.97 GWh submitted on '
                      '19 Aug 2026 and now ties to it exactly in the app as well, via a disclosed -460.8 MWh true-up overlay on Aug-Dec.')
ws.cell(2,1).font = Font(italic=True, size=9)
col = 2
for y in (2026, 2027, 2028):
    ws.cell(4,col).value = 'FY%d' % y
    ws.cell(4,col).font = Font(bold=True, size=11)
    ws.cell(4,col).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+12)
    for i, mn in enumerate(months):
        c = ws.cell(5, col+i); c.value = mn; c.fill = HDR; c.font = W
        c.alignment = Alignment(horizontal='center')
    t = ws.cell(5, col+12); t.value = 'FY%d Total' % y; t.fill = HDR; t.font = W
    col += 13
ws.cell(5,1).value = 'Rate Class'; ws.cell(5,1).fill = HDR; ws.cell(5,1).font = W

tot = {y: [0.0]*12 for y in (2026, 2027, 2028)}
for i, rc in enumerate(CLS):
    r = 6 + i
    ws.cell(r,1).value = rc; ws.cell(r,1).font = Font(bold=True)
    col = 2
    for y in (2026, 2027, 2028):
        for m in range(12):
            v = series[rc][y][m]; tot[y][m] += v
            c = ws.cell(r, col+m); c.value = round(v, 0); c.number_format = '#,##0'
        tc = ws.cell(r, col+12); tc.value = round(sum(series[rc][y]), 0)
        tc.number_format = '#,##0'; tc.font = Font(bold=True); tc.fill = YR
        col += 13
gr = 6 + len(CLS)
ws.cell(gr,1).value = 'Grand Total'; ws.cell(gr,1).font = Font(bold=True)
col = 2
for y in (2026, 2027, 2028):
    for m in range(12):
        c = ws.cell(gr, col+m); c.value = round(tot[y][m], 0)
        c.number_format = '#,##0'; c.font = Font(bold=True)
    t = ws.cell(gr, col+12); t.value = round(sum(tot[y]), 0)
    t.number_format = '#,##0'; t.font = Font(bold=True); t.fill = YR
    col += 13

ws.cell(gr+1,1).value = 'Losses % (monthly)'; ws.cell(gr+1,1).font = Font(bold=True)
ws.cell(gr+2,1).value = 'Rolling 12-mth loss %'; ws.cell(gr+2,1).font = Font(bold=True)
ws.cell(gr+3,1).value = 'Net Generation'; ws.cell(gr+3,1).font = Font(bold=True)
col = 2
for yi, y in enumerate((2026, 2027, 2028)):
    for m in range(12):
        t = yi*12 + m
        c = ws.cell(gr+1, col+m); c.value = pct[t]/100.0; c.number_format = '0.00%'
        if t >= 12:
            r2 = ws.cell(gr+2, col+m); r2.value = rchk[t-12]/100.0; r2.number_format = '0.00%'
        g = ws.cell(gr+3, col+m); g.value = round(ng[t], 0); g.number_format = '#,##0'
    tc = ws.cell(gr+3, col+12); tc.value = round(sum(ng[yi*12:yi*12+12]), 0)
    tc.number_format = '#,##0'; tc.font = Font(bold=True); tc.fill = YR
    col += 13
ws.freeze_panes = 'B6'; ws.column_dimensions['A'].width = 20
for c in range(2, 42): ws.column_dimensions[get_column_letter(c)].width = 13

d = wb.create_sheet('Demand Wide')
def rv(r): return [cell(dv, r, c) for c in range(2, 14)]
G = {'RT40': (827.11/767.00, 819.28/767.00),
     'RT50': (393.83/342.94, 425.81/342.94),
     'RT70': (188.91/268.31, 179.01/268.31)}
blocks = [('Total billed kVA', [(6,'RT40','RT40'), (7,'RT50','RT50'), (8,'RT70','RT70')]),
          ('RT40 - Load Shape', [(12,'RT40','Standard'), (13,'RT40','On-Peak'), (14,'RT40','Partial-Peak'), (15,'RT40','Off-Peak')]),
          ('RT50 - Load Shape', [(19,'RT50','Standard'), (20,'RT50','On-Peak'), (21,'RT50','Partial-Peak'), (22,'RT50','Off-Peak')]),
          ('RT70 - Load Shape', [(26,'RT70','Standard'), (27,'RT70','On-Peak'), (28,'RT70','Partial-Peak'), (29,'RT70','Off-Peak')])]
d.cell(1,1).value = 'JPS Peak Demand Forecast - FY2026 to FY2028, billed kVA'
d.cell(1,1).font = Font(bold=True, size=13)
d.cell(2,1).value = ('FY2027/28 scaled from FY2026 by each class kWh growth at constant load factor. '
                     'jps_demand_actuals has no per-account grain, so demand cannot run through the Driver Forecast engine.')
d.cell(2,1).font = Font(italic=True, size=9)
r = 4
for title, rows in blocks:
    d.cell(r,1).value = title; d.cell(r,1).font = Font(bold=True, size=11); r += 1
    hy, hm = r, r+1
    d.cell(hm,1).value = 'Demand Type' if 'Shape' in title else 'Rate Class'
    d.cell(hm,1).fill = HDR; d.cell(hm,1).font = W
    col = 2
    for y in (2026, 2027, 2028):
        d.cell(hy,col).value = 'FY%d' % y
        d.cell(hy,col).font = Font(bold=True, size=11)
        d.cell(hy,col).alignment = Alignment(horizontal='center')
        d.merge_cells(start_row=hy, start_column=col, end_row=hy, end_column=col+12)
        for i, mn in enumerate(months):
            c = d.cell(hm, col+i); c.value = mn; c.fill = HDR; c.font = W
            c.alignment = Alignment(horizontal='center')
        t = d.cell(hm, col+12); t.value = 'Avg'; t.fill = HDR; t.font = W
        col += 13
    r = hm + 1
    for srow, rc, lab in rows:
        base = rv(srow); d.cell(r,1).value = lab; col = 2
        for yi, y in enumerate((2026, 2027, 2028)):
            g = 1.0 if y == 2026 else G[rc][yi-1]
            vv = [bb*g for bb in base]
            for m in range(12):
                c = d.cell(r, col+m); c.value = round(vv[m], 0); c.number_format = '#,##0'
            t = d.cell(r, col+12); t.value = round(sum(vv)/12.0, 0)
            t.number_format = '#,##0'; t.font = Font(bold=True); t.fill = YR
            col += 13
        r += 1
    r += 1
d.freeze_panes = 'B6'; d.column_dimensions['A'].width = 18
for c in range(2, 42): d.column_dimensions[get_column_letter(c)].width = 12

wb.save(dst)
print('saved', dst)
for y, off in ((2026,0), (2027,12), (2028,24)):
    S = sales[off:off+12].sum(); L = full[off:off+12].sum()
    print('FY%d: sales %.1f  losses %.1f  netgen %.1f  loss%% %.2f'
          % (y, S/1e6, L/1e6, (S+L)/1e6, L/(S+L)*100))
print('monthly loss range %.2f to %.2f pct' % (min(pct[12:]), max(pct[12:])))
print('max rolling deviation %.2f pp' % max(abs(rchk[k]-roll[k]*100) for k in range(24)))
