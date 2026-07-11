# -*- coding: utf-8 -*-
# Budget (Feb 2026 LE) monthly by class & component for 2026 -> budget_monthly.json
# Revenue components in J$000; volume in MWh.
import openpyxl, json, datetime
wb=openpyxl.load_workbook('FEB2026_LE.xlsm', read_only=True, data_only=True)
def cls_of(lbl):
    u=lbl.upper()
    for k,t in [('MT10','RT10'),('RATE 10','RT10'),('MT20','RT20'),('RATE 20','RT20'),('MT40','RT40'),('RATE 40','RT40'),
                ('MT60','RT60'),('RATE 60','RT60'),('STREET','RT60'),('TRAFFIC','RT60'),('CEMENT','RT50'),('MT50','RT50'),
                ('RATE 50','RT50'),('PORT AUTH','RT50'),('MT70','RT70'),('RATE 70','RT70')]:
        if k in u: return t
    return None
CLASSES=['RT10','RT20','RT40','RT50','RT60','RT70']
# Revenue sheet
ws=wb['Revenue']; rows=list(ws.iter_rows(values_only=True))
h=rows[4]; cmap=[(c,v.strftime('%Y-%m')) for c,v in enumerate(h) if isinstance(v,datetime.datetime) and v.year==2026]
months=[m for _,m in cmap]; cols=[c for c,_ in cmap]
sections={'FUEL':(5,56),'CUSTOMER':(58,110),'ENERGY':(112,164),'DEMAND':(188,240),'TOTAL':(242,292)}
comp={}
for name,(a,b) in sections.items():
    byc={cl:[0.0]*len(cols) for cl in CLASSES}
    for i in range(a,b+1):
        r=rows[i]; lbl=' '.join(str(r[c]).strip() for c in range(3,7) if r[c] is not None and str(r[c]).strip())
        cl=cls_of(lbl)
        if cl is None: continue
        for j,c in enumerate(cols):
            if isinstance(r[c],(int,float)): byc[cl][j]+=r[c]
    comp[name]=byc
# Demand sheet -> volume MWh by class
ws2=wb['Demand']; rr=list(ws2.iter_rows(values_only=True))
h2=rr[6]; cmap2=[(c,v.strftime('%Y-%m')) for c,v in enumerate(h2) if isinstance(v,datetime.datetime) and v.year==2026]
cols2=[c for c,_ in cmap2]
volmap={'RT10':[7,8,9],'RT20':[10,11],'RT50':[12,13,14,26,27,28,29,30,31,32,33,34,35,36],
        'RT40':[16,17,18,19,21,22,23,24],'RT60':[37,38],'RT70':[40,41,42,43]}
vol={cl:[round(sum((rr[i][c] or 0) for i in idxs)) for c in cols2] for cl,idxs in volmap.items()}
wb.close()
out={'months':months,'classes':CLASSES,'components':comp,'VOLUME_MWH':vol}
json.dump(out,open('budget_monthly.json','w'))
# sanity
ytd=lambda series: round(sum(series[:5]))
print('months:',months)
for name in ['FUEL','ENERGY','DEMAND','CUSTOMER','TOTAL']:
    print('%-9s YTD JanMay J$M:'%name, {cl:ytd([v/1000 for v in comp[name][cl]]) for cl in CLASSES if sum(comp[name][cl][:5])>1000})
print('VOLUME YTD MWh:',{cl:ytd(vol[cl]) for cl in CLASSES})
