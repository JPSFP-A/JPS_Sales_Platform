# -*- coding: utf-8 -*-
# JPS Demand/Energy/Variance workbook.
#  - Dashboard: month dropdown + Month/YTD toggle drive the by-class table & level charts.
#  - Variance shown as TREND LINES (monthly Actual-Budget) on Dashboard + Charts.
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
V=json.load(open('varset.json'))
CL=V['classes']; BMON=V['budget_months']; AMON=V['actual_months']
MLAB={f'2026-{m:02d}':lab for m,lab in zip(range(1,13),['Jan-26','Feb-26','Mar-26','Apr-26','May-26','Jun-26','Jul-26','Aug-26','Sep-26','Oct-26','Nov-26','Dec-26'])}
A=V['actual']; B=V['budget']
NAVY='1F3864'; RED='C00000'; LBLUE='9DC3E6'; LORG='F4B183'; GREEN='70AD47'; GREY='808080'
F=lambda **k: Font(name='Arial',size=10,**k)
HEADf=Font(name='Arial',size=10,bold=True,color='FFFFFF'); HEADfill=PatternFill('solid',fgColor=NAVY)
TITLEf=Font(name='Arial',size=13,bold=True,color=NAVY); CTRLfill=PatternFill('solid',fgColor='FFF2CC')
def hrow(ws,r,hdrs,c0=1):
    for j,h in enumerate(hdrs,c0):
        c=ws.cell(r,j,h); c.font=HEADf; c.fill=HEADfill; c.alignment=Alignment(horizontal='center',wrap_text=True)
def title(ws,t,sub):
    ws['A1']=t; ws['A1'].font=TITLEf; ws['A2']=sub; ws['A2'].font=Font(name='Arial',size=9,italic=True,color=GREY)
def styleline(ch,series,sheet,r0,last,catcol=1):
    ch.set_categories(Reference(sheet,min_col=catcol,min_row=r0+1,max_row=last)); ch.legend.position='b'
    ch.x_axis.delete=False; ch.y_axis.delete=False
    for k,(col,color,dsh) in enumerate(series):
        sx=ch.series[k]; sx.graphicalProperties.line.solidFill=color; sx.graphicalProperties.line.width=28000; sx.smooth=False
        if dsh: sx.graphicalProperties.line.dashStyle=dsh
W=Workbook()

# ===== RAW (hidden) =====
raw=W.active; raw.title='Raw'
raw.cell(1,1,'src'); raw.cell(1,2,'metric'); raw.cell(1,3,'class')
for j,mo in enumerate(BMON): raw.cell(1,4+j,MLAB[mo])
MET=['demand','energy','rev','vol']
def aval(c,k,mo):
    if mo not in AMON: return None
    v=A[c][k].get(mo,0); return v/1000.0 if k=='vol' else v
def bval(c,k,mo):
    v=B[c][k][mo]; return v/1000.0 if k=='vol' else v
rowmap={}; r=2
for src in ['ACT','BUD']:
    for k in MET:
        for c in CL:
            raw.cell(r,1,src); raw.cell(r,2,k); raw.cell(r,3,c)
            for j,mo in enumerate(BMON):
                val=aval(c,k,mo) if src=='ACT' else bval(c,k,mo)
                raw.cell(r,4+j, round(val,2) if val is not None else None)
            rowmap[(src,k,c)]=r; r+=1
raw.sheet_state='hidden'
MHDR='Raw!$D$1:$'+get_column_letter(3+len(BMON))+'$1'

# ===== DATA (total monthly + variance) =====
ws=W.create_sheet('Data')
title(ws,'Total Company — Demand / Energy & Variance (monthly)','J$ millions · Actual vs Budget (Feb 2026 LE) · Variance = Actual − Budget')
hrow(ws,4,['Month','Actual Demand','Actual Energy','Budget Demand','Budget Energy','Demand Var','Energy Var','Revenue Var'])
r0=4
for i,mo in enumerate(BMON):
    rr=r0+1+i; ws.cell(rr,1,MLAB[mo]).font=F()
    ad=sum(A[c]['demand'].get(mo,0) for c in CL) if mo in AMON else None
    ae=sum(A[c]['energy'].get(mo,0) for c in CL) if mo in AMON else None
    bd=sum(B[c]['demand'][mo] for c in CL); be=sum(B[c]['energy'][mo] for c in CL)
    arev=sum(A[c]['rev'].get(mo,0) for c in CL) if mo in AMON else None
    brev=sum(B[c]['rev'][mo] for c in CL)
    dvar=(ad-bd) if ad is not None else None
    evar=(ae-be) if ae is not None else None
    rvar=(arev-brev) if arev is not None else None
    for col,val,fmt in [(2,ad,'#,##0.0'),(3,ae,'#,##0.0'),(4,round(bd,1),'#,##0.0'),(5,round(be,1),'#,##0.0'),
                        (6,dvar,'#,##0.0;(#,##0.0)'),(7,evar,'#,##0.0;(#,##0.0)'),(8,rvar,'#,##0.0;(#,##0.0)')]:
        cc=ws.cell(rr,col, round(val,1) if val is not None else None); cc.font=F(); cc.number_format=fmt
last2=r0+len(BMON); ws.column_dimensions['A'].width=10
for col in 'BCDEFGH': ws.column_dimensions[col].width=13

# ===== CHARTS (trends) =====
chS=W.create_sheet('Charts')
def line_on_charts(title_,series,anchor):
    ch=LineChart(); ch.title=title_; ch.height=8.5; ch.width=20; ch.y_axis.title='J$ millions'; ch.x_axis.title='Month'
    for col,color,dsh in series: ch.add_data(Reference(ws,min_col=col,min_row=r0,max_row=last2),titles_from_data=True)
    styleline(ch,series,ws,r0,last2); chS.add_chart(ch,anchor)
line_on_charts('Demand & Energy — Actual (solid) vs Budget (dashed)',[(2,RED,None),(4,RED,'sysDash'),(3,NAVY,None),(5,NAVY,'sysDash')],'A1')
line_on_charts('Variance Trend — Actual − Budget (J$M)',[(6,RED,None),(7,NAVY,None),(8,GREEN,None)],'A19')

# ===== DASHBOARD (interactive) =====
db=W.create_sheet('Dashboard',0)
title(db,'JPS — Demand / Energy / Revenue: Actual vs Budget (Feb 2026 LE)','Interactive · pick a month & Month/YTD for the by-class table and level charts · variance shown as monthly trend lines · J$M, revenue ex-IPP')
db['A2']='Month →'; db['A2'].font=F(bold=True); db['B2']=MLAB[AMON[-1]]
db['B2'].fill=CTRLfill; db['B2'].font=Font(name='Arial',size=11,bold=True,color=NAVY); db['B2'].alignment=Alignment(horizontal='center')
db['A3']='View →'; db['A3'].font=F(bold=True); db['B3']='Month'
db['B3'].fill=CTRLfill; db['B3'].font=Font(name='Arial',size=11,bold=True,color=NAVY); db['B3'].alignment=Alignment(horizontal='center')
dvM=DataValidation(type='list',formula1='"'+','.join(MLAB[m] for m in AMON)+'"',allow_blank=False); db.add_data_validation(dvM); dvM.add(db['B2'])
dvV=DataValidation(type='list',formula1='"Month,YTD"',allow_blank=False); db.add_data_validation(dvV); dvV.add(db['B3'])
db['D2']='◄ choose month / Month vs YTD — table & the two level charts update; variance trend below shows all months'; db['D2'].font=Font(name='Arial',size=9,italic=True,color=GREY)
def dyn(src,k,c):
    row=rowmap[(src,k,c)]; rng=f'Raw!$D${row}:${get_column_letter(3+len(BMON))}${row}'
    return (f'=IF($B$3="YTD",SUM(OFFSET(Raw!$D${row},0,0,1,MATCH($B$2,{MHDR},0))),INDEX({rng},MATCH($B$2,{MHDR},0)))')
R0=5
hrow(db,R0,['Rate Class','Act Demand','Bud Demand','Act Energy','Bud Energy','Act Rev','Bud Rev','Dem Var','Engy Var','Rev Var'])
for i,c in enumerate(CL):
    rr=R0+1+i; db.cell(rr,1,c).font=F(bold=True)
    db.cell(rr,2,dyn('ACT','demand',c)); db.cell(rr,3,dyn('BUD','demand',c))
    db.cell(rr,4,dyn('ACT','energy',c)); db.cell(rr,5,dyn('BUD','energy',c))
    db.cell(rr,6,dyn('ACT','rev',c));    db.cell(rr,7,dyn('BUD','rev',c))
    db.cell(rr,8,f'=B{rr}-C{rr}'); db.cell(rr,9,f'=D{rr}-E{rr}'); db.cell(rr,10,f'=F{rr}-G{rr}')
    for col in range(2,11):
        cc=db.cell(rr,col); cc.font=F(); cc.number_format='#,##0.0;(#,##0.0)'
last=R0+len(CL); tr=last+1; db.cell(tr,1,'TOTAL').font=F(bold=True)
for col in range(2,11):
    L=get_column_letter(col); cc=db.cell(tr,col,f'=SUM({L}{R0+1}:{L}{last})'); cc.font=F(bold=True); cc.number_format='#,##0.0;(#,##0.0)'
db.column_dimensions['A'].width=12
for col in 'BCDEFGHIJ': db.column_dimensions[col].width=11
# level bars (selected month) by class
catsC=Reference(db,min_col=1,min_row=R0+1,max_row=last)
def bar(title_,cols,anchor):
    ch=BarChart(); ch.type='col'; ch.title=title_; ch.height=8; ch.width=11.5; ch.y_axis.title='J$M'
    ch.x_axis.delete=False; ch.y_axis.delete=False
    for col,color in cols: ch.add_data(Reference(db,min_col=col,min_row=R0,max_row=last),titles_from_data=True)
    ch.set_categories(catsC); ch.legend.position='b'
    for k,(col,color) in enumerate(cols): ch.series[k].graphicalProperties.solidFill=color
    db.add_chart(ch,anchor)
bar('Demand — Actual vs Budget (selected month, by class)',[(2,RED),(3,LORG)],'A16')
bar('Energy — Actual vs Budget (selected month, by class)',[(4,NAVY),(5,LBLUE)],'L16')
# variance TREND lines (all months) — references Data sheet
vt=LineChart(); vt.title='Variance Trend — Actual − Budget (monthly, total company, J$M)'; vt.height=9; vt.width=23
vt.y_axis.title='J$M (Actual − Budget)'; vt.x_axis.title='Month'
vser=[(6,RED,None),(7,NAVY,None),(8,GREEN,None)]
for col,color,dsh in vser: vt.add_data(Reference(ws,min_col=col,min_row=r0,max_row=last2),titles_from_data=True)
styleline(vt,vser,ws,r0,last2); db.add_chart(vt,'A33')

W.save('JPS_Demand_vs_Energy_Actual_Budget.xlsx')
print('wrote workbook | variance = trend lines | months:',[MLAB[m] for m in AMON])
