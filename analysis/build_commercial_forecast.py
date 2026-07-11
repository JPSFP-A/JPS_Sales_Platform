import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"D:\Projects\Sales_Platform\analysis\commercial_actuals_2024_2026.csv"
OUT = r"D:\Projects\Sales_Platform\analysis\Commercial_RateClass_Forecast_2026.xlsx"

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF")
YELLOW = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="B7C6D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUMFMT = "#,##0;(#,##0);\"-\""
PCTFMT = "0.0%"

rows = list(csv.DictReader(open(SRC, encoding="utf-8")))

# key = (jps_ac, rate_class) -> {name, parish, {(year,month): kwh}}
cust = {}
for r in rows:
    k = (r["jps_ac"], r["rate_class"])
    if k not in cust:
        cust[k] = {"name": r["name"], "parish": r["parish"], "kwh": {}}
    y, m = int(r["year"]), int(r["month"])
    cust[k]["kwh"][(y, m)] = float(r["kwh"] or 0)
    if r["name"]:
        cust[k]["name"] = r["name"]
    if r["parish"]:
        cust[k]["parish"] = r["parish"]

# scope = accounts present in 2026 Jan-Jun
keys = sorted(k for k in cust if any((2026, m) in cust[k]["kwh"] for m in range(1, 7)))
print("accounts in scope:", len(keys))

RC_INFO = [
    ("RT40", "LV Commercial"),
    ("RT50", "MV Commercial"),
    ("RT60-ST", "Street Lights"),
    ("RT70", "Large Industrial"),
]
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

wb = Workbook()

# ── README ──────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Read Me"
ws.column_dimensions["A"].width = 100
lines = [
    ("Commercial Rate Class Forecast — 2026", BOLD),
    ("", None),
    ("Source: jps_actuals table, Supabase project bhrswnbenkvflpdjhfpa. Pulled 2026-07-06.", BLACK),
    ("Scope: Commercial rate classes only — RT40 (LV Commercial), RT50 (MV Commercial), RT60-ST (Street Lights), RT70 (Large Industrial).", BLACK),
    ("Jan-Jun 2026: real actuals. Jul-Dec 2026: formula-driven projection (see Methodology).", BLACK),
    ("", None),
    ("DATA QUALITY FLAG — NOT FIXED, EXCLUDED FROM THIS WORKBOOK:", Font(name=FONT, bold=True, color="C00000")),
    ("jps_actuals for 2026-06 contains 23,780 rows tagged rate_class='RT20', segment='Commercial' that do not belong there.", Font(name=FONT, color="C00000")),
    ("Inspection shows this is really residential Prepaid/Zero-usage bucket data (jps_ac='Prepaid'/'Zero', real parish-level customer counts,", Font(name=FONT, color="C00000")),
    ("null revenue, blank consumption_bucket) mistagged as individual commercial RT20 accounts. Inserted 2026-07-03 14:45-14:48 UTC with no", Font(name=FONT, color="C00000")),
    ("matching entry in sync_runs, so it did not go through the normal sync tool path. RT20 is excluded entirely from this workbook.", Font(name=FONT, color="C00000")),
    ("This needs cleanup in jps_actuals directly before RT20 or residential 2026-06 figures can be trusted anywhere else in the platform.", Font(name=FONT, color="C00000")),
    ("", None),
    ("Methodology (Jul-Dec 2026 projection, per account):", BOLD),
    ("  seasonal base = average of that account's real kWh in the same calendar month, 2024 and 2025 (2025 only if 2024 has no data)", BLACK),
    ("  projected kWh = seasonal base x (1 + trailing growth% + Growth Adjustment toggle) + (Other Monthly Adjustment toggle x account's share of Jun-2026 kWh within its rate class)", BLACK),
    ("  trailing growth% = (rate class Apr-Jun 2026 actual) / (rate class Apr-Jun 2025 actual) - 1, computed live from the Commercial Customers sheet", BLACK),
    ("  Customer Count Growth toggle is applied at the rate-class level only (Rate Class Summary sheet), as an add-on for new accounts, not spread across existing accounts", BLACK),
    ("", None),
    ("Tabs:", BOLD),
    ("  Drivers — toggle inputs (blue cells). Change these to update the whole forecast.", BLACK),
    ("  Rate Class Summary — Jan-Jun actual + Jul-Dec projected by rate class, plus Customer Count Growth add-on.", BLACK),
    ("  Commercial Customers — per-account detail: 2024/2025 history, 2026 actuals, 2026 projection.", BLACK),
]
r = 1
for text, font in lines:
    c = ws.cell(row=r, column=1, value=text)
    if font:
        c.font = font
    r += 1

# ── DRIVERS ─────────────────────────────────────────────────────────
dv = wb.create_sheet("Drivers")
headers = ["Rate Class", "Description", "Trailing Growth % (Apr-Jun YoY, calc)", "Growth Adjustment % (toggle)",
           "Customer Count Growth % (toggle)", "Other Monthly Adjustment (kWh, toggle)"]
for i, h in enumerate(headers, 1):
    c = dv.cell(row=1, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = BORDER
dv.row_dimensions[1].height = 40
widths = [12, 20, 26, 22, 24, 28]
for i, w in enumerate(widths, 1):
    dv.column_dimensions[get_column_letter(i)].width = w

CC = "'Commercial Customers'"
for i, (rc, desc) in enumerate(RC_INFO):
    row = i + 2
    dv.cell(row=row, column=1, value=rc).font = BLACK
    dv.cell(row=row, column=2, value=desc).font = BLACK
    # trailing growth = sum(2026 Apr+May+Jun) / sum(2025 Apr+May+Jun) - 1
    f = (f"=IFERROR((SUMIF({CC}!$C:$C,$A{row},{CC}!$H:$H)+SUMIF({CC}!$C:$C,$A{row},{CC}!$I:$I)+SUMIF({CC}!$C:$C,$A{row},{CC}!$J:$J))"
         f"/(SUMIF({CC}!$C:$C,$A{row},{CC}!$K:$K)+SUMIF({CC}!$C:$C,$A{row},{CC}!$L:$L)+SUMIF({CC}!$C:$C,$A{row},{CC}!$M:$M))-1,0)")
    cell = dv.cell(row=row, column=3, value=f)
    cell.font = BLACK
    cell.number_format = PCTFMT
    d = dv.cell(row=row, column=4, value=0.0)
    d.font = BLUE
    d.number_format = PCTFMT
    d.fill = YELLOW
    e = dv.cell(row=row, column=5, value=0.0)
    e.font = BLUE
    e.number_format = PCTFMT
    e.fill = YELLOW
    g = dv.cell(row=row, column=6, value=0)
    g.font = BLUE
    g.number_format = NUMFMT
    g.fill = YELLOW
    for col in range(1, 7):
        dv.cell(row=row, column=col).border = BORDER

DRV = "Drivers"

# ── COMMERCIAL CUSTOMERS ────────────────────────────────────────────
cs = wb.create_sheet("Commercial Customers")
cols = (["Account #", "Name", "Rate Class", "Parish"]
        + [f"2026 {m}" for m in MONTHS[:6]]
        + [f"2025 {m}" for m in MONTHS[3:6]]
        + [f"2025 {m}" for m in MONTHS[6:12]]
        + [f"2024 {m}" for m in MONTHS[6:12]]
        + [f"2026 {m} (proj)" for m in MONTHS[6:12]]
        + ["FY2026 Total (calc)"])
for i, h in enumerate(cols, 1):
    c = cs.cell(row=1, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = BORDER
cs.row_dimensions[1].height = 30
cs.column_dimensions["A"].width = 14
cs.column_dimensions["B"].width = 26
cs.column_dimensions["C"].width = 10
cs.column_dimensions["D"].width = 14
for i in range(5, len(cols) + 1):
    cs.column_dimensions[get_column_letter(i)].width = 12
cs.freeze_panes = "E2"

# column index map
# A=1 acct,B=2 name,C=3 rc,D=4 parish
# E-J (5-10) = 2026 Jan-Jun
# K-M (11-13) = 2025 Apr-Jun
# N-S (14-19) = 2025 Jul-Dec
# T-Y (20-25) = 2024 Jul-Dec
# Z-AE (26-31) = 2026 Jul-Dec projected
# AF (32) = FY total
COL_2026_JANJUN = list(range(5, 11))
COL_2025_AMJ = list(range(11, 14))
COL_2025_JD = list(range(14, 20))
COL_2024_JD = list(range(20, 26))
COL_2026_PROJ = list(range(26, 32))
COL_TOTAL = 32

for ridx, k in enumerate(keys):
    row = ridx + 2
    d = cust[k]
    cs.cell(row=row, column=1, value=k[0]).font = BLACK
    cs.cell(row=row, column=2, value=d["name"]).font = BLACK
    cs.cell(row=row, column=3, value=k[1]).font = BLACK
    cs.cell(row=row, column=4, value=d["parish"]).font = BLACK

    for j, m in enumerate(range(1, 7)):
        v = d["kwh"].get((2026, m), 0)
        cell = cs.cell(row=row, column=COL_2026_JANJUN[j], value=round(v))
        cell.font = BLACK
        cell.number_format = NUMFMT
        cell.border = BORDER

    for j, m in enumerate(range(4, 7)):
        v = d["kwh"].get((2025, m), 0)
        cell = cs.cell(row=row, column=COL_2025_AMJ[j], value=round(v))
        cell.font = BLACK
        cell.number_format = NUMFMT

    for j, m in enumerate(range(7, 13)):
        v = d["kwh"].get((2025, m), 0)
        cell = cs.cell(row=row, column=COL_2025_JD[j], value=round(v))
        cell.font = BLACK
        cell.number_format = NUMFMT

    for j, m in enumerate(range(7, 13)):
        v = d["kwh"].get((2024, m), 0)
        cell = cs.cell(row=row, column=COL_2024_JD[j], value=round(v))
        cell.font = BLACK
        cell.number_format = NUMFMT

    jun2026_col = get_column_letter(COL_2026_JANJUN[5])  # J
    rc_col = get_column_letter(3)
    for j in range(6):
        hist25_col = get_column_letter(COL_2025_JD[j])
        hist24_col = get_column_letter(COL_2024_JD[j])
        proj_col_idx = COL_2026_PROJ[j]
        f = (f"=IF({hist24_col}{row}=0,{hist25_col}{row},AVERAGE({hist25_col}{row},{hist24_col}{row}))"
             f"*(1+VLOOKUP($C{row},{DRV}!$A:$F,3,0)+VLOOKUP($C{row},{DRV}!$A:$F,4,0))"
             f"+VLOOKUP($C{row},{DRV}!$A:$F,6,0)*IFERROR({jun2026_col}{row}/SUMIF({rc_col}:{rc_col},$C{row},{jun2026_col}:{jun2026_col}),0)")
        cell = cs.cell(row=row, column=proj_col_idx, value=f)
        cell.font = BLACK
        cell.number_format = NUMFMT
        cell.border = BORDER

    first_actual = get_column_letter(COL_2026_JANJUN[0])
    last_actual = get_column_letter(COL_2026_JANJUN[-1])
    first_proj = get_column_letter(COL_2026_PROJ[0])
    last_proj = get_column_letter(COL_2026_PROJ[-1])
    tot = cs.cell(row=row, column=COL_TOTAL,
                  value=f"=SUM({first_actual}{row}:{last_actual}{row})+SUM({first_proj}{row}:{last_proj}{row})")
    tot.font = BOLD
    tot.number_format = NUMFMT
    tot.border = BORDER

LAST_ROW = len(keys) + 1

# ── RATE CLASS SUMMARY ──────────────────────────────────────────────
rcs = wb.create_sheet("Rate Class Summary", 1)
rcs.column_dimensions["A"].width = 24
for i in range(2, 15):
    rcs.column_dimensions[get_column_letter(i)].width = 12
hdr = ["Rate Class"] + [f"2026 {m}" for m in MONTHS]
for i, h in enumerate(hdr, 1):
    c = rcs.cell(row=1, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.border = BORDER

CCref = "'Commercial Customers'"
for i, (rc, desc) in enumerate(RC_INFO):
    row = i + 2
    rcs.cell(row=row, column=1, value=f"{rc} — {desc}").font = BOLD
    for j in range(6):
        col_letter = get_column_letter(COL_2026_JANJUN[j])
        f = f"=SUMIF({CCref}!$C:$C,\"{rc}\",{CCref}!${col_letter}:${col_letter})"
        cell = rcs.cell(row=row, column=j + 2, value=f)
        cell.font = GREEN
        cell.number_format = NUMFMT
    for j in range(6):
        col_letter = get_column_letter(COL_2026_PROJ[j])
        f = f"=SUMIF({CCref}!$C:$C,\"{rc}\",{CCref}!${col_letter}:${col_letter})"
        cell = rcs.cell(row=row, column=8 + j, value=f)
        cell.font = GREEN
        cell.number_format = NUMFMT
    for col in range(1, 14):
        rcs.cell(row=row, column=col).border = BORDER

cust_growth_row = len(RC_INFO) + 3
rcs.cell(row=cust_growth_row, column=1, value="Customer Count Growth Add-On (Jul-Dec, kWh)").font = Font(name=FONT, italic=True)
for i, (rc, desc) in enumerate(RC_INFO):
    pass
for j in range(6):
    col_idx = 8 + j
    parts = []
    for rc, desc in RC_INFO:
        parts.append(
            f"COUNTIF({CCref}!$C:$C,\"{rc}\")*VLOOKUP(\"{rc}\",{DRV}!$A:$F,5,0)*IFERROR(SUMIF({CCref}!$C:$C,\"{rc}\",{CCref}!$J:$J)/COUNTIF({CCref}!$C:$C,\"{rc}\"),0)"
        )
    f = "=" + "+".join(parts)
    cell = rcs.cell(row=cust_growth_row, column=col_idx, value=f)
    cell.font = BLACK
    cell.number_format = NUMFMT

total_row = cust_growth_row + 1
rcs.cell(row=total_row, column=1, value="TOTAL (incl. Count Growth)").font = BOLD
for j in range(6):
    col = get_column_letter(2 + j)
    cell = rcs.cell(row=total_row, column=2 + j, value=f"=SUM({col}2:{col}{cust_growth_row-2})")
    cell.font = BOLD
    cell.number_format = NUMFMT
for j in range(6):
    col = get_column_letter(8 + j)
    cell = rcs.cell(row=total_row, column=8 + j,
                     value=f"=SUM({col}2:{col}{cust_growth_row-1})")
    cell.font = BOLD
    cell.number_format = NUMFMT
for col in range(1, 14):
    rcs.cell(row=total_row, column=col).border = BORDER

wb.move_sheet("Drivers", offset=-100)
wb.active = 0
wb.save(OUT)
print("saved", OUT, "rows:", LAST_ROW - 1)
