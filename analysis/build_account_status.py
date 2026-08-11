# -*- coding: utf-8 -*-
# Backfills a lightweight jps_account_status table (jps_ac, year, month, status,
# name, title) from the raw Billing Details Report's Account_Status column --
# jps_actuals has no status field at all, so "was this account closed or did it
# just not process" can't be answered from the DB alone (see analysis that led
# to this: Caribbean Broilers/NWC dropped out of jps_actuals for Jul 2026 while
# still Account_Status='A' in the raw file -- a pipeline gap, not a closure,
# vs. e.g. Ibex Global which really is Account_Status='I').
#
# Scope: only the premise-level classes the Sales Explorer's "Not billed" tab
# tracks (RT20/RT40/RT50/RT60/RT70) -- RT10 mass-market rows are skipped (huge
# volume, no per-account identity used anywhere in the app). Reuses the exact
# same title_of(rate_class, srat) resolution as corrected_scan.py so a row
# classifies identically everywhere in this pipeline.
#
# Coverage is whatever raw Billing Details Report files exist locally -- as of
# this run that's 11 of the 18 months the Explorer has actuals for (May 2025,
# Oct/Nov/Dec 2025, Jan-Jul 2026). Months with no raw file simply get no rows
# here; the Explorer must show "status unknown" rather than pretend a gap.
import glob, os, re, csv, json, sys, time, requests

DL = r'C:\Users\jwilson\Downloads'
MONNUM = {'JAN':1,'FEB':2,'MAR':3,'APR':4,'MAY':5,'JUN':6,'JUL':7,'AUG':8,'SEP':9,'OCT':10,'NOV':11,'DEC':12}

def discover_billing_files():
    files = {}
    cands = sorted(set(glob.glob(os.path.join(DL, 'Billing Details Report*.xls*')) + glob.glob('Billing Details Report*.xls*')))
    for fp in cands:
        base = os.path.basename(fp)
        m = re.search(r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*[ _-]*(\d{4})', base.upper())
        if not m:
            continue
        key = '%s-%02d' % (m.group(2), MONNUM[m.group(1)])
        files.setdefault(key, fp if os.path.exists(fp) else os.path.join(DL, base))
    return dict(sorted(files.items()))

def norm(s): return str(s).strip().upper()

RMAP = {}; RCMAP = {}
with open(os.path.join(DL, 'Rate categorry Data mapping.csv'), encoding='utf-8-sig') as f:
    for r in list(csv.reader(f))[1:]:
        if len(r) >= 3:
            RMAP[norm(r[2])] = r[0]
            if norm(r[1]) not in RCMAP:
                RCMAP[norm(r[1])] = r[0]

def title_of(rc, srat):
    return RCMAP.get(norm(rc)) or RMAP.get(norm(srat))

KEEP = {'RT20', 'RT40', 'RT50', 'RT60', 'RT60-ST', 'RT70'}

def is_zip_xlsx(path):
    with open(path, 'rb') as f:
        return f.read(2) == b'PK'

def rows_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    detail = None
    for sn in wb.sheetnames:
        ws = wb[sn]
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if r and r[0] == 'Cust_Code':
                detail = sn; break
            if i > 8: break
        if detail: break
    ws = wb[detail]
    for r in ws.iter_rows(values_only=True):
        yield r
    wb.close()

def rows_tsv(path):
    with open(path, encoding='latin-1', errors='replace') as f:
        for line in f:
            yield tuple(line.rstrip('\r\n').split('\t'))

def proc(path, y, mo):
    rows = rows_xlsx(path) if is_zip_xlsx(path) else rows_tsv(path)
    hdr = None
    for r in rows:
        if r and r[0] == 'Cust_Code':
            hdr = list(r); break
    I = {n: k for k, n in enumerate(hdr) if n}
    gv = lambda r, n: (r[I[n]] if n in I and I[n] < len(r) else None)
    out = {}
    n_seen = 0
    for r in rows:
        code = str(gv(r, 'Cust_Code') or '').strip().replace(',', '')
        if code in ('', 'Cust_Code'):
            continue
        srat = str(gv(r, 'Srat_Code') or ''); rc = str(gv(r, 'rate_class') or '')
        title = title_of(rc, srat)
        if title not in KEEP:
            continue
        if title == 'RT60-ST':
            title = 'RT60'
        prem = str(gv(r, 'Prem_Code') or '').strip()
        full_ac = code + '-' + prem if prem else code
        status = str(gv(r, 'Account_Status') or '').strip()
        name = str(gv(r, 'Name') or '').strip()
        out[full_ac] = {'jps_ac': full_ac, 'year': y, 'month': mo, 'status': status, 'name': name, 'title': title}
        n_seen += 1
    return list(out.values()), n_seen

SECRET = open(r'D:\Projects\DataManager\.env').read().split('=', 1)[1].strip()
URL = 'https://bhrswnbenkvflpdjhfpa.supabase.co/rest/v1/jps_account_status'
HDRS = {'apikey': SECRET, 'Authorization': 'Bearer ' + SECRET, 'Content-Type': 'application/json',
        'Prefer': 'resolution=merge-duplicates,return=minimal'}
DRY_RUN = '--dry-run' in sys.argv

def push(rows_out):
    if DRY_RUN:
        return len(rows_out)
    BATCH = 500
    pushed = 0
    for i in range(0, len(rows_out), BATCH):
        chunk = rows_out[i:i + BATCH]
        ok = False; last = None
        for attempt in range(5):
            try:
                resp = requests.post(URL + '?on_conflict=jps_ac,year,month', headers=HDRS, data=json.dumps(chunk), timeout=60)
                if resp.status_code < 300:
                    ok = True; break
                last = f'{resp.status_code} {resp.text[:300]}'
            except Exception as e:
                last = repr(e)
            time.sleep(2)
        if not ok:
            print('BATCH FAILED', last, flush=True)
            continue
        pushed += len(chunk)
    return pushed

DONE_LOG = '_account_status_done.json'
done = set(json.load(open(DONE_LOG))) if os.path.exists(DONE_LOG) else set()

if __name__ == '__main__':
    files = discover_billing_files()
    print('discovered', len(files), 'raw billing files:', list(files.keys()), flush=True)
    print('already done (skipping):', sorted(done), flush=True)
    grand = 0
    for key, path in files.items():
        if key in done:
            continue
        y, mo = int(key[:4]), int(key[5:7])
        t = time.time()
        try:
            rows_out, n_seen = proc(path, y, mo)
            n = push(rows_out)
        except Exception as e:
            import traceback
            print(f'{key}: FAILED -- {e!r}', flush=True)
            traceback.print_exc()
            continue
        grand += n
        by_status = {}
        for r in rows_out:
            by_status[r['status']] = by_status.get(r['status'], 0) + 1
        print(f'{key}: raw_rows_scanned_for_target_classes={n_seen} unique_accounts={len(rows_out)} pushed={n} '
              f'by_status={by_status} in {round(time.time()-t)}s', flush=True)
        if not DRY_RUN:
            done.add(key)
            json.dump(sorted(done), open(DONE_LOG, 'w'))
    print('GRAND TOTAL PUSHED:', grand, flush=True)
